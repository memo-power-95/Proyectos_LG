"""
Sistema de Retrabajo y Clasificación de PCBs
=============================================
Reusa la logica de mapeo tipo "ajedrez" del remapeo_4.0.1.py, pero aplicada
al flujo de retrabajo: escaneo con validacion de duplicados, cruce contra el
Excel que regresa el ERP (modelo/proceso por Lot ID), vista de clasificacion
por color (modelo o proceso) e historial persistente.

Autor: prototipo generado con Claude para Guillermo.
"""

import os
import sys
import sqlite3
import string
import tkinter as tk
import datetime as _dt
from datetime import datetime
from tkinter import ttk, messagebox, filedialog

import pandas as pd

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_DISPONIBLE = True
except ImportError:
    MATPLOTLIB_DISPONIBLE = False

def _carpeta_base_persistente():
    """
    Devuelve una carpeta ESTABLE donde guardar la base de datos, sin importar
    si el programa corre como script .py o como .exe empaquetado (PyInstaller).

    IMPORTANTE: cuando PyInstaller empaqueta en modo "onefile", el programa se
    descomprime en una carpeta temporal (sys._MEIPASS) que Windows borra al
    cerrar la app. Si la BD se guardaba ahi (usando __file__ directamente),
    TODO se perdia al cerrar el programa - eso es lo que estaba pasando.

    Aqui, si el programa esta "congelado" (frozen, o sea corriendo como .exe),
    se usa la carpeta donde vive el .exe (que si es permanente). Si corre como
    script normal, se usa la carpeta del .py como siempre.
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


DB_PATH = os.path.join(_carpeta_base_persistente(), "retrabajo_pcb.db")

# Paleta fija para que los colores sean consistentes entre sesiones.
COLOR_PALETTE = [
    "#4CAF50", "#2196F3", "#FF9800", "#9C27B0", "#F44336",
    "#00BCD4", "#8BC34A", "#E91E63", "#3F51B5", "#FFC107",
    "#795548", "#607D8B",
]
COLOR_PENDIENTE = "#D9D9D9"   # gris: sin escanear
COLOR_ESCANEADO = "#A5D6A7"   # verde suave: escaneado, aun sin cruzar con Excel
COLOR_SIN_MATCH = "#FFF176"   # amarillo: escaneado pero el Excel no trajo dato


# --------------------------------------------------------------------------
# Capa de datos
# --------------------------------------------------------------------------
class Database:
    def __init__(self, path=DB_PATH):
        self.conn = sqlite3.connect(path)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._crear_tablas()

    def _crear_tablas(self):
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS charolas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE,
                filas INTEGER,
                columnas INTEGER,
                volteada INTEGER DEFAULT 0,
                auto_volteo_180 INTEGER DEFAULT 0,
                volteo_180_aplicado INTEGER DEFAULT 0,
                creada TEXT
            )
        """)
        # Migracion suave por si la DB ya existia sin estas columnas
        cur.execute("PRAGMA table_info(charolas)")
        columnas_existentes = [c[1] for c in cur.fetchall()]
        if "volteada" not in columnas_existentes:
            cur.execute("ALTER TABLE charolas ADD COLUMN volteada INTEGER DEFAULT 0")
        if "auto_volteo_180" not in columnas_existentes:
            cur.execute("ALTER TABLE charolas ADD COLUMN auto_volteo_180 INTEGER DEFAULT 0")
        if "volteo_180_aplicado" not in columnas_existentes:
            cur.execute("ALTER TABLE charolas ADD COLUMN volteo_180_aplicado INTEGER DEFAULT 0")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS pcbs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                charola_id INTEGER,
                posicion TEXT,
                lot_id TEXT,
                modelo TEXT,
                proceso TEXT,
                separada INTEGER DEFAULT 0,
                escaneado_en TEXT,
                actualizado_en TEXT,
                FOREIGN KEY(charola_id) REFERENCES charolas(id),
                UNIQUE(charola_id, posicion)
            )
        """)
        cur.execute("PRAGMA table_info(pcbs)")
        columnas_pcbs = [c[1] for c in cur.fetchall()]
        if "separada" not in columnas_pcbs:
            cur.execute("ALTER TABLE pcbs ADD COLUMN separada INTEGER DEFAULT 0")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS historial (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                evento TEXT,
                detalle TEXT,
                fecha TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS equivalencias_modelo (
                modelo_variante TEXT PRIMARY KEY,
                modelo_base TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS posiciones_bloqueadas (
                charola_id INTEGER,
                posicion TEXT,
                PRIMARY KEY (charola_id, posicion),
                FOREIGN KEY(charola_id) REFERENCES charolas(id)
            )
        """)
        # --- tablas del modulo de Remapeo (fusionado desde remapeo_4.0.1) ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS remapeos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                turno INTEGER,
                hora TEXT,
                carrier_base TEXT,
                carrier_id TEXT,
                lot_id TEXT,
                creado_en TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS remapeo_liberados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                carrier_id TEXT,
                lot_id TEXT,
                fecha_liberado TEXT,
                UNIQUE(carrier_id, lot_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS catalogo_modelos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                familia TEXT,
                nombre TEXT,
                lgit_model TEXT,
                pcb_type TEXT,
                degrees TEXT,
                top_housing TEXT,
                housing_mess TEXT,
                bottom_cover TEXT,
                opal TEXT
            )
        """)
        # --- tablas del modulo de Inventario por Racks (gestion de almacen) ---
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventario_racks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE,
                num_trays INTEGER,
                creado TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventario_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rack_id INTEGER,
                tray_numero INTEGER,
                lot_id TEXT,
                escaneado_en TEXT,
                FOREIGN KEY(rack_id) REFERENCES inventario_racks(id),
                UNIQUE(rack_id, lot_id)
            )
        """)
        self.conn.commit()

    # --- charolas ---
    def siguiente_nombre_automatico(self, capacidad=999):
        """
        Genera el siguiente nombre automatico de charola: A1, A2, A3... hasta
        'capacidad', luego B1, B2, B3... y asi sucesivamente (reutiliza el mismo
        esquema de letras tipo Excel que usan las columnas del grid, asi que
        nunca se acaban las combinaciones: ...Z1..Z999, AA1..AA999, AB1...).

        Se basa en el contador AUTOINCREMENT de la tabla (siempre creciente y
        unico), no en parsear nombres existentes, asi que funciona sin problema
        aunque el usuario haya nombrado charolas anteriores manualmente.
        """
        cur = self.conn.cursor()
        cur.execute("SELECT COALESCE(MAX(id), 0) FROM charolas")
        ultimo_id = cur.fetchone()[0]
        idx = ultimo_id  # 0-based: la proxima charola tendra id = ultimo_id + 1
        letra = etiqueta_columna(idx // capacidad)
        numero = (idx % capacidad) + 1
        return f"{letra}{numero}"

    def crear_charola(self, nombre, filas, columnas, volteada=0, auto_volteo_180=0):
        cur = self.conn.cursor()
        cur.execute(
            """INSERT INTO charolas (nombre, filas, columnas, volteada, auto_volteo_180, volteo_180_aplicado, creada)
               VALUES (?,?,?,?,?,0,?)""",
            (nombre, filas, columnas, int(volteada), int(auto_volteo_180), datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()
        etiqueta_volteo = " [volteada]" if volteada else ""
        etiqueta_auto = " [auto-volteo 180° al clasificar]" if auto_volteo_180 else ""
        self.registrar_historial("charola_creada", f"{nombre} ({filas}x{columnas}){etiqueta_volteo}{etiqueta_auto}")
        return cur.lastrowid

    def listar_charolas(self):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id, nombre, filas, columnas, volteada, auto_volteo_180, volteo_180_aplicado "
            "FROM charolas ORDER BY id DESC"
        )
        return cur.fetchall()

    def obtener_charola(self, charola_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id, nombre, filas, columnas, volteada, auto_volteo_180, volteo_180_aplicado "
            "FROM charolas WHERE id=?",
            (charola_id,),
        )
        return cur.fetchone()

    # --- pcbs / escaneo ---
    def lot_id_ya_escaneado(self, lot_id):
        """Valida duplicados en TODO el sistema, no solo en la charola actual."""
        cur = self.conn.cursor()
        cur.execute("SELECT charola_id, posicion FROM pcbs WHERE lot_id=?", (lot_id,))
        return cur.fetchone()

    def posicion_ocupada(self, charola_id, posicion):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT lot_id FROM pcbs WHERE charola_id=? AND posicion=?",
            (charola_id, posicion),
        )
        return cur.fetchone()

    def escanear(self, charola_id, posicion, lot_id):
        cur = self.conn.cursor()
        ahora = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            """INSERT INTO pcbs (charola_id, posicion, lot_id, escaneado_en, actualizado_en)
               VALUES (?,?,?,?,?)""",
            (charola_id, posicion, lot_id, ahora, ahora),
        )
        self.conn.commit()
        self.registrar_historial("pcb_escaneado", f"{lot_id} -> charola {charola_id} pos {posicion}")

    def undo_ultimo_escaneo(self, charola_id, posicion):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM pcbs WHERE charola_id=? AND posicion=?", (charola_id, posicion))
        self.conn.commit()
        self.registrar_historial("undo_escaneo", f"charola {charola_id} pos {posicion}")

    def pcbs_de_charola(self, charola_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT posicion, lot_id, modelo, proceso, separada FROM pcbs WHERE charola_id=?",
            (charola_id,),
        )
        return {
            row[0]: {"lot_id": row[1], "modelo": row[2], "proceso": row[3], "separada": bool(row[4])}
            for row in cur.fetchall()
        }

    def marcar_separada(self, charola_id, posicion, valor=True):
        cur = self.conn.cursor()
        cur.execute(
            "UPDATE pcbs SET separada=? WHERE charola_id=? AND posicion=?",
            (int(bool(valor)), charola_id, posicion),
        )
        self.conn.commit()

    def pcbs_agrupados_por_modelo(self, charola_id, usar_modelo_base=False):
        """Para la lista de separacion: agrupa las piezas ya clasificadas por modelo.

        Si usar_modelo_base=True, agrupa usando la tabla de equivalencias
        (modelo_variante -> modelo_base). Si un modelo no tiene equivalencia
        registrada, se agrupa tal cual viene del ERP (no se pierde ni se inventa nada).
        """
        cur = self.conn.cursor()
        cur.execute(
            """SELECT posicion, lot_id, modelo, proceso, separada FROM pcbs
               WHERE charola_id=? AND modelo IS NOT NULL
               ORDER BY modelo, posicion""",
            (charola_id,),
        )
        filas = cur.fetchall()

        equivalencias = {}
        if usar_modelo_base:
            equivalencias = dict(self.listar_equivalencias())

        grupos = {}
        for pos, lot_id, modelo, proceso, separada in filas:
            clave = equivalencias.get(modelo, modelo) if usar_modelo_base else modelo
            grupos.setdefault(clave, []).append(
                {
                    "posicion": pos, "lot_id": lot_id, "modelo": modelo,
                    "proceso": proceso, "separada": bool(separada),
                }
            )
        return grupos

    # --- equivalencias de modelo base ---
    def agregar_equivalencia(self, variante, base):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT OR REPLACE INTO equivalencias_modelo (modelo_variante, modelo_base) VALUES (?,?)",
            (variante.strip(), base.strip()),
        )
        self.conn.commit()
        self.registrar_historial("equivalencia_modelo", f"{variante} -> {base}")

    def eliminar_equivalencia(self, variante):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM equivalencias_modelo WHERE modelo_variante=?", (variante,))
        self.conn.commit()
        self.registrar_historial("equivalencia_eliminada", variante)

    def listar_equivalencias(self):
        """Devuelve todas las equivalencias modelo_variante -> modelo_base."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT modelo_variante, modelo_base FROM equivalencias_modelo ORDER BY modelo_variante"
        )
        return cur.fetchall()

    # =========================================================================
    # CATALOGO DE MODELOS LGIT (tablas de referencia tipo Thunder Trinity/Cheetah)
    # =========================================================================

    def agregar_catalogo_modelo(self, familia, nombre, lgit_model, pcb_type, degrees,
                                  top_housing, housing_mess, bottom_cover, opal):
        cur = self.conn.cursor()
        cur.execute(
            """INSERT INTO catalogo_modelos
               (familia, nombre, lgit_model, pcb_type, degrees, top_housing, housing_mess, bottom_cover, opal)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (familia, nombre, lgit_model, pcb_type, degrees, top_housing, housing_mess, bottom_cover, opal),
        )
        self.conn.commit()
        return cur.lastrowid

    def eliminar_catalogo_modelo(self, catalogo_id):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM catalogo_modelos WHERE id=?", (catalogo_id,))
        self.conn.commit()

    def listar_catalogo_modelos(self, familia=None):
        cur = self.conn.cursor()
        if familia and familia != "Todas":
            cur.execute(
                """SELECT id, familia, nombre, lgit_model, pcb_type, degrees, top_housing,
                          housing_mess, bottom_cover, opal
                   FROM catalogo_modelos WHERE familia=? ORDER BY nombre""",
                (familia,),
            )
        else:
            cur.execute(
                """SELECT id, familia, nombre, lgit_model, pcb_type, degrees, top_housing,
                          housing_mess, bottom_cover, opal
                   FROM catalogo_modelos ORDER BY familia, nombre"""
            )
        return cur.fetchall()

    def buscar_catalogo_modelo(self, texto):
        """Busca por LGIT model o nombre (coincidencia parcial, insensible a mayusculas)."""
        cur = self.conn.cursor()
        patron = f"%{texto.strip().upper()}%"
        cur.execute(
            """SELECT id, familia, nombre, lgit_model, pcb_type, degrees, top_housing,
                      housing_mess, bottom_cover, opal
               FROM catalogo_modelos
               WHERE UPPER(lgit_model) LIKE ? OR UPPER(nombre) LIKE ?
               ORDER BY familia, nombre""",
            (patron, patron),
        )
        return cur.fetchall()

    def listar_familias_catalogo(self):
        cur = self.conn.cursor()
        cur.execute("SELECT DISTINCT familia FROM catalogo_modelos ORDER BY familia")
        return [r[0] for r in cur.fetchall()]

    def importar_catalogo_desde_texto(self, texto, familia_default=""):
        """Pega masiva: cada linea = nombre, lgit_model, pcb_type, degrees, top_housing,
        [housing_mess], bottom_cover, [opal] separados por tab. Columnas opcionales
        se rellenan vacias si no vienen. Retorna cuantas filas se insertaron."""
        insertadas = 0
        for linea in texto.strip().split("\n"):
            linea = linea.strip()
            if not linea:
                continue
            partes = [p.strip() for p in linea.split("\t")]
            if len(partes) < 5:
                partes = [p.strip() for p in linea.split() if p.strip()]
            if len(partes) < 5:
                continue
            nombre = partes[0]
            lgit_model = partes[1]
            pcb_type = partes[2]
            degrees = partes[3]
            top_housing = partes[4]
            housing_mess = partes[5] if len(partes) > 7 else ""
            bottom_cover = partes[6] if len(partes) > 7 else (partes[5] if len(partes) > 5 else "")
            opal = partes[7] if len(partes) > 7 else ""
            self.agregar_catalogo_modelo(
                familia_default, nombre, lgit_model, pcb_type, degrees,
                top_housing, housing_mess, bottom_cover, opal,
            )
            insertadas += 1
        return insertadas

    def cargar_catalogo_ejemplo_thunder(self):
        """Carga las filas del catalogo Thunder Trinity/Cheetah que se pudieron leer
        claramente de la foto proporcionada. Solo inserta si la tabla esta vacia,
        para no duplicar ni sobreescribir ediciones posteriores del usuario.
        ADVERTENCIA: un par de celdas en la foto original estaban borrosas/con reflejo
        (ej. 'Everest (Fascia)' y el codigo LGIT del grupo SRC de Cheetah) - revisa esas
        filas manualmente despues de cargar."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM catalogo_modelos")
        if cur.fetchone()[0] > 0:
            return 0

        trinity = [
            ("MY (Bi-Cam)", "ACA05S001X", "HORIZONTAL", "83°", "CORN", "", "B KEY", ""),
            ("MY (Bi-Cam)", "ACA05S002X", "HORIZONTAL", "46°", "CORN", "", "A KEY", ""),
            ("MY (SRC-L)", "ACS05S002L", "HORIZONTAL", "83°", "FLAT", "", "C KEY", ""),
            ("MY (SRC-R)", "ACS05S002R", "HORIZONTAL", "83°", "FLAT", "", "C KEY", ""),
            ("MY (B-Pillar-L)", "ACS05S001L", "VERTICAL", "83°", "CORN", "", "B KEY", ""),
            ("MY (B-Pillar-R)", "ACS05S001R", "VERTICAL", "83°", "CORN", "", "B KEY", ""),
            ("MY (Rear)", "ACR05S001X", "VERTICAL", "180°", "WIDE ARM", "", "D KEY", ""),
            ("Everest (Bi-Cam)", "ACA05S004X", "HORIZONTAL", "83°", "CORN", "", "B KEY", ""),
            ("Everest (Bi-Cam)", "ACA05S005X", "HORIZONTAL", "46°", "CORN", "", "A KEY", ""),
            # Everest (Fascia): codigo LGIT borroso/con reflejo en la foto original - VERIFICAR
            ("Everest (Fascia) [VERIFICAR CODIGO]", "ACA05C00?", "HORIZONTAL", "180°", "FLAT", "", "E KEY", ""),
            ("Everest (Fender-L)", "ACS05S003L", "HORIZONTAL", "83°", "CORN", "", "B KEY", ""),
            ("Everest (Fender-R)", "ACS05S003R", "HORIZONTAL", "83°", "CORN", "", "B KEY", ""),
            ("Everest (B-Pillar-L)", "ACS05S001L", "VERTICAL", "83°", "CORN", "", "B KEY", ""),
            ("Everest (B-Pillar-R)", "ACS05S001R", "VERTICAL", "83°", "CORN", "", "B KEY", ""),
            ("Everest (Rear)", "ACR05S002X", "VERTICAL", "180°", "WIDE ARM", "", "D KEY", ""),
            ("Highland (Bi-Cam)", "ACA05S003X", "HORIZONTAL", "83°", "CORN", "", "B KEY", ""),
            ("Highland (Bi-Cam)", "ACA05S006X", "HORIZONTAL", "46°", "CORN", "", "A KEY", ""),
            ("Highland (SRC-L)", "ACS05S002L", "HORIZONTAL", "83°", "FLAT", "", "C KEY", ""),
            ("Highland (SRC-R)", "ACS05S002R", "HORIZONTAL", "83°", "FLAT", "", "C KEY", ""),
            ("Highland (B-Pillar-L)", "ACS05S001L", "VERTICAL", "83°", "CORN", "", "B KEY", ""),
            ("Highland (B-Pillar-R)", "ACS05S001R", "VERTICAL", "83°", "CORN", "", "B KEY", ""),
            ("Highland (Rear)", "ACR05S001XMM10", "VERTICAL", "180°", "WIDE ARM", "", "D KEY", ""),
        ]
        for nombre, lgit, pcb_type, deg, top_h, mess, bottom, opal in trinity:
            self.agregar_catalogo_modelo("Thunder Trinity", nombre, lgit, pcb_type, deg, top_h, mess, bottom, opal)

        cheetah = [
            ("(Bi-Cam) ACA05S0014X10", "ACA05S0014XM10", "HORIZONTAL", "46°", "CORN 26A", "ACA1X00", "A KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0014X10", "ACA05S0014XM11", "HORIZONTAL", "46°", "CORN 26A", "ACA1X00", "A KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0014X10", "ACA05S0014XM12", "HORIZONTAL", "46°", "CORN 26A", "ACA1X00", "A KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0014X10", "ACA05S0014XM13", "HORIZONTAL", "46°", "CORN 26A", "ACA1X00", "A KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0014X10", "ACA05S0014XM14", "HORIZONTAL", "46°", "CORN 26A", "ACA1X00", "A KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0013X10", "ACA05S0013XM10", "HORIZONTAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0013X10", "ACA05S0013XM11", "HORIZONTAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0013X10", "ACA05S0013XM12", "HORIZONTAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0013X10", "ACA05S0013XM13", "HORIZONTAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            ("(Bi-Cam) ACA05S0013X10", "ACA05S0013XM14", "HORIZONTAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            # (SRC): codigo LGIT del grupo (columna izquierda) borroso en la foto - VERIFICAR
            ("(SRC) [VERIFICAR CODIGO GRUPO]", "ACS05S007LM00", "HORIZONTAL", "83°", "FLAT/PLAIN", "ACR5XM10", "C KEY", "HEATER"),
            ("(SRC) [VERIFICAR CODIGO GRUPO]", "ACS05S007LM10", "HORIZONTAL", "83°", "FLAT/PLAIN", "ACR5XM10", "C KEY", "HEATER"),
            ("(SRC) [VERIFICAR CODIGO GRUPO]", "ACR05S005XM10", "HORIZONTAL", "180°", "FLAT/PLAIN", "ACR5XM10", "A KEY", "HEATER"),
            ("(B-Pillar) ACS05S006L10", "ACS05S006LM10", "VERTICAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            ("(B-Pillar) ACS05S006L10", "ACS05S006RM10", "VERTICAL", "83°", "CORN 26A", "ACA1X00", "B KEY", "NO HEATER"),
            ("(FASCIA) ACA05S0015X10", "ACR05S005X11", "VERTICAL", "180°", "WIDE ARM", "ACA9X00", "E KEY", "HEATER"),
            ("(FASCIA) ACA05S0015X10", "ACA05S0015X10", "VERTICAL", "180°", "WIDE ARM", "ACA9X00", "E KEY", "HEATER"),
            ("(FASCIA) ACA05S0015X10", "ACA05S0015X11", "VERTICAL", "180°", "WIDE ARM", "ACA9X00", "E KEY", "HEATER"),
            ("(FASCIA) ACA05S0015X10", "ACA05S0015X12", "VERTICAL", "180°", "WIDE ARM", "ACA9X00", "E KEY", "HEATER"),
        ]
        for nombre, lgit, pcb_type, deg, top_h, mess, bottom, opal in cheetah:
            self.agregar_catalogo_modelo("Thunder Cheetah", nombre, lgit, pcb_type, deg, top_h, mess, bottom, opal)

        return len(trinity) + len(cheetah)

    # =========================================================================
    # INVENTARIO POR RACKS (gestion de almacen, formato Rack -> Trays -> Lot ID)
    # Reusa la misma logica de escaneo con deteccion de duplicados que Escaneo,
    # y el mismo historial unificado de toda la app.
    # =========================================================================

    def siguiente_nombre_rack_automatico(self):
        """Genera el siguiente nombre de rack: R1, R2, R3... (secuencial, basado
        en el contador AUTOINCREMENT, igual de robusto que el de charolas)."""
        cur = self.conn.cursor()
        cur.execute("SELECT COALESCE(MAX(id), 0) FROM inventario_racks")
        return f"R{cur.fetchone()[0] + 1}"

    def crear_rack(self, nombre, num_trays):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO inventario_racks (nombre, num_trays, creado) VALUES (?,?,?)",
            (nombre, num_trays, datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()
        self.registrar_historial("rack_creado", f"{nombre} ({num_trays} trays)")
        return cur.lastrowid

    def listar_racks(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, nombre, num_trays FROM inventario_racks ORDER BY id DESC")
        return cur.fetchall()

    def obtener_rack(self, rack_id):
        cur = self.conn.cursor()
        cur.execute("SELECT id, nombre, num_trays FROM inventario_racks WHERE id=?", (rack_id,))
        return cur.fetchone()

    def lot_id_ya_en_rack(self, rack_id, lot_id):
        """Deteccion de duplicados: un Lot ID solo puede aparecer UNA vez en
        todo el rack, sin importar en que tray se intente escanear."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT tray_numero FROM inventario_items WHERE rack_id=? AND lot_id=?", (rack_id, lot_id)
        )
        row = cur.fetchone()
        return row[0] if row else None

    def escanear_item_inventario(self, rack_id, tray_numero, lot_id):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO inventario_items (rack_id, tray_numero, lot_id, escaneado_en) VALUES (?,?,?,?)",
            (rack_id, tray_numero, lot_id, datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()
        self.registrar_historial("inventario_escaneado", f"rack {rack_id} tray {tray_numero}: {lot_id}")

    def items_de_tray(self, rack_id, tray_numero):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT lot_id FROM inventario_items WHERE rack_id=? AND tray_numero=? ORDER BY id",
            (rack_id, tray_numero),
        )
        return [r[0] for r in cur.fetchall()]

    def qty_por_tray(self, rack_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT tray_numero, COUNT(*) FROM inventario_items WHERE rack_id=? GROUP BY tray_numero",
            (rack_id,),
        )
        return dict(cur.fetchall())

    def deshacer_ultimo_item_inventario(self, rack_id, tray_numero):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id, lot_id FROM inventario_items WHERE rack_id=? AND tray_numero=? ORDER BY id DESC LIMIT 1",
            (rack_id, tray_numero),
        )
        row = cur.fetchone()
        if not row:
            return None
        item_id, lot_id = row
        cur.execute("DELETE FROM inventario_items WHERE id=?", (item_id,))
        self.conn.commit()
        self.registrar_historial("inventario_deshacer", f"rack {rack_id} tray {tray_numero}: {lot_id}")
        return lot_id

    def eliminar_item_inventario(self, rack_id, lot_id):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM inventario_items WHERE rack_id=? AND lot_id=?", (rack_id, lot_id))
        self.conn.commit()
        return cur.rowcount > 0

    def exportar_rack_a_matriz(self, rack_id):
        """Arma la estructura de datos (sin tocar Excel aun) para exportar un
        rack en el formato de la hoja: por cada tray, su QTY y su lista de
        Lot ID en orden de escaneo. Columnas se emparejan a la altura del
        tray con mas piezas (el resto se deja en blanco)."""
        rack = self.obtener_rack(rack_id)
        if not rack:
            return None
        _id, nombre, num_trays = rack
        qtys = self.qty_por_tray(rack_id)
        trays = []
        for t in range(1, num_trays + 1):
            items = self.items_de_tray(rack_id, t)
            trays.append({"tray_numero": t, "qty": len(items), "items": items})
        total_qty = sum(t["qty"] for t in trays)
        return {"nombre": nombre, "total_qty": total_qty, "trays": trays}

    def exportar_todos_los_racks_a_matriz(self):
        return [self.exportar_rack_a_matriz(rid) for rid, _n, _nt in self.listar_racks()]

    # =========================================================================

    def guardar_remapeo(self, carrier_base, mapeo_cid_a_lot):
        """Guarda un pallet completo remapeado. mapeo_cid_a_lot: dict {carrier_id: lot_id}.
        Comparte el mismo historial que retrabajo (registrar_historial)."""
        fecha, turno = obtener_turno()
        hora = datetime.now().strftime("%H:%M:%S")
        ahora = datetime.now().isoformat(timespec="seconds")
        cur = self.conn.cursor()
        for cid, lot in mapeo_cid_a_lot.items():
            if not lot:
                continue
            cur.execute(
                """INSERT INTO remapeos (fecha, turno, hora, carrier_base, carrier_id, lot_id, creado_en)
                   VALUES (?,?,?,?,?,?,?)""",
                (str(fecha), turno, hora, carrier_base, cid, lot, ahora),
            )
        self.conn.commit()
        self.registrar_historial(
            "remapeo_pallet", f"{carrier_base}: {sum(1 for v in mapeo_cid_a_lot.values() if v)} piezas remapeadas"
        )

    def contar_veces_remapeado(self, lot_id):
        """Cuántas veces este Lot ID ha sido remapeado en TODA la historia."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM remapeos WHERE lot_id=?", (lot_id,))
        return cur.fetchone()[0]

    def veces_retrabajado(self, lot_id):
        """Cuántas veces este Lot ID ha pasado por el modulo de retrabajo (escaneado).
        Esto permite la regla de negocio: tras 2 remapeos, la pieza DEBE pasar por
        retrabajo antes de poder remapearse de nuevo (si no, se scrapea)."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM pcbs WHERE lot_id=?", (lot_id,))
        return cur.fetchone()[0]

    def marcar_liberado(self, carrier_id, lot_id):
        cur = self.conn.cursor()
        try:
            cur.execute(
                "INSERT INTO remapeo_liberados (carrier_id, lot_id, fecha_liberado) VALUES (?,?,?)",
                (carrier_id, lot_id, datetime.now().isoformat(timespec="seconds")),
            )
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False  # ya estaba liberado

    def esta_liberado(self, carrier_id, lot_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT 1 FROM remapeo_liberados WHERE carrier_id=? AND lot_id=?", (carrier_id, lot_id)
        )
        return cur.fetchone() is not None

    def buscar_remapeos(self, filtro_carrier="", desde=None, hasta=None, solo_estado=None):
        """Busca remapeos con filtros. solo_estado: None|'Pendientes'|'Liberados'."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT fecha, turno, hora, carrier_id, lot_id FROM remapeos ORDER BY fecha, hora"
        )
        filas = cur.fetchall()
        resultado = []
        for fecha, turno, hora, carrier_id, lot_id in filas:
            if filtro_carrier and filtro_carrier.upper() not in carrier_id.upper():
                continue
            try:
                fecha_d = _dt.date.fromisoformat(fecha)
            except ValueError:
                continue
            if desde and fecha_d < desde:
                continue
            if hasta and fecha_d > hasta:
                continue
            liberado = self.esta_liberado(carrier_id, lot_id)
            if solo_estado == "Pendientes" and liberado:
                continue
            if solo_estado == "Liberados" and not liberado:
                continue
            resultado.append((fecha, turno, hora, carrier_id, lot_id, "Liberado" if liberado else "Pendiente"))
        return resultado

    def remapeos_del_turno(self, fecha, turno):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT carrier_id, lot_id FROM remapeos WHERE fecha=? AND turno=?",
            (str(fecha), turno),
        )
        return cur.fetchall()

    def listar_carriers_agrupados(self, filtro=""):
        """Agrupa remapeos por carrier base (sin el sufijo de posicion)."""
        cur = self.conn.cursor()
        cur.execute("SELECT fecha, turno, hora, carrier_id, lot_id FROM remapeos")
        filas = cur.fetchall()
        agrupado = {}
        for fecha, turno, hora, carrier_id, lot_id in filas:
            base = "-".join(carrier_id.split("-")[:-1]) if "-" in carrier_id else carrier_id
            if filtro and filtro.upper() not in base.upper():
                continue
            agrupado.setdefault(base, []).append((fecha, turno, hora, carrier_id, lot_id))
        return agrupado

    def metricas_remapeo_por_turno(self):
        cur = self.conn.cursor()
        cur.execute("SELECT fecha, turno FROM remapeos")
        metricas = {}
        for fecha, turno in cur.fetchall():
            metricas.setdefault(fecha, {1: 0, 2: 0, 3: 0})
            metricas[fecha][turno] += 1
        return metricas

    def historial_completo_lot_id(self, lot_id):
        """SIMBIOSIS: historial cruzado de una pieza a traves de AMBOS modulos
        (remapeo + retrabajo), ordenado cronologicamente. Esto es lo que
        conecta de verdad los dos programas en una sola linea de tiempo."""
        eventos = []

        cur = self.conn.cursor()
        cur.execute(
            "SELECT fecha, hora, carrier_id, turno FROM remapeos WHERE lot_id=? ORDER BY fecha, hora",
            (lot_id,),
        )
        for fecha, hora, carrier_id, turno in cur.fetchall():
            eventos.append((f"{fecha} {hora}", "Remapeo", f"Carrier {carrier_id}, turno {turno}"))

        cur.execute(
            """SELECT p.escaneado_en, c.nombre, p.posicion, p.modelo, p.proceso, p.separada
               FROM pcbs p JOIN charolas c ON p.charola_id = c.id
               WHERE p.lot_id=? ORDER BY p.escaneado_en""",
            (lot_id,),
        )
        for escaneado_en, charola_nombre, posicion, modelo, proceso, separada in cur.fetchall():
            detalle = f"Charola {charola_nombre} pos {posicion}"
            if modelo:
                detalle += f", modelo {modelo}"
            if proceso:
                detalle += f", proceso {proceso}"
            if separada:
                detalle += " (separada)"
            eventos.append((escaneado_en or "", "Retrabajo/Escaneo", detalle))

        eventos.sort(key=lambda e: e[0])
        return eventos

    def listar_todas_las_charolas_clasificadas(self):
        """Lista charolas que tienen al menos algunos datos clasificados."""
        cur = self.conn.cursor()
        cur.execute(
            """SELECT DISTINCT c.id, c.nombre, c.filas, c.columnas, c.volteada
               FROM charolas c
               INNER JOIN pcbs p ON c.id = p.charola_id
               WHERE p.modelo IS NOT NULL
               ORDER BY c.id"""
        )
        return cur.fetchall()

    def contar_sin_clasificar_en_charola(self, charola_id):
        """Cuántas piezas en charola sin modelo/proceso."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM pcbs WHERE charola_id=? AND (modelo IS NULL OR proceso IS NULL)",
            (charola_id,),
        )
        return cur.fetchone()[0]

    def contar_totales_en_charola(self, charola_id):
        """Total de piezas escaneadas en charola."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM pcbs WHERE charola_id=?", (charola_id,))
        return cur.fetchone()[0]

    def actualizar_modelo_proceso_batch(self, mapeo_lot_id):
        """Actualiza modelo/proceso para múltiples Lot ID de una sola vez.
        
        mapeo_lot_id: dict {lot_id: (modelo, proceso)}
        Retorna: (actualizados, no_encontrados)
        """
        cur = self.conn.cursor()
        actualizados = 0
        no_encontrados = []
        ahora = datetime.now().isoformat(timespec="seconds")

        for lot_id, (modelo, proceso) in mapeo_lot_id.items():
            resultado = cur.execute(
                "UPDATE pcbs SET modelo=?, proceso=?, actualizado_en=? WHERE lot_id=?",
                (modelo, proceso, ahora, lot_id),
            )
            if resultado.rowcount > 0:
                actualizados += 1
            else:
                no_encontrados.append(lot_id)

        self.conn.commit()
        return actualizados, no_encontrados

    def exportar_charola_a_dict(self, charola_id):
        """Exporta una charola completa como dict."""
        cur = self.conn.cursor()
        cur.execute("SELECT nombre, filas, columnas FROM charolas WHERE id=?", (charola_id,))
        charola_info = cur.fetchone()
        if not charola_info:
            return None
        nombre, filas, columnas = charola_info
        
        datos = self.pcbs_de_charola(charola_id)
        return {
            "nombre": nombre,
            "filas": filas,
            "columnas": columnas,
            "posiciones": datos,
        }

    def exportar_todas_charolas_a_dict(self):
        """Exporta TODAS las charolas con datos (para el Excel consolidado)."""
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM charolas ORDER BY id")
        charolas_ids = [row[0] for row in cur.fetchall()]

        resultado = {}
        for cid in charolas_ids:
            exp = self.exportar_charola_a_dict(cid)
            if exp and exp["posiciones"]:  # solo si tiene datos
                resultado[cid] = exp
        return resultado

    def voltear_charola_180(self, charola_id):
        """
        Gira 180 grados TODA la charola de una sola vez (fila y columna invertidas),
        preservando lot_id, modelo, proceso, separada y fechas de cada pieza.
        Usa un DELETE + re-INSERT para evitar choques con la restriccion UNIQUE
        de posicion mientras se reasignan las nuevas posiciones.
        """
        cur = self.conn.cursor()
        cur.execute("SELECT filas, columnas FROM charolas WHERE id=?", (charola_id,))
        row = cur.fetchone()
        if not row:
            return 0
        filas, columnas = row

        cur.execute(
            """SELECT posicion, lot_id, modelo, proceso, separada, escaneado_en, actualizado_en
               FROM pcbs WHERE charola_id=?""",
            (charola_id,),
        )
        piezas = cur.fetchall()
        if not piezas:
            return 0

        cur.execute("DELETE FROM pcbs WHERE charola_id=?", (charola_id,))
        for posicion, lot_id, modelo, proceso, separada, escaneado_en, actualizado_en in piezas:
            nueva_pos = voltear_posicion_180_completo(posicion, filas, columnas)
            cur.execute(
                """INSERT INTO pcbs (charola_id, posicion, lot_id, modelo, proceso, separada,
                                      escaneado_en, actualizado_en)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (charola_id, nueva_pos, lot_id, modelo, proceso, separada, escaneado_en, actualizado_en),
            )
        self.conn.commit()
        self.registrar_historial("charola_volteada_180", f"Charola {charola_id}: {len(piezas)} piezas giradas 180°")
        return len(piezas)

    def marcar_volteo_180_aplicado(self, charola_id):
        cur = self.conn.cursor()
        cur.execute("UPDATE charolas SET volteo_180_aplicado=1 WHERE id=?", (charola_id,))
        self.conn.commit()

    def charolas_pendientes_de_volteo_automatico(self):
        """Charolas marcadas para volteo automatico de 180° que aun no se les ha aplicado.
        Se usa justo cuando se pasan/clasifican datos (importar Excel o pegar texto),
        para voltearlas UNA sola vez sin que el usuario tenga que acordarse del boton."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM charolas WHERE auto_volteo_180=1 AND volteo_180_aplicado=0"
        )
        return [r[0] for r in cur.fetchall()]

    def aplicar_volteos_automaticos_pendientes(self):
        """Aplica voltear_charola_180 a todas las charolas pendientes de volteo
        automatico, y las marca como aplicadas para no voltearlas dos veces.
        Retorna la lista de (charola_id, nombre, piezas_volteadas) para poder avisar."""
        resultados = []
        for charola_id in self.charolas_pendientes_de_volteo_automatico():
            n = self.voltear_charola_180(charola_id)
            self.marcar_volteo_180_aplicado(charola_id)
            if n:
                cur = self.conn.cursor()
                cur.execute("SELECT nombre FROM charolas WHERE id=?", (charola_id,))
                nombre = cur.fetchone()[0]
                resultados.append((charola_id, nombre, n))
        return resultados

    def limpiar_charola_completa(self, charola_id):
        """Elimina TODOS los PCBs escaneados de una charola (pero conserva la charola)."""
        cur = self.conn.cursor()
        cur.execute("DELETE FROM pcbs WHERE charola_id=?", (charola_id,))
        self.conn.commit()
        self.registrar_historial("charola_limpiada", f"Charola {charola_id}: {cur.rowcount} piezas eliminadas")

    def bloquear_posicion(self, charola_id, posicion):
        """Bloquea una posición para que no pueda ser escaneada."""
        cur = self.conn.cursor()
        try:
            cur.execute(
                "INSERT INTO posiciones_bloqueadas (charola_id, posicion) VALUES (?,?)",
                (charola_id, posicion),
            )
            self.conn.commit()
            self.registrar_historial("posicion_bloqueada", f"Charola {charola_id}, pos {posicion}")
        except sqlite3.IntegrityError:
            pass  # ya existe

    def desbloquear_posicion(self, charola_id, posicion):
        """Desbloquea una posición."""
        cur = self.conn.cursor()
        cur.execute(
            "DELETE FROM posiciones_bloqueadas WHERE charola_id=? AND posicion=?",
            (charola_id, posicion),
        )
        self.conn.commit()
        if cur.rowcount > 0:
            self.registrar_historial("posicion_desbloqueada", f"Charola {charola_id}, pos {posicion}")

    def posiciones_bloqueadas_de_charola(self, charola_id):
        """Retorna lista de posiciones bloqueadas."""
        cur = self.conn.cursor()
        cur.execute("SELECT posicion FROM posiciones_bloqueadas WHERE charola_id=?", (charola_id,))
        return {row[0] for row in cur.fetchall()}

    def actualizar_modelo_proceso_desde_texto(self, texto):
        """Parsea texto 'espagueti' del RFC (múltiples espacios/tabs).
        
        Normaliza espacios y detecta automáticamente columnas.
        Retorna: dict {lot_id: (modelo, proceso)}, list errores
        """
        lineas = texto.strip().split("\n")
        mapeo = {}
        errores = []
        
        # normalizar cada linea: reemplaza tabs y espacios múltiples con |
        lineas_normalizadas = []
        for linea in lineas:
            # reemplaza tabs y espacios múltiples con separador consistente
            import re
            linea_norm = re.sub(r'[\t ]+', '|', linea.strip())
            lineas_normalizadas.append(linea_norm)
        
        # detectar estructura: probablemente columnas son Lot_ID | Producto | ... | Proceso | ...
        # heuristica: Lot_ID suele ser alfanumerico con numeros, Proceso puede contener palabras como "Install", "Screw"
        
        for i, linea in enumerate(lineas_normalizadas, start=1):
            if not linea.strip():
                continue
            
            partes = linea.split("|")
            partes = [p.strip() for p in partes if p.strip()]  # eliminar vacios
            
            if len(partes) < 2:
                errores.append(f"Línea {i}: insuficientes datos (esperaba al menos 2 campos)")
                continue
            
            # heuristica: primer campo suele ser Lot_ID (alfanumérico con números)
            lot_id = partes[0]
            
            # validar que sea un Lot_ID razonable (no "Product", "MASS", etc)
            if any(palabra in lot_id.lower() for palabra in ["product", "mass", "process", "hold", "array"]):
                continue  # es una linea de encabezado o basura
            
            # el resto de campos: buscar uno que parezca descripcion (producto/modelo)
            # y otro que parezca proceso
            modelo_cand = None
            proceso_cand = None
            
            for j in range(1, len(partes)):
                parte = partes[j]
                # buscar campo que contenga patrones de código de producto (puntos, guiones)
                if "." in parte or "-" in parte:
                    modelo_cand = parte
                # buscar proceso: Install, Screw, etc
                elif any(proc in parte for proc in ["Install", "Screw", "AVI", "TOP", "WAIT", "PROCESSING"]):
                    proceso_cand = parte
            
            # fallback: si no encontró, usar posiciones fijas
            if not modelo_cand and len(partes) > 1:
                modelo_cand = partes[1]
            if not proceso_cand and len(partes) > 2:
                proceso_cand = partes[2]
            
            modelo = modelo_cand or "N/D"
            proceso = proceso_cand or "N/D"
            
            mapeo[lot_id] = (modelo, proceso)
        
        return mapeo, errores

    def parsear_y_mostrar_preview(self, texto):
        """Parsea texto y retorna lista de (lot_id, modelo, proceso) para previsualizar."""
        mapeo, errores = self.actualizar_modelo_proceso_desde_texto(texto)
        preview = [(lot_id, mod, proc) for lot_id, (mod, proc) in mapeo.items()]
        return preview, errores

    def actualizar_modelo_proceso(self, lot_id, modelo, proceso):
        cur = self.conn.cursor()
        ahora = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            "UPDATE pcbs SET modelo=?, proceso=?, actualizado_en=? WHERE lot_id=?",
            (modelo, proceso, ahora, lot_id),
        )
        self.conn.commit()
        return cur.rowcount

    def todos_los_lot_ids_pendientes_de_cruce(self):
        cur = self.conn.cursor()
        cur.execute("SELECT DISTINCT lot_id FROM pcbs WHERE modelo IS NULL")
        return [r[0] for r in cur.fetchall()]

    def todos_los_lot_ids(self):
        cur = self.conn.cursor()
        cur.execute("SELECT DISTINCT lot_id FROM pcbs")
        return [r[0] for r in cur.fetchall()]

    # --- historial ---
    def registrar_historial(self, evento, detalle):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO historial (evento, detalle, fecha) VALUES (?,?,?)",
            (evento, detalle, datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()

    def leer_historial(self, limite=200):
        cur = self.conn.cursor()
        cur.execute("SELECT evento, detalle, fecha FROM historial ORDER BY id DESC LIMIT ?", (limite,))
        return cur.fetchall()


# --------------------------------------------------------------------------
# Utilidades de mapeo tipo ajedrez
# --------------------------------------------------------------------------
def obtener_turno():
    """Calcula fecha y turno (1/2/3) segun la hora actual. Portado de remapeo_4.0.1.py."""
    ahora = datetime.now()
    hora = ahora.hour
    fecha = ahora.date()
    if 7 <= hora < 15:
        turno = 1
    elif 15 <= hora < 23:
        turno = 2
    else:
        turno = 3
        fecha = fecha - _dt.timedelta(days=1)
    return fecha, turno


def es_formato_carrier(texto, longitud_carrier):
    """Un Carrier tiene guiones y la longitud configurada."""
    return "-" in texto and len(texto) == longitud_carrier


def es_formato_pcb_remapeo(texto, longitud_pcb):
    """Un PCB (para remapeo) es alfanumerico, sin guiones, longitud fija."""
    return "-" not in texto and texto.isalnum() and len(texto) == longitud_pcb


def etiqueta_columna(idx):
    """0 -> A, 1 -> B ... 25 -> Z, 26 -> AA ..."""
    letras = string.ascii_uppercase
    resultado = ""
    idx += 1
    while idx > 0:
        idx, resto = divmod(idx - 1, 26)
        resultado = letras[resto] + resultado
    return resultado


def generar_posiciones(filas, columnas):
    """Genera posiciones canonicas tipo ajedrez recorriendo fila por fila."""
    posiciones = []
    for f in range(filas):
        for c in range(columnas):
            posiciones.append(f"{etiqueta_columna(c)}{f + 1}")
    return posiciones


def voltear_posicion(pos, filas):
    """
    Traduce una posicion cuando la charola se voltea 180 grados de arriba-abajo
    para escanear el QR que queda boca abajo (tipico en charolas de 36).

    La columna se mantiene igual; la fila se invierte:
        fila_nueva = (filas + 1) - fila_original
    Ej. con 6 filas: A1 <-> A6, C3 <-> C4, C5 <-> C2.
    """
    col_letra = "".join(ch for ch in pos if ch.isalpha())
    fila_num = int("".join(ch for ch in pos if ch.isdigit()))
    fila_nueva = (filas + 1) - fila_num
    return f"{col_letra}{fila_nueva}"


def voltear_posicion_180_completo(pos, filas, columnas):
    """
    Gira la charola COMPLETA 180 grados (como girar un plato sobre la mesa):
    tanto la fila como la columna se invierten.

    fila_nueva = (filas + 1) - fila_original
    columna_nueva = (columnas + 1) - columna_original (por indice de letra)

    Ej. con 6 filas x 8 columnas (A-H): A1 <-> H6, B1 <-> G6, D3 <-> E4, etc.
    Distinto de voltear_posicion(): aqui ambos ejes se invierten, no solo la fila.
    """
    col_letra = "".join(ch for ch in pos if ch.isalpha())
    fila_num = int("".join(ch for ch in pos if ch.isdigit()))
    idx_col = _col_letra_a_indice(col_letra)
    idx_col_nuevo = columnas - 1 - idx_col
    fila_nueva = (filas + 1) - fila_num
    return f"{etiqueta_columna(idx_col_nuevo)}{fila_nueva}"


def orden_escaneo(filas, columnas, volteada):
    """
    Devuelve la lista de posiciones CANONICAS en el orden real en que se van
    llenando conforme el operador escanea fisicamente la charola.

    Si la charola esta volteada, el operador recorre la charola en su orden
    fisico normal (fila por fila) pero cada pieza que toca en realidad
    corresponde a la posicion canonica opuesta verticalmente.
    """
    fisico = generar_posiciones(filas, columnas)
    if not volteada:
        return fisico
    return [voltear_posicion(p, filas) for p in fisico]


# --------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Retrabajo y Clasificacion de PCBs")
        self.geometry("1150x720")
        self.db = Database()

        # atributos generales
        self.charola_actual_id = None
        self.filas = 6
        self.columnas = 6
        self.volteada_actual = 0
        self.celdas_sep_seleccionadas = set()
        self.combo_charola = None
        self.combo_charola_sep = None
        self.orden_estricto = tk.BooleanVar(value=False)
        self.vista_color = tk.StringVar(value="proceso")
        self.usar_modelo_base = tk.BooleanVar(value=False)
        self.vista_color_sep = tk.StringVar(value="modelo")
        self.celdas = {}
        self.celdas_sep = {}
        self.color_map_modelo = {}
        self.color_map_proceso = {}
        
        # atributos de UI que se crean en _construir_tab_*
        self.text_pega = None
        self.entry_lot_id = None
        self.entry_posicion_bloqueo = None
        self.frame_grid = None
        self.frame_grid_separacion = None
        self.frame_picking = None
        self.frame_leyenda = None
        self.frame_leyenda_sep = None
        self.lbl_progreso = None
        self.lbl_import_status = None
        self.canvas_picking = None
        self.tree_historial = None
        self.tree_equivalencias = None
        self.combo_charola_sep = None

        # --- estado del modulo de Remapeo (fusionado desde remapeo_4.0.1.py) ---
        self.LONGITUD_CARRIER = 10   # ej. LG1T-15-07
        self.LONGITUD_PCB_REMAPEO = 15  # ej. D42673OVB100207
        self.remap_mapeo = {}            # carrier_id -> lot_id o None
        self.remap_carrier_ids = []
        self.remap_carrier_base = None
        self.remap_bloqueados = set()
        self.remap_reincidentes = set()  # carrier_ids cuyo lot_id ya estaba remapeado antes
        self.remap_orden_escaneos = []   # pila para deshacer: (cid, lot_id)
        self.remap_posiciones_canvas = {}  # rect_id (canvas) -> carrier_id
        self.tabla_remapeo = None
        self.canvas_remapeo = None
        self.entry_carrier_remapeo = None
        self.entry_pcb_remapeo = None
        self.lbl_estado_remapeo = None
        self.lbl_contador_remapeo = None
        self.lbl_pallets_remapeo = None
        self.lbl_reloj_remapeo = None

        # --- estado del modulo de Inventario por Racks ---
        self.rack_actual_id = None
        self.combo_rack = None
        self.combo_tray_actual = None
        self.entry_lot_inventario = None
        self.tree_resumen_racks = None
        self.lbl_estado_inventario = None

        self._construir_ui()

    # ---------------- UI construction ----------------
    def _construir_ui(self):
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)

        self.tab_remapeo = ttk.Frame(notebook)
        self.tab_escaneo = ttk.Frame(notebook)
        self.tab_clasificacion = ttk.Frame(notebook)
        self.tab_separacion = ttk.Frame(notebook)
        self.tab_modelos_base = ttk.Frame(notebook)
        self.tab_inventario = ttk.Frame(notebook)
        self.tab_historial = ttk.Frame(notebook)

        notebook.add(self.tab_remapeo, text="Remapeo")
        notebook.add(self.tab_escaneo, text="Escaneo")
        notebook.add(self.tab_clasificacion, text="Clasificacion / Excel")
        notebook.add(self.tab_separacion, text="Separacion")
        notebook.add(self.tab_modelos_base, text="Modelos base")
        notebook.add(self.tab_inventario, text="Inventario (Racks)")
        notebook.add(self.tab_historial, text="Historial")

        self._construir_tab_remapeo()
        self._construir_tab_escaneo()
        self._construir_tab_clasificacion()
        self._construir_tab_separacion()
        self._construir_tab_modelos_base()
        self._construir_tab_inventario()
        self._construir_tab_historial()
        
        # refrescar combos DESPUES de que todas las pestañas esten construidas
        self._refrescar_combo_charolas()
        self._actualizar_reloj_remapeo()
        self._actualizar_contador_pallets_remapeo()
        self._dibujar_mapa_remapeo()
        self._refrescar_combo_racks()

    def _construir_tab_remapeo(self):
        frame_izq = ttk.Frame(self.tab_remapeo, padding=10)
        frame_izq.pack(side="left", fill="y")

        frame_der = ttk.Frame(self.tab_remapeo, padding=10)
        frame_der.pack(side="left", fill="both", expand=True)

        # --- columna izquierda: reloj, mapa visual, leyenda, contadores ---
        self.lbl_reloj_remapeo = ttk.Label(frame_izq, text="", font=("Segoe UI", 10, "bold"))
        self.lbl_reloj_remapeo.pack(pady=(0, 8))

        self.canvas_remapeo = tk.Canvas(frame_izq, width=500, height=132, bg="white")
        self.canvas_remapeo.pack(pady=5)
        self.canvas_remapeo.bind("<Button-1>", self._click_mapa_remapeo)

        self.lbl_contador_remapeo = ttk.Label(frame_izq, text="Ocupados: 0 | Libres: 0")
        self.lbl_contador_remapeo.pack(pady=5)

        self.lbl_pallets_remapeo = ttk.Label(frame_izq, text="Pallets completados (Turno -): 0")
        self.lbl_pallets_remapeo.pack(pady=(0, 8))

        frame_leyenda_remap = ttk.Frame(frame_izq)
        frame_leyenda_remap.pack(pady=5)
        leyenda_items = [
            ("red", "Libre"),
            ("green", "Ocupado"),
            ("orange", "PCB ya remapeado antes"),
            ("dimgray", "Bloqueado"),
            ("yellow", "Recién asignado"),
        ]
        for color, texto in leyenda_items:
            item = ttk.Frame(frame_leyenda_remap)
            item.pack(anchor="w", pady=1)
            tk.Canvas(item, width=14, height=14, bg=color, highlightthickness=1, highlightbackground="black").pack(
                side="left"
            )
            ttk.Label(item, text=texto, font=("Segoe UI", 8)).pack(side="left", padx=4)

        # --- columna derecha: tabla, entradas, estado, botones ---
        self.tabla_remapeo = ttk.Treeview(
            frame_der, columns=("Carrier ID", "LOT ID"), show="headings", height=8
        )
        self.tabla_remapeo.heading("Carrier ID", text="Carrier ID")
        self.tabla_remapeo.heading("LOT ID", text="LOT ID")
        self.tabla_remapeo.pack(pady=5, fill="x")

        ttk.Label(frame_der, text=f"Escanea Carrier Base ({self.LONGITUD_CARRIER} caracteres, ej. LG1T-15-07):").pack()
        self.entry_carrier_remapeo = ttk.Entry(frame_der)
        self.entry_carrier_remapeo.pack(pady=5, fill="x")
        self.entry_carrier_remapeo.bind("<Return>", self._escanear_carrier_remapeo)

        ttk.Label(frame_der, text=f"Escanea PCB / Lot ID ({self.LONGITUD_PCB_REMAPEO} caracteres):").pack()
        self.entry_pcb_remapeo = ttk.Entry(frame_der)
        self.entry_pcb_remapeo.pack(pady=5, fill="x")
        self.entry_pcb_remapeo.bind("<Return>", self._escanear_pcb_remapeo)

        self.lbl_estado_remapeo = ttk.Label(frame_der, text="", font=("Segoe UI", 9, "bold"))
        self.lbl_estado_remapeo.pack(pady=3)

        frame_botones = ttk.Frame(frame_der)
        frame_botones.pack(pady=5)

        ttk.Button(frame_botones, text="Copiar al portapapeles", command=self._copiar_portapapeles_remapeo).grid(
            row=0, column=0, padx=3, pady=3
        )
        ttk.Button(frame_botones, text="Deshacer último escaneo", command=self._deshacer_ultimo_remapeo).grid(
            row=0, column=1, padx=3, pady=3
        )
        ttk.Button(frame_botones, text="Reiniciar mapa", command=self._reiniciar_mapa_remapeo).grid(
            row=1, column=0, padx=3, pady=3
        )
        ttk.Button(frame_botones, text="Ver historial del día/turno", command=self._ver_historial_remapeo).grid(
            row=1, column=1, padx=3, pady=3
        )
        ttk.Button(frame_botones, text="Buscar / Liberar historial", command=self._buscar_historial_remapeo_ventana).grid(
            row=2, column=0, columnspan=2, padx=3, pady=3, sticky="we"
        )
        ttk.Button(
            frame_botones, text="Ver Carriers escaneados (Línea 1 / 2)",
            command=self._ventana_carriers_escaneados,
        ).grid(row=3, column=0, columnspan=2, padx=3, pady=3, sticky="we")
        if MATPLOTLIB_DISPONIBLE:
            ttk.Button(frame_botones, text="Mostrar gráficas", command=self._mostrar_graficas_remapeo).grid(
                row=4, column=0, columnspan=2, padx=3, pady=3, sticky="we"
            )
        else:
            ttk.Label(
                frame_botones, text="(matplotlib no instalado: gráficas deshabilitadas)",
                foreground="#888", font=("Segoe UI", 8),
            ).grid(row=4, column=0, columnspan=2, pady=3)

    # =========================================================================
    # Comportamiento del modulo de Remapeo (fusionado desde remapeo_4.0.1.py)
    # =========================================================================

    def _mostrar_estado_remapeo(self, msg, color="black", ms=4000, beep=False):
        self.lbl_estado_remapeo.config(text=msg, foreground=color)
        if beep:
            self.bell()
        if ms:
            self.after(ms, lambda: self.lbl_estado_remapeo.config(text=""))

    def _actualizar_reloj_remapeo(self):
        fecha, turno = obtener_turno()
        ahora = datetime.now().strftime("%H:%M:%S")
        self.lbl_reloj_remapeo.config(text=f"{ahora}  |  Turno {turno}  ({fecha})")
        self.after(1000, self._actualizar_reloj_remapeo)

    def _actualizar_contador_pallets_remapeo(self):
        fecha, turno = obtener_turno()
        filas = self.db.remapeos_del_turno(fecha, turno)
        pallets = {"-".join(cid.split("-")[:-1]) for cid, _lot in filas}
        self.lbl_pallets_remapeo.config(text=f"Pallets completados (Turno {turno}): {len(pallets)}")

    def _dibujar_mapa_remapeo(self, resaltar=None):
        canvas = self.canvas_remapeo
        canvas.delete("all")
        self.remap_posiciones_canvas = {}
        ancho, alto, espacio = 70, 46, 10
        ocupados = sum(1 for v in self.remap_mapeo.values() if v)
        libres = len(self.remap_mapeo) - ocupados
        self.lbl_contador_remapeo.config(text=f"Ocupados: {ocupados} | Libres: {libres}")

        for fila in range(2):
            for col in range(6):
                idx = fila * 6 + col + 1
                cid = None
                if not self.remap_carrier_ids:
                    texto, color = f"{idx:02d}", "gray"
                else:
                    cid = self.remap_carrier_ids[idx - 1]
                    texto = cid.split("-")[-1]
                    if cid in self.remap_bloqueados:
                        color = "dimgray"
                    elif self.remap_mapeo.get(cid):
                        texto = self.remap_mapeo[cid]
                        color = "orange" if cid in self.remap_reincidentes else "green"
                    else:
                        color = "red"
                    if resaltar == cid:
                        color = "yellow"
                x, y = col * (ancho + espacio), fila * (alto + espacio)
                rect = canvas.create_rectangle(x, y, x + ancho, y + alto, fill=color)
                canvas.create_text(x + ancho / 2, y + alto / 2, text=texto, fill="white", font=("Arial", 9, "bold"))
                if cid is not None:
                    self.remap_posiciones_canvas[rect] = cid

    def _click_mapa_remapeo(self, event):
        cercano = self.canvas_remapeo.find_closest(event.x, event.y)
        if not cercano:
            return
        rect_id = cercano[0]
        cid = self.remap_posiciones_canvas.get(rect_id)
        if cid:
            if cid in self.remap_bloqueados:
                self.remap_bloqueados.remove(cid)
            else:
                self.remap_bloqueados.add(cid)
            self._dibujar_mapa_remapeo()

    def _escanear_carrier_remapeo(self, event=None):
        carrier_base = self.entry_carrier_remapeo.get().strip()
        if not carrier_base:
            return

        if es_formato_pcb_remapeo(carrier_base, self.LONGITUD_PCB_REMAPEO):
            self._mostrar_estado_remapeo("⚠ Eso es un PCB, no un Carrier. Escanéalo en el campo de PCB.", "red", beep=True)
            self.entry_carrier_remapeo.delete(0, tk.END)
            return

        if len(carrier_base) != self.LONGITUD_CARRIER:
            messagebox.showerror("Error", f"El Carrier debe tener {self.LONGITUD_CARRIER} caracteres.")
            self.entry_carrier_remapeo.delete(0, tk.END)
            return

        prefix_parts = carrier_base.split("-")
        base = f"{prefix_parts[0]}-{prefix_parts[1]}" if len(prefix_parts) >= 2 else carrier_base
        self.remap_carrier_base = carrier_base
        self.remap_carrier_ids = [f"{base}-{str(i).zfill(2)}" for i in range(1, 13)]
        self.remap_mapeo = {cid: None for cid in self.remap_carrier_ids}
        self.remap_bloqueados.clear()
        self.remap_reincidentes.clear()
        self.remap_orden_escaneos.clear()

        for item in self.tabla_remapeo.get_children():
            self.tabla_remapeo.delete(item)
        for cid in self.remap_carrier_ids:
            self.tabla_remapeo.insert("", "end", iid=cid, values=(cid, ""))

        self._dibujar_mapa_remapeo()
        self.entry_carrier_remapeo.delete(0, tk.END)
        self.entry_pcb_remapeo.focus_set()

    def _escanear_pcb_remapeo(self, event=None):
        lot_id = self.entry_pcb_remapeo.get().strip()
        if not lot_id:
            return

        if not self.remap_carrier_ids:
            self._mostrar_estado_remapeo("⚠ Primero escanea el Carrier del pallet.", "red", beep=True)
            self.entry_pcb_remapeo.delete(0, tk.END)
            return

        if es_formato_carrier(lot_id, self.LONGITUD_CARRIER):
            self._mostrar_estado_remapeo(
                "⚠ Eso es un número de Carrier, no de PCB. Verifica lo que escaneaste.", "red", beep=True
            )
            self.entry_pcb_remapeo.delete(0, tk.END)
            return

        if not es_formato_pcb_remapeo(lot_id, self.LONGITUD_PCB_REMAPEO):
            self._mostrar_estado_remapeo(
                f"⚠ El PCB debe tener {self.LONGITUD_PCB_REMAPEO} caracteres alfanuméricos (sin guiones).",
                "red", beep=True,
            )
            self.entry_pcb_remapeo.delete(0, tk.END)
            return

        if lot_id in self.remap_mapeo.values():
            self._mostrar_estado_remapeo(f"⚠ El PCB {lot_id} ya está asignado en este pallet.", "red", beep=True)
            self.entry_pcb_remapeo.delete(0, tk.END)
            return

        # === SIMBIOSIS: regla de negocio compartida entre remapeo y retrabajo ===
        # Una pieza solo puede remapearse 2 veces; a la 3a debe pasar primero por
        # retrabajo (modulo Escaneo) o se scrapea. Aqui se bloquea la 3a asignacion
        # a menos que ya haya pasado por retrabajo despues de su ultimo remapeo.
        veces_remapeado = self.db.contar_veces_remapeado(lot_id)
        veces_retrabajado = self.db.veces_retrabajado(lot_id)
        if veces_remapeado >= 2 and veces_retrabajado == 0:
            messagebox.showerror(
                "Límite de remapeo alcanzado",
                f"El PCB {lot_id} ya fue remapeado {veces_remapeado} veces y NUNCA ha pasado por retrabajo.\n\n"
                "Según la política, tras 2 remapeos la pieza debe pasar primero por el módulo "
                "de Escaneo/Retrabajo antes de remapearse de nuevo, o se debe scrapear.",
            )
            self.entry_pcb_remapeo.delete(0, tk.END)
            return
        elif veces_remapeado >= 2:
            # ya paso por retrabajo despues del ultimo remapeo -> se permite, pero se avisa
            self._mostrar_estado_remapeo(
                f"ℹ PCB {lot_id}: {veces_remapeado} remapeos previos, ya pasó por retrabajo. Continuando.",
                "#cc7a00",
            )

        ya_remapeado_antes = veces_remapeado > 0

        for cid in self.remap_carrier_ids:
            if self.remap_mapeo[cid] is None and cid not in self.remap_bloqueados:
                self.remap_mapeo[cid] = lot_id
                self.remap_orden_escaneos.append((cid, lot_id))
                if ya_remapeado_antes:
                    self.remap_reincidentes.add(cid)
                    self._mostrar_estado_remapeo(
                        f"ℹ PCB {lot_id} ya había sido remapeado antes (posición {cid}, en naranja).", "#cc7a00"
                    )
                self.tabla_remapeo.set(cid, "LOT ID", lot_id)
                self._dibujar_mapa_remapeo(resaltar=cid)
                break
        else:
            self._mostrar_estado_remapeo("⚠ No hay posiciones libres/desbloqueadas en este pallet.", "red", beep=True)

        self.entry_pcb_remapeo.delete(0, tk.END)

        if self.remap_carrier_ids and all(self.remap_mapeo[c] is not None for c in self.remap_carrier_ids):
            self.after(700, self._cerrar_pallet_automatico_remapeo)

    def _deshacer_ultimo_remapeo(self):
        if not self.remap_orden_escaneos:
            self._mostrar_estado_remapeo("No hay ningún escaneo de PCB para deshacer.", "red", beep=True)
            return
        cid, lot_id = self.remap_orden_escaneos.pop()
        self.remap_mapeo[cid] = None
        self.remap_reincidentes.discard(cid)
        self.tabla_remapeo.set(cid, "LOT ID", "")
        self._dibujar_mapa_remapeo()
        self._mostrar_estado_remapeo(f"↩ Deshecho: {lot_id} en {cid}.", "#555555")
        self.entry_pcb_remapeo.focus_set()

    def _copiar_portapapeles_remapeo(self):
        if not self.remap_mapeo:
            return
        texto = "\n".join(f"{cid}\t{lot}" for cid, lot in self.remap_mapeo.items() if lot)
        self.clipboard_clear()
        self.clipboard_append(texto)
        self.db.guardar_remapeo(self.remap_carrier_base, self.remap_mapeo)
        self._actualizar_contador_pallets_remapeo()
        self._mostrar_estado_remapeo("✔ Copiado al portapapeles y guardado en historial.", "green")
        self.entry_carrier_remapeo.focus_set()
        self._refrescar_historial()

    def _cerrar_pallet_automatico_remapeo(self):
        if not self.remap_carrier_ids or not all(self.remap_mapeo.get(c) for c in self.remap_carrier_ids):
            return
        texto = "\n".join(f"{cid}\t{lot}" for cid, lot in self.remap_mapeo.items() if lot)
        self.clipboard_clear()
        self.clipboard_append(texto)
        self.db.guardar_remapeo(self.remap_carrier_base, self.remap_mapeo)
        self._actualizar_contador_pallets_remapeo()
        self._mostrar_estado_remapeo("✔ Pallet completo (12/12): copiado y guardado automáticamente.", "green", ms=5000)
        self.remap_mapeo, self.remap_carrier_ids = {}, []
        self.remap_bloqueados.clear()
        self.remap_reincidentes.clear()
        self.remap_orden_escaneos.clear()
        for item in self.tabla_remapeo.get_children():
            self.tabla_remapeo.delete(item)
        self._dibujar_mapa_remapeo()
        self.entry_carrier_remapeo.focus_set()
        self._refrescar_historial()

    def _reiniciar_mapa_remapeo(self):
        ocupados = sum(1 for v in self.remap_mapeo.values() if v)
        if ocupados > 0:
            if not messagebox.askyesno(
                "Confirmar reinicio",
                f"Hay {ocupados} posición(es) ocupada(s) sin verificar que se hayan copiado/guardado.\n"
                "¿Seguro que quieres reiniciar el mapa?",
            ):
                return
        self.remap_mapeo, self.remap_carrier_ids = {}, []
        self.remap_bloqueados.clear()
        self.remap_reincidentes.clear()
        self.remap_orden_escaneos.clear()
        for item in self.tabla_remapeo.get_children():
            self.tabla_remapeo.delete(item)
        self._dibujar_mapa_remapeo()
        self.entry_carrier_remapeo.focus_set()

    def _ver_historial_remapeo(self):
        fecha, turno = obtener_turno()
        filas = self.db.remapeos_del_turno(fecha, turno)
        ventana = tk.Toplevel(self)
        ventana.title(f"Historial {fecha} - Turno {turno}")
        tabla_hist = ttk.Treeview(ventana, columns=("Carrier", "PCB"), show="headings")
        for col in ("Carrier", "PCB"):
            tabla_hist.heading(col, text=col)
        tabla_hist.pack(fill="both", expand=True)
        for carrier, lot in filas:
            tabla_hist.insert("", "end", values=(carrier, lot))
        ttk.Label(ventana, text=f"Turno {turno} ({fecha}): {len(filas)} piezas").pack(pady=5)

    def _buscar_historial_remapeo_ventana(self):
        ventana = tk.Toplevel(self)
        ventana.title("Buscar en historial de remapeos")

        frame_filtros = ttk.Frame(ventana, padding=8)
        frame_filtros.pack(fill="x")

        ttk.Label(frame_filtros, text="Carrier (parcial o completo):").grid(row=0, column=0, sticky="w")
        entry_buscar_carrier = ttk.Entry(frame_filtros, width=18)
        entry_buscar_carrier.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="Desde (YYYY-MM-DD):").grid(row=0, column=2, sticky="w")
        entry_desde = ttk.Entry(frame_filtros, width=12)
        entry_desde.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="Hasta (YYYY-MM-DD):").grid(row=0, column=4, sticky="w")
        entry_hasta = ttk.Entry(frame_filtros, width=12)
        entry_hasta.grid(row=0, column=5, padx=5)

        ttk.Label(frame_filtros, text="Estado:").grid(row=1, column=0, sticky="w", pady=(6, 0))
        combo_estado = ttk.Combobox(frame_filtros, values=["Todos", "Pendientes", "Liberados"], width=12, state="readonly")
        combo_estado.set("Todos")
        combo_estado.grid(row=1, column=1, pady=(6, 0), sticky="w")

        lbl_resultado_info = ttk.Label(ventana, text="")
        lbl_resultado_info.pack(pady=(4, 0))

        tabla_res = ttk.Treeview(
            ventana, columns=("Fecha", "Turno", "Carrier", "PCB", "Estado"), show="headings", height=15
        )
        for col in ("Fecha", "Turno", "Carrier", "PCB", "Estado"):
            tabla_res.heading(col, text=col)
        tabla_res.pack(fill="both", expand=True, padx=8, pady=8)
        tabla_res.config(selectmode="extended")

        def ejecutar_busqueda():
            for item in tabla_res.get_children():
                tabla_res.delete(item)
            filtro_carrier = entry_buscar_carrier.get().strip()
            desde_txt = entry_desde.get().strip()
            hasta_txt = entry_hasta.get().strip()
            estado_filtro = combo_estado.get()
            estado_param = None if estado_filtro == "Todos" else estado_filtro

            try:
                desde = _dt.date.fromisoformat(desde_txt) if desde_txt else None
                hasta = _dt.date.fromisoformat(hasta_txt) if hasta_txt else None
            except ValueError:
                messagebox.showerror("Error", "Formato de fecha inválido. Usa YYYY-MM-DD.")
                return

            filas = self.db.buscar_remapeos(filtro_carrier, desde, hasta, estado_param)
            for fecha, turno, hora, carrier, lot, estado_txt in filas:
                tabla_res.insert("", "end", values=(fecha, turno, carrier, lot, estado_txt))
            lbl_resultado_info.config(text=f"{len(filas)} resultado(s) encontrado(s).")

        def marcar_seleccion_liberado():
            seleccion = tabla_res.selection()
            if not seleccion:
                messagebox.showinfo("Liberar", "Selecciona una o más filas para marcar como liberadas.")
                return
            n = 0
            for iid in seleccion:
                _fecha, _turno, carrier, lot, _estado = tabla_res.item(iid, "values")
                if self.db.marcar_liberado(carrier, lot):
                    n += 1
            ejecutar_busqueda()
            if n:
                self._mostrar_estado_remapeo(f"✔ {n} registro(s) marcado(s) como liberado.", "green")
            else:
                self._mostrar_estado_remapeo("Los registros seleccionados ya estaban liberados.", "#555555")

        frame_btns_busq = ttk.Frame(ventana)
        frame_btns_busq.pack(pady=(0, 8))
        ttk.Button(frame_btns_busq, text="Buscar", command=ejecutar_busqueda).grid(row=0, column=0, padx=5)
        ttk.Button(
            frame_btns_busq, text="Marcar seleccionados como liberado", command=marcar_seleccion_liberado
        ).grid(row=0, column=1, padx=5)

        ejecutar_busqueda()

    def _ventana_carriers_escaneados(self):
        def base_de(carrier):
            partes = carrier.split("-")
            return "-".join(partes[:-1]) if len(partes) > 1 else carrier

        def linea_de(base):
            b = base.upper()
            if b.startswith("LG1"):
                return "Línea 1"
            elif b.startswith("LG2"):
                return "Línea 2"
            return "Otra"

        ventana = tk.Toplevel(self)
        ventana.title("Carriers escaneados (Línea 1 / Línea 2)")

        frame_top = ttk.Frame(ventana, padding=8)
        frame_top.pack(fill="x")
        ttk.Label(frame_top, text="Buscar número de Carrier:").pack(side="left")
        entry_filtro_carrier = ttk.Entry(frame_top, width=20)
        entry_filtro_carrier.pack(side="left", padx=5)

        frame_tablas = ttk.Frame(ventana, padding=(8, 0))
        frame_tablas.pack(fill="both", expand=True)

        ttk.Label(frame_tablas, text="Carriers escaneados (agrupados por pallet)").pack(anchor="w")
        tabla_carriers = ttk.Treeview(
            frame_tablas, columns=("Linea", "Carrier", "Remapeos", "Ultima fecha"), show="headings", height=10
        )
        for col in ("Linea", "Carrier", "Remapeos", "Ultima fecha"):
            tabla_carriers.heading(col, text=col)
        tabla_carriers.pack(fill="both", expand=True, pady=(2, 10))

        ttk.Label(frame_tablas, text="Detalle: todas las veces que se remapeó el carrier seleccionado").pack(anchor="w")
        tabla_detalle = ttk.Treeview(
            frame_tablas, columns=("Fecha", "Turno", "Hora", "Posicion", "LOT ID"), show="headings", height=10
        )
        for col in ("Fecha", "Turno", "Hora", "Posicion", "LOT ID"):
            tabla_detalle.heading(col, text=col)
        tabla_detalle.pack(fill="both", expand=True, pady=(2, 8))

        lbl_info = ttk.Label(ventana, text="")
        lbl_info.pack(pady=(0, 4))

        def refrescar_carriers():
            for item in tabla_carriers.get_children():
                tabla_carriers.delete(item)
            for item in tabla_detalle.get_children():
                tabla_detalle.delete(item)
            filtro = entry_filtro_carrier.get().strip()
            agrupado = self.db.listar_carriers_agrupados(filtro)
            for base in sorted(agrupado.keys()):
                regs = agrupado[base]
                ultima = max(regs, key=lambda r: (r[0], r[2]))[0]
                tabla_carriers.insert("", "end", iid=base, values=(linea_de(base), base, len(regs), ultima))

        def mostrar_detalle(event=None):
            for item in tabla_detalle.get_children():
                tabla_detalle.delete(item)
            seleccion = tabla_carriers.selection()
            if not seleccion:
                return
            base = seleccion[0]
            agrupado = self.db.listar_carriers_agrupados()
            detalle = sorted(agrupado.get(base, []), key=lambda r: (r[0], r[2]))
            for fecha, turno, hora, carrier, lot in detalle:
                tabla_detalle.insert("", "end", values=(fecha, turno, hora, carrier, lot))
            lbl_info.config(text=f"{base}: {len(detalle)} remapeo(s) registrado(s) en total.")

        def copiar_lot_ids():
            seleccion = tabla_carriers.selection()
            if not seleccion:
                messagebox.showinfo("Copiar", "Selecciona primero un Carrier de la lista.")
                return
            base = seleccion[0]
            agrupado = self.db.listar_carriers_agrupados()
            lots = [r[4] for r in agrupado.get(base, [])]
            if not lots:
                return
            self.clipboard_clear()
            self.clipboard_append("\n".join(lots))
            self._mostrar_estado_remapeo(
                f"✔ {len(lots)} LOT ID del carrier {base} copiados al portapapeles (listos para el MES).", "green"
            )

        entry_filtro_carrier.bind("<KeyRelease>", lambda e: refrescar_carriers())
        tabla_carriers.bind("<<TreeviewSelect>>", mostrar_detalle)

        frame_btns = ttk.Frame(ventana)
        frame_btns.pack(pady=(0, 10))
        ttk.Button(frame_btns, text="Copiar LOT ID del carrier seleccionado", command=copiar_lot_ids).pack()

        refrescar_carriers()

    def _mostrar_graficas_remapeo(self):
        if not MATPLOTLIB_DISPONIBLE:
            messagebox.showwarning("Gráficas", "matplotlib no está instalado.")
            return
        metricas = self.db.metricas_remapeo_por_turno()
        if not metricas:
            messagebox.showwarning("Historial", "No hay historial registrado aún.")
            return
        conteo_turnos = {1: 0, 2: 0, 3: 0}
        for turnos in metricas.values():
            for t, n in turnos.items():
                conteo_turnos[t] += n
        plt.figure(figsize=(8, 4))
        plt.bar(conteo_turnos.keys(), conteo_turnos.values(), color=["blue", "green", "red"])
        plt.xticks([1, 2, 3], ["Turno 1", "Turno 2", "Turno 3"])
        plt.title("Remapeos por turno")
        plt.show()
        plt.figure(figsize=(6, 6))
        plt.pie(conteo_turnos.values(), labels=["Turno 1", "Turno 2", "Turno 3"], autopct="%1.1f%%")
        plt.title("Distribución de remapeos")
        plt.show()

    # =========================================================================
    # Modulo de Inventario por Racks (gestion de almacen)
    # =========================================================================

    def _construir_tab_inventario(self):
        top = ttk.Frame(self.tab_inventario, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Rack:").grid(row=0, column=0, sticky="w")
        self.combo_rack = ttk.Combobox(top, width=25, state="readonly")
        self.combo_rack.grid(row=0, column=1, padx=5)
        self.combo_rack.bind("<<ComboboxSelected>>", self._on_seleccionar_rack)

        ttk.Button(top, text="Nuevo rack", command=self._dialogo_nuevo_rack).grid(row=0, column=2, padx=5)

        ttk.Label(top, text="Tray actual:").grid(row=0, column=3, sticky="w", padx=(15, 0))
        self.combo_tray_actual = ttk.Combobox(top, width=6, state="readonly")
        self.combo_tray_actual.grid(row=0, column=4, padx=5)

        entry_frame = ttk.Frame(self.tab_inventario, padding=(10, 0))
        entry_frame.pack(fill="x")
        ttk.Label(entry_frame, text="Escanear Lot ID:").pack(side="left")
        self.entry_lot_inventario = ttk.Entry(entry_frame, width=35)
        self.entry_lot_inventario.pack(side="left", padx=6)
        self.entry_lot_inventario.bind("<Return>", lambda e: self._escanear_item_inventario())
        ttk.Button(entry_frame, text="Escanear", command=self._escanear_item_inventario).pack(side="left", padx=4)
        ttk.Button(entry_frame, text="Deshacer último", command=self._deshacer_item_inventario).pack(
            side="left", padx=4
        )

        self.lbl_estado_inventario = ttk.Label(self.tab_inventario, text="", font=("Segoe UI", 9, "bold"))
        self.lbl_estado_inventario.pack(anchor="w", padx=10, pady=(4, 0))

        # resumen de QTY por tray
        resumen_frame = ttk.LabelFrame(self.tab_inventario, text="Resumen del rack (QTY por tray)")
        resumen_frame.pack(fill="both", expand=True, padx=10, pady=10)
        cols = ("tray", "qty")
        self.tree_resumen_racks = ttk.Treeview(resumen_frame, columns=cols, show="headings", height=10)
        self.tree_resumen_racks.heading("tray", text="# Tray")
        self.tree_resumen_racks.heading("qty", text="QTY")
        self.tree_resumen_racks.column("tray", width=100)
        self.tree_resumen_racks.column("qty", width=100)
        self.tree_resumen_racks.pack(fill="both", expand=True, padx=8, pady=8)

        btn_export_frame = ttk.Frame(self.tab_inventario)
        btn_export_frame.pack(pady=(0, 10))
        ttk.Button(
            btn_export_frame, text="Exportar ESTE rack a Excel", command=self._exportar_rack_actual_excel
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_export_frame, text="Exportar TODOS los racks a Excel", command=self._exportar_todos_racks_excel
        ).pack(side="left", padx=5)

    def _refrescar_combo_racks(self):
        racks = self.db.listar_racks()
        valores = [f"{rid} - {nombre} ({nt} trays)" for rid, nombre, nt in racks]
        if self.combo_rack is not None:
            self.combo_rack["values"] = valores
            if valores and not self.combo_rack.get():
                self.combo_rack.current(0)
                self._on_seleccionar_rack()

    def _dialogo_nuevo_rack(self):
        ventana = tk.Toplevel(self)
        ventana.title("Nuevo rack")
        ventana.geometry("300x220")

        siguiente_auto = self.db.siguiente_nombre_rack_automatico()
        auto_var = tk.BooleanVar(value=True)
        lbl_preview = ttk.Label(ventana, text=f"Se nombrará: {siguiente_auto}", foreground="#2a7a2a")

        ttk.Label(ventana, text="Nombre del rack:").pack(pady=(10, 0))
        entry_nombre = ttk.Entry(ventana)
        entry_nombre.pack(pady=4)
        entry_nombre.config(state="disabled")

        def on_toggle_auto():
            if auto_var.get():
                entry_nombre.config(state="disabled")
                lbl_preview.config(text=f"Se nombrará: {self.db.siguiente_nombre_rack_automatico()}")
                lbl_preview.pack(pady=(0, 4))
            else:
                entry_nombre.config(state="normal")
                lbl_preview.pack_forget()

        ttk.Checkbutton(
            ventana, text="Nombrar automáticamente (R1, R2, R3...)", variable=auto_var, command=on_toggle_auto
        ).pack(pady=(4, 0))
        lbl_preview.pack(pady=(0, 4))

        ttk.Label(ventana, text="Número de trays en este rack:").pack(pady=(10, 0))
        entry_trays = ttk.Entry(ventana, width=6)
        entry_trays.insert(0, "6")
        entry_trays.pack(pady=4)

        def confirmar():
            if auto_var.get():
                nombre = self.db.siguiente_nombre_rack_automatico()
            else:
                nombre = entry_nombre.get().strip()
                if not nombre:
                    messagebox.showwarning("Falta nombre", "Ingresa un nombre para el rack.")
                    return
            try:
                num_trays = int(entry_trays.get())
                if num_trays < 1 or num_trays > 50:
                    messagebox.showwarning("Inválido", "El número de trays debe estar entre 1 y 50.")
                    return
            except ValueError:
                messagebox.showwarning("Inválido", "El número de trays debe ser un número.")
                return
            try:
                nuevo_id = self.db.crear_rack(nombre, num_trays)
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Ya existe un rack con ese nombre.")
                return
            ventana.destroy()
            self._refrescar_combo_racks()
            self.combo_rack.set(f"{nuevo_id} - {nombre} ({num_trays} trays)")
            self._on_seleccionar_rack()

        ttk.Button(ventana, text="Crear", command=confirmar).pack(pady=15)

    def _on_seleccionar_rack(self, event=None):
        seleccion = self.combo_rack.get()
        if not seleccion:
            return
        rack_id = int(seleccion.split(" - ")[0])
        self.rack_actual_id = rack_id
        row = self.db.obtener_rack(rack_id)
        if not row:
            return
        _id, _nombre, num_trays = row
        self.combo_tray_actual["values"] = [str(t) for t in range(1, num_trays + 1)]
        self.combo_tray_actual.current(0)
        self._refrescar_resumen_racks()

    def _refrescar_resumen_racks(self):
        for item in self.tree_resumen_racks.get_children():
            self.tree_resumen_racks.delete(item)
        if self.rack_actual_id is None:
            return
        row = self.db.obtener_rack(self.rack_actual_id)
        if not row:
            return
        _id, _nombre, num_trays = row
        qtys = self.db.qty_por_tray(self.rack_actual_id)
        for t in range(1, num_trays + 1):
            self.tree_resumen_racks.insert("", "end", values=(t, qtys.get(t, 0)))

    def _escanear_item_inventario(self):
        if self.rack_actual_id is None:
            messagebox.showwarning("Sin rack", "Selecciona o crea un rack primero.")
            return
        tray_str = self.combo_tray_actual.get()
        if not tray_str:
            messagebox.showwarning("Sin tray", "Selecciona el tray actual.")
            return
        tray_numero = int(tray_str)

        lot_id = self.entry_lot_inventario.get().strip()
        self.entry_lot_inventario.delete(0, tk.END)
        if not lot_id:
            return

        # deteccion de duplicados en TODO el rack (misma logica que Escaneo)
        tray_existente = self.db.lot_id_ya_en_rack(self.rack_actual_id, lot_id)
        if tray_existente is not None:
            self.lbl_estado_inventario.config(
                text=f"⚠ {lot_id} ya está en este rack (tray {tray_existente}).", foreground="red"
            )
            self.bell()
            return

        self.db.escanear_item_inventario(self.rack_actual_id, tray_numero, lot_id)
        self.lbl_estado_inventario.config(text=f"✔ {lot_id} agregado al tray {tray_numero}.", foreground="green")
        self._refrescar_resumen_racks()
        self.entry_lot_inventario.focus_set()

    def _deshacer_item_inventario(self):
        if self.rack_actual_id is None or not self.combo_tray_actual.get():
            return
        tray_numero = int(self.combo_tray_actual.get())
        lot_id = self.db.deshacer_ultimo_item_inventario(self.rack_actual_id, tray_numero)
        if lot_id:
            self.lbl_estado_inventario.config(text=f"↩ Deshecho: {lot_id} (tray {tray_numero}).", foreground="#555555")
            self._refrescar_resumen_racks()
        else:
            self.lbl_estado_inventario.config(text="No hay nada que deshacer en este tray.", foreground="red")

    def _exportar_rack_actual_excel(self):
        if self.rack_actual_id is None:
            messagebox.showwarning("Sin rack", "Selecciona un rack primero.")
            return
        matriz = self.db.exportar_rack_a_matriz(self.rack_actual_id)
        if not matriz or matriz["total_qty"] == 0:
            messagebox.showinfo("Sin datos", "Este rack no tiene piezas escaneadas.")
            return
        path = self._escribir_excel_inventario([matriz])
        messagebox.showinfo("Exportado", f"Archivo guardado como:\n{path}")

    def _exportar_todos_racks_excel(self):
        matrices = [m for m in self.db.exportar_todos_los_racks_a_matriz() if m and m["total_qty"] > 0]
        if not matrices:
            messagebox.showinfo("Sin datos", "No hay racks con piezas escaneadas.")
            return
        path = self._escribir_excel_inventario(matrices)
        messagebox.showinfo("Exportado", f"Archivo guardado como:\n{path}")

    def _escribir_excel_inventario(self, matrices):
        """Escribe el Excel en el mismo formato de la hoja de referencia:
        fila 'NUMERO DE RACK', fila '# TRAY', fila 'QTY' (resaltada), y debajo
        la lista de Lot ID de cada tray, un rack tras otro en columnas."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        relleno_qty = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        negrita = Font(bold=True)

        ws.cell(row=1, column=1, value="TOTAL QTY").font = negrita
        ws.cell(row=2, column=1, value="").font = negrita
        ws.cell(row=3, column=1, value="").font = negrita

        col = 2  # columna B en adelante (columna A es TOTAL QTY del primer rack)
        max_filas_items = 0
        primer_rack = True

        for matriz in matrices:
            nombre_rack = matriz["nombre"]
            total_qty = matriz["total_qty"]
            if primer_rack:
                ws.cell(row=4, column=1, value=total_qty)
                primer_rack = False

            for tray in matriz["trays"]:
                ws.cell(row=1, column=col, value=nombre_rack).font = negrita
                ws.cell(row=2, column=col, value=f"{nombre_rack}-{tray['tray_numero']}").font = negrita
                celda_qty = ws.cell(row=3, column=col, value=tray["qty"])
                celda_qty.font = negrita
                celda_qty.fill = relleno_qty
                for i, lot_id in enumerate(tray["items"]):
                    ws.cell(row=4 + i, column=col, value=lot_id)
                max_filas_items = max(max_filas_items, len(tray["items"]))
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
                col += 1

        ws.cell(row=1, column=1, value="NUMERO DE RACK").font = negrita
        ws.cell(row=2, column=1, value="# TRAY").font = negrita
        ws.cell(row=3, column=1, value="QTY").font = negrita
        ws.column_dimensions["A"].width = 16

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path_salida = f"inventario_racks_{timestamp}.xlsx"
        wb.save(path_salida)
        return path_salida

    def _construir_tab_escaneo(self):
        top = ttk.Frame(self.tab_escaneo)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="Charola:").grid(row=0, column=0, sticky="w")
        self.combo_charola = ttk.Combobox(top, width=30, state="readonly")
        self.combo_charola.grid(row=0, column=1, padx=5)
        self.combo_charola.bind("<<ComboboxSelected>>", self._on_seleccionar_charola)

        ttk.Button(top, text="Nueva charola", command=self._dialogo_nueva_charola).grid(row=0, column=2, padx=5)
        ttk.Checkbutton(top, text="Forzar orden estricto de escaneo", variable=self.orden_estricto).grid(
            row=0, column=3, padx=15
        )

        entry_frame = ttk.Frame(self.tab_escaneo)
        entry_frame.pack(fill="x", padx=10, pady=4)
        ttk.Label(entry_frame, text="Escanear Lot ID:").pack(side="left")
        self.entry_lot_id = ttk.Entry(entry_frame, width=35)
        self.entry_lot_id.pack(side="left", padx=6)
        self.entry_lot_id.bind("<Return>", lambda e: self._escanear())
        ttk.Button(entry_frame, text="Escanear", command=self._escanear).pack(side="left", padx=4)
        ttk.Button(entry_frame, text="Deshacer ultima celda", command=self._undo_celda).pack(side="left", padx=4)
        ttk.Button(entry_frame, text="Copiar todos los Lot ID", command=self._copiar_lot_ids).pack(
            side="left", padx=12
        )
        ttk.Button(entry_frame, text="Limpiar charola", command=self._limpiar_charola_dialogo).pack(
            side="left", padx=4
        )
        ttk.Button(
            entry_frame, text="Voltear charola completa (180°)", command=self._voltear_charola_180_dialogo
        ).pack(side="left", padx=4)

        self.lbl_progreso = ttk.Label(self.tab_escaneo, text="0 / 0 posiciones escaneadas")
        self.lbl_progreso.pack(anchor="w", padx=10)

        # opciones de bloqueo
        block_frame = ttk.Frame(self.tab_escaneo)
        block_frame.pack(fill="x", padx=10, pady=4)
        ttk.Label(block_frame, text="Bloquear posición:").pack(side="left")
        self.entry_posicion_bloqueo = ttk.Entry(block_frame, width=8)
        self.entry_posicion_bloqueo.pack(side="left", padx=4)
        ttk.Button(block_frame, text="Bloquear", command=self._bloquear_posicion).pack(side="left", padx=2)
        ttk.Button(block_frame, text="Desbloquear", command=self._desbloquear_posicion).pack(side="left", padx=2)
        ttk.Label(block_frame, text="(ej: A1, B6)", foreground="#666").pack(side="left", padx=4)

        self.frame_grid = ttk.Frame(self.tab_escaneo)
        self.frame_grid.pack(padx=10, pady=10)

    def _construir_tab_clasificacion(self):
        top = ttk.Frame(self.tab_clasificacion)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Button(top, text="Importar Excel del ERP (BATCH - todas charolas)", command=self._importar_excel).pack(side="left")
        ttk.Button(top, text="Exportar consolidado (Excel)", command=self._exportar_consolidado).pack(side="left", padx=10)

        ttk.Label(top, text="Ver por:").pack(side="left", padx=(20, 4))
        ttk.Radiobutton(
            top, text="Modelo", variable=self.vista_color, value="modelo", command=self._redibujar_grid
        ).pack(side="left")
        ttk.Radiobutton(
            top, text="Proceso", variable=self.vista_color, value="proceso", command=self._redibujar_grid
        ).pack(side="left")

        self.lbl_import_status = ttk.Label(self.tab_clasificacion, text="Sin archivo importado todavia.")
        self.lbl_import_status.pack(anchor="w", padx=10)

        # area de pega masiva
        pega_frame = ttk.LabelFrame(self.tab_clasificacion, text="Pegar datos masivos (Lot ID + Proceso +Modelo)")
        pega_frame.pack(fill="x", padx=10, pady=(8, 4))
        ttk.Label(
            pega_frame,
            text="Pega los datos aqui (tab/espacio separados: Lot_ID Modelo Proceso). Una linea por pieza.",
            font=("Segoe UI", 8), foreground="#666",
        ).pack(anchor="w", padx=8, pady=(4, 2))

        self.text_pega = tk.Text(pega_frame, height=6, width=80, font=("Courier New", 9))
        self.text_pega.pack(fill="x", padx=8, pady=(0, 4))

        pega_btn_frame = ttk.Frame(pega_frame)
        pega_btn_frame.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(pega_btn_frame, text="Mapear desde portapapeles", command=self._mapear_desde_texto).pack(
            side="left", padx=2
        )
        ttk.Button(pega_btn_frame, text="Limpiar texto", command=lambda: self.text_pega.delete("1.0", tk.END)).pack(
            side="left", padx=2
        )

        cuerpo = ttk.Frame(self.tab_clasificacion)
        cuerpo.pack(fill="both", expand=True, padx=10, pady=10)

        self.frame_leyenda = ttk.LabelFrame(cuerpo, text="Leyenda")
        self.frame_leyenda.pack(side="right", fill="y", padx=(10, 0))

    def _construir_tab_separacion(self):
        top = ttk.Frame(self.tab_separacion)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="Charola:").pack(side="left")
        self.combo_charola_sep = ttk.Combobox(top, width=30, state="readonly")
        self.combo_charola_sep.pack(side="left", padx=5)
        self.combo_charola_sep.bind("<<ComboboxSelected>>", lambda e: self._refrescar_picking_list())
        ttk.Button(top, text="Actualizar", command=self._refrescar_picking_list).pack(side="left", padx=5)

        self.usar_modelo_base = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            top, text="Agrupar por modelo base (usa tabla de equivalencias)",
            variable=self.usar_modelo_base, command=self._refrescar_picking_list,
        ).pack(side="left", padx=15)

        # --- Vista grafica tipo mapa (igual que Clasificacion, pero con esta charola) ---
        barra_vista = ttk.Frame(self.tab_separacion)
        barra_vista.pack(fill="x", padx=10, pady=(4, 0))
        ttk.Label(barra_vista, text="Mapa de charola - ver por:").pack(side="left")
        self.vista_color_sep = tk.StringVar(value="modelo")
        ttk.Radiobutton(
            barra_vista, text="Modelo", variable=self.vista_color_sep, value="modelo",
            command=self._dibujar_grid_separacion,
        ).pack(side="left")
        ttk.Radiobutton(
            barra_vista, text="Proceso", variable=self.vista_color_sep, value="proceso",
            command=self._dibujar_grid_separacion,
        ).pack(side="left", padx=6)
        ttk.Button(barra_vista, text="Copiar seleccionados", command=self._copiar_seleccionados_sep).pack(
            side="left", padx=15
        )

        panel_mapa = ttk.Frame(self.tab_separacion)
        panel_mapa.pack(fill="x", padx=10, pady=6)
        self.frame_grid_separacion = ttk.Frame(panel_mapa)
        self.frame_grid_separacion.pack(side="left")
        self.frame_leyenda_sep = ttk.LabelFrame(panel_mapa, text="Leyenda")
        self.frame_leyenda_sep.pack(side="left", padx=15, fill="y")

        ttk.Label(
            self.tab_separacion,
            text=(
                "Marca cada pieza conforme la muevas fisicamente a su charola de destino por modelo. "
                "El check se guarda al instante."
            ),
            foreground="#555555",
        ).pack(anchor="w", padx=10, pady=(0, 6))

        # Area con scroll para las listas de picking por modelo
        contenedor = ttk.Frame(self.tab_separacion)
        contenedor.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        canvas = tk.Canvas(contenedor, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(contenedor, orient="vertical", command=canvas.yview)
        self.frame_picking = ttk.Frame(canvas)
        self.frame_picking.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.frame_picking, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.canvas_picking = canvas

    def _dibujar_grid_separacion(self):
        """Dibuja el mapa visual de charola en Separacion, mostrando modelo/proceso por celda.
        Las celdas son clickeables para seleccionar/deseleccionar para copiar.
        """
        for w in self.frame_grid_separacion.winfo_children():
            w.destroy()

        seleccion = self.combo_charola_sep.get()
        if not seleccion or self.charola_actual_id != int(seleccion.split(" - ")[0]):
            return

        charola_id = self.charola_actual_id
        filas, columnas = self.filas, self.columnas
        volteada = self.volteada_actual
        criterio = self.vista_color_sep.get()

        datos = self.db.pcbs_de_charola(charola_id)
        posiciones = generar_posiciones(filas, columnas)

        # generar mapas de color
        color_map = {}
        for pos, info in datos.items():
            valor = info.get(criterio)
            if valor and valor not in color_map:
                color_map[valor] = COLOR_PALETTE[len(color_map) % len(COLOR_PALETTE)]

        # encabezados columna
        for c in range(columnas):
            ttk.Label(self.frame_grid_separacion, text=etiqueta_columna(c), width=12, anchor="center").grid(
                row=0, column=c + 1
            )
        for f in range(filas):
            ttk.Label(self.frame_grid_separacion, text=str(f + 1), width=4, anchor="center").grid(
                row=f + 1, column=0
            )

        # celdas (clickeables)
        self.celdas_sep = {}
        for pos in posiciones:
            f = int("".join(ch for ch in pos if ch.isdigit())) - 1
            col_letra = "".join(ch for ch in pos if ch.isalpha())
            c = _col_letra_a_indice(col_letra)
            info = datos.get(pos)

            if not info:
                color = COLOR_PENDIENTE
                texto = pos
                fg = "black"
            else:
                valor = info.get(criterio)
                prefijo = "✓ " if info.get("separada") else ""
                texto = prefijo + (valor if valor else "N/D")
                color = color_map.get(valor, COLOR_SIN_MATCH) if valor else COLOR_ESCANEADO
                fg = "black"

            # resaltado si está seleccionada para copiar
            relief = "sunken" if pos in self.celdas_sep_seleccionadas else "ridge"
            width_label = 14 if pos not in self.celdas_sep_seleccionadas else 14
            bd = 3 if pos in self.celdas_sep_seleccionadas else 1

            lbl = tk.Label(
                self.frame_grid_separacion, text=texto, width=width_label, height=2,
                relief=relief, bg=color, fg=fg, wraplength=100, font=("Segoe UI", 8),
                bd=bd, cursor="hand2",
            )
            lbl.grid(row=f + 1, column=c + 1, padx=2, pady=2)
            lbl.bind("<Button-1>", lambda e, pos=pos: self._toggle_celda_sep(pos))
            self.celdas_sep[pos] = lbl

        # leyenda (clickeable también)
        self._dibujar_leyenda_separacion(color_map)

    def _dibujar_leyenda_separacion(self, mapa_color):
        for w in self.frame_leyenda_sep.winfo_children():
            w.destroy()
        if not mapa_color:
            ttk.Label(self.frame_leyenda_sep, text="(sin datos importados)").pack(padx=8, pady=8)
            return
        
        ttk.Label(self.frame_leyenda_sep, text="Haz click para seleccionar/deseleccionar", 
                  font=("Segoe UI", 7), foreground="#666").pack(padx=4, pady=2)
        
        for nombre, color in mapa_color.items():
            fila = tk.Frame(self.frame_leyenda_sep, bg="white", relief="raised", bd=1, cursor="hand2")
            fila.pack(fill="x", padx=4, pady=2, ipady=2)
            fila.bind("<Button-1>", lambda e, cat=nombre: self._toggle_color_sep(cat, mapa_color))
            
            tk.Label(fila, bg=color, width=2, height=1).pack(side="left", padx=2)
            tk.Label(fila, text=nombre, bg="white", fg="black", cursor="hand2").pack(side="left", padx=6, fill="x", expand=True)
            fila.bind("<Button-1>", lambda e, cat=nombre: self._toggle_color_sep(cat, mapa_color))

    def _toggle_celda_sep(self, posicion):
        """Marca/desmarca una celda individual para copiar."""
        if posicion in self.celdas_sep_seleccionadas:
            self.celdas_sep_seleccionadas.discard(posicion)
        else:
            self.celdas_sep_seleccionadas.add(posicion)
        self._dibujar_grid_separacion()

    def _toggle_color_sep(self, categoria, mapa_color):
        """Selecciona/deselecciona todas las celdas de un color (modelo/proceso)."""
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        criterio = self.vista_color_sep.get()
        
        # encuentra todas las posiciones que tienen este color
        posiciones_color = [pos for pos, info in datos.items() if info.get(criterio) == categoria]
        
        # si todas ya están seleccionadas, deselecciona; si no, selecciona todas
        si_todas_marcadas = all(pos in self.celdas_sep_seleccionadas for pos in posiciones_color)
        
        if si_todas_marcadas:
            for pos in posiciones_color:
                self.celdas_sep_seleccionadas.discard(pos)
        else:
            for pos in posiciones_color:
                self.celdas_sep_seleccionadas.add(pos)
        
        self._dibujar_grid_separacion()

    def _copiar_seleccionados_sep(self):
        """Copia los Lot ID de las celdas seleccionadas al portapapeles."""
        if not self.celdas_sep_seleccionadas:
            messagebox.showinfo("Sin selección", "Selecciona al menos una pieza haciendo click en el mapa.")
            return
        
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        lot_ids = [datos[pos]["lot_id"] for pos in self.celdas_sep_seleccionadas if pos in datos]
        
        if not lot_ids:
            messagebox.showinfo("Sin datos", "Las posiciones seleccionadas no tienen Lot ID.")
            return
        
        self.clipboard_clear()
        self.clipboard_append("\n".join(lot_ids))
        messagebox.showinfo("Copiado", f"{len(lot_ids)} Lot ID copiados al portapapeles.")

    def _refrescar_picking_list(self):
        for w in self.frame_picking.winfo_children():
            w.destroy()

        seleccion = self.combo_charola_sep.get()
        if not seleccion:
            ttk.Label(self.frame_picking, text="Selecciona una charola arriba.").pack(padx=10, pady=10)
            return

        charola_id = int(seleccion.split(" - ")[0])
        self.charola_actual_id = charola_id
        row = self.db.obtener_charola(charola_id)
        if row:
            _, _, self.filas, self.columnas, self.volteada_actual, _auto_v, _v_aplicado = row
        grupos = self.db.pcbs_agrupados_por_modelo(charola_id, usar_modelo_base=self.usar_modelo_base.get())

        if not grupos:
            ttk.Label(
                self.frame_picking,
                text="Esta charola aun no tiene piezas clasificadas.\n"
                     "Importa el Excel del ERP en la pestana 'Clasificacion / Excel' primero.",
                justify="left",
            ).pack(padx=10, pady=10, anchor="w")
            return

        for modelo, piezas in grupos.items():
            total = len(piezas)
            separadas = sum(1 for p in piezas if p["separada"])
            titulo = f"Modelo: {modelo}   ({separadas}/{total} separadas)"
            grupo_frame = ttk.LabelFrame(self.frame_picking, text=titulo)
            grupo_frame.pack(fill="x", padx=6, pady=6, anchor="w")

            barra = ttk.Frame(grupo_frame)
            barra.pack(fill="x", pady=(2, 6))
            ttk.Button(
                barra, text="Marcar todas (todo el modelo)",
                command=lambda cid=charola_id, m=modelo: self._marcar_grupo(cid, m, True),
            ).pack(side="left", padx=4)
            ttk.Button(
                barra, text="Desmarcar todas",
                command=lambda cid=charola_id, m=modelo: self._marcar_grupo(cid, m, False),
            ).pack(side="left", padx=4)
            ttk.Button(
                barra, text="Copiar Lot ID de todo el modelo",
                command=lambda piezas=piezas: self._copiar_grupo(piezas),
            ).pack(side="left", padx=12)

            # --- submenu: subdividir por proceso dentro de este modelo ---
            subgrupos_proceso = {}
            for pieza in piezas:
                clave_proceso = pieza["proceso"] if pieza["proceso"] else "(sin proceso)"
                subgrupos_proceso.setdefault(clave_proceso, []).append(pieza)

            for proceso, piezas_proceso in subgrupos_proceso.items():
                total_p = len(piezas_proceso)
                separadas_p = sum(1 for p in piezas_proceso if p["separada"])
                sub_frame = ttk.LabelFrame(
                    grupo_frame, text=f"Proceso: {proceso}   ({separadas_p}/{total_p} separadas)"
                )
                sub_frame.pack(fill="x", padx=16, pady=4, anchor="w")

                barra_sub = ttk.Frame(sub_frame)
                barra_sub.pack(fill="x", pady=(2, 4))
                ttk.Button(
                    barra_sub, text="Marcar todas de este proceso",
                    command=lambda cid=charola_id, m=modelo, p=proceso: self._marcar_subgrupo_proceso(cid, m, p, True),
                ).pack(side="left", padx=4)
                ttk.Button(
                    barra_sub, text="Desmarcar",
                    command=lambda cid=charola_id, m=modelo, p=proceso: self._marcar_subgrupo_proceso(cid, m, p, False),
                ).pack(side="left", padx=4)
                ttk.Button(
                    barra_sub, text="Copiar Lot ID de este proceso",
                    command=lambda piezas_proceso=piezas_proceso: self._copiar_grupo(piezas_proceso),
                ).pack(side="left", padx=8)

                for pieza in piezas_proceso:
                    var = tk.BooleanVar(value=pieza["separada"])
                    texto = f'{pieza["posicion"]}   {pieza["lot_id"]}'
                    ttk.Checkbutton(
                        sub_frame, text=texto, variable=var,
                        command=lambda cid=charola_id, pos=pieza["posicion"], var=var: self._toggle_separada(
                            cid, pos, var
                        ),
                    ).pack(anchor="w", padx=20, pady=1)

        # validaciones de error proofing
        total_charola = self.db.contar_totales_en_charola(charola_id)
        sin_clasificar = self.db.contar_sin_clasificar_en_charola(charola_id)
        if total_charola > 0 and sin_clasificar > 0:
            porcentaje_sin = (sin_clasificar / total_charola) * 100
            if porcentaje_sin > 50:
                messagebox.showwarning(
                    "Alerta: mucho sin clasificar",
                    f"Esta charola tiene {sin_clasificar}/{total_charola} piezas ({porcentaje_sin:.1f}%) sin clasificar.\n"
                    "Asegúrate de que el Excel del ERP se importó correctamente.",
                )
            elif porcentaje_sin > 10:
                self.db.registrar_historial("validacion_alerta", f"Charola {charola_id}: {porcentaje_sin:.1f}% sin clasificar")

        # redibuja el grid visual tras refrescar la lista de picking
        self._dibujar_grid_separacion()

    def _toggle_separada(self, charola_id, posicion, var):
        self.db.marcar_separada(charola_id, posicion, var.get())
        estado = "separada" if var.get() else "sin separar"
        self.db.registrar_historial("separacion", f"charola {charola_id} pos {posicion} -> {estado}")
        self._actualizar_titulos_picking()
        self._redibujar_grid()
        self._dibujar_grid_separacion()  # actualiza el grid visual de Separacion

    def _marcar_grupo(self, charola_id, modelo, valor):
        grupos = self.db.pcbs_agrupados_por_modelo(charola_id, usar_modelo_base=self.usar_modelo_base.get())
        for pieza in grupos.get(modelo, []):
            self.db.marcar_separada(charola_id, pieza["posicion"], valor)
        estado = "todas separadas" if valor else "todas sin separar"
        self.db.registrar_historial("separacion_grupo", f"charola {charola_id} modelo {modelo} -> {estado}")
        self._refrescar_picking_list()
        self._redibujar_grid()

    def _marcar_subgrupo_proceso(self, charola_id, modelo, proceso, valor):
        """Marca/desmarca solo las piezas de un modelo Y proceso especificos (submenu)."""
        grupos = self.db.pcbs_agrupados_por_modelo(charola_id, usar_modelo_base=self.usar_modelo_base.get())
        for pieza in grupos.get(modelo, []):
            clave_proceso = pieza["proceso"] if pieza["proceso"] else "(sin proceso)"
            if clave_proceso == proceso:
                self.db.marcar_separada(charola_id, pieza["posicion"], valor)
        estado = "separadas" if valor else "sin separar"
        self.db.registrar_historial(
            "separacion_subgrupo", f"charola {charola_id} modelo {modelo} proceso {proceso} -> {estado}"
        )
        self._refrescar_picking_list()
        self._redibujar_grid()

    def _actualizar_titulos_picking(self):
        # simplemente redibuja toda la lista para refrescar los contadores (N/N separadas)
        self._refrescar_picking_list()

    def _copiar_grupo(self, piezas):
        lot_ids = [p["lot_id"] for p in piezas]
        if not lot_ids:
            return
        self.clipboard_clear()
        self.clipboard_append("\n".join(lot_ids))
        messagebox.showinfo("Copiado", f"{len(lot_ids)} Lot ID de este modelo copiados al portapapeles.")

    def _construir_tab_modelos_base(self):
        sub_notebook = ttk.Notebook(self.tab_modelos_base)
        sub_notebook.pack(fill="both", expand=True)

        self.subtab_equivalencias = ttk.Frame(sub_notebook)
        self.subtab_catalogo = ttk.Frame(sub_notebook)
        sub_notebook.add(self.subtab_equivalencias, text="Equivalencias (modelo base)")
        sub_notebook.add(self.subtab_catalogo, text="Catálogo LGIT (referencia)")

        self._construir_subtab_equivalencias()
        self._construir_subtab_catalogo()

    def _construir_subtab_equivalencias(self):
        ttk.Label(
            self.subtab_equivalencias,
            text=(
                "Tabla de equivalencias: registra aqui como se reduce cada modelo variante a su "
                "modelo base al relocalizar en InstalPCB (ej. 6L10 -> 6L). Confirma la regla exacta "
                "con tu equipo antes de llenarla; mientras tanto, la separacion sigue funcionando "
                "con el modelo tal cual lo entrega el ERP."
            ),
            wraplength=780, justify="left", foreground="#555555",
        ).pack(anchor="w", padx=10, pady=10)

        form = ttk.Frame(self.subtab_equivalencias)
        form.pack(fill="x", padx=10, pady=4)
        ttk.Label(form, text="Modelo variante (ERP):").grid(row=0, column=0, sticky="w")
        self.entry_variante = ttk.Entry(form, width=20)
        self.entry_variante.grid(row=0, column=1, padx=6)
        ttk.Label(form, text="Modelo base:").grid(row=0, column=2, sticky="w", padx=(15, 0))
        self.entry_base = ttk.Entry(form, width=20)
        self.entry_base.grid(row=0, column=3, padx=6)
        ttk.Button(form, text="Agregar / actualizar", command=self._agregar_equivalencia).grid(
            row=0, column=4, padx=10
        )

        cols = ("variante", "base")
        self.tree_equivalencias = ttk.Treeview(self.subtab_equivalencias, columns=cols, show="headings", height=12)
        self.tree_equivalencias.heading("variante", text="Modelo variante")
        self.tree_equivalencias.heading("base", text="Modelo base")
        self.tree_equivalencias.column("variante", width=250)
        self.tree_equivalencias.column("base", width=250)
        self.tree_equivalencias.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Button(
            self.subtab_equivalencias, text="Eliminar seleccionada", command=self._eliminar_equivalencia
        ).pack(anchor="w", padx=10, pady=(0, 10))

        self._refrescar_equivalencias()

    def _construir_subtab_catalogo(self):
        ttk.Label(
            self.subtab_catalogo,
            text=(
                "Catálogo de referencia LGIT: consulta qué housing, grados y modelo base lleva cada "
                "código LGIT (ej. tablas 'Thunder Trinity' / 'Thunder Cheetah'). Escribe un código o "
                "nombre para buscar."
            ),
            wraplength=900, justify="left", foreground="#555555",
        ).pack(anchor="w", padx=10, pady=(10, 4))

        # barra de busqueda y carga de ejemplo
        barra_top = ttk.Frame(self.subtab_catalogo)
        barra_top.pack(fill="x", padx=10, pady=4)
        ttk.Label(barra_top, text="Buscar (código LGIT o nombre):").pack(side="left")
        self.entry_buscar_catalogo = ttk.Entry(barra_top, width=30)
        self.entry_buscar_catalogo.pack(side="left", padx=6)
        self.entry_buscar_catalogo.bind("<KeyRelease>", lambda e: self._refrescar_catalogo())
        ttk.Button(barra_top, text="Cargar catálogo Thunder (ejemplo)", command=self._cargar_catalogo_ejemplo).pack(
            side="left", padx=15
        )

        # tabla de resultados
        cols_cat = ("familia", "nombre", "lgit_model", "pcb_type", "degrees", "top_housing",
                    "housing_mess", "bottom_cover", "opal")
        self.tree_catalogo = ttk.Treeview(self.subtab_catalogo, columns=cols_cat, show="headings", height=14)
        encabezados = {
            "familia": "Familia", "nombre": "Nombre", "lgit_model": "LGIT Model",
            "pcb_type": "PCB Type", "degrees": "Degrees", "top_housing": "Top Housing",
            "housing_mess": "Housing (mess)", "bottom_cover": "Bottom Cover", "opal": "Opal",
        }
        anchos = {
            "familia": 100, "nombre": 170, "lgit_model": 130, "pcb_type": 90, "degrees": 60,
            "top_housing": 90, "housing_mess": 90, "bottom_cover": 90, "opal": 80,
        }
        for c in cols_cat:
            self.tree_catalogo.heading(c, text=encabezados[c])
            self.tree_catalogo.column(c, width=anchos[c])
        self.tree_catalogo.pack(fill="both", expand=True, padx=10, pady=(4, 6))

        ttk.Button(
            self.subtab_catalogo, text="Eliminar fila seleccionada", command=self._eliminar_catalogo_fila
        ).pack(anchor="w", padx=10, pady=(0, 8))

        # --- pegar/agregar filas manualmente ---
        form_cat = ttk.LabelFrame(self.subtab_catalogo, text="Agregar fila manualmente")
        form_cat.pack(fill="x", padx=10, pady=(0, 6))
        etiquetas_campos = [
            ("Familia", 0), ("Nombre", 1), ("LGIT Model", 2), ("PCB Type", 3), ("Degrees", 4),
            ("Top Housing", 5), ("Housing (mess)", 6), ("Bottom Cover", 7), ("Opal", 8),
        ]
        self.entries_catalogo_manual = {}
        for i, (etiqueta, col) in enumerate(etiquetas_campos):
            ttk.Label(form_cat, text=etiqueta + ":").grid(row=i // 3, column=(i % 3) * 2, sticky="w", padx=4, pady=2)
            entry = ttk.Entry(form_cat, width=16)
            entry.grid(row=i // 3, column=(i % 3) * 2 + 1, padx=4, pady=2)
            self.entries_catalogo_manual[etiqueta] = entry
        ttk.Button(form_cat, text="Agregar fila", command=self._agregar_catalogo_manual).grid(
            row=3, column=0, columnspan=6, pady=6
        )

        # --- pega masiva ---
        pega_cat_frame = ttk.LabelFrame(self.subtab_catalogo, text="Pegar tabla completa (una fila por línea, separado por tab)")
        pega_cat_frame.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Label(
            pega_cat_frame,
            text="Formato: Nombre  LGIT_Model  PCB_Type  Degrees  Top_Housing  [Housing_mess]  Bottom_Cover  [Opal]",
            font=("Segoe UI", 8), foreground="#666",
        ).pack(anchor="w", padx=6, pady=(4, 2))
        self.text_pega_catalogo = tk.Text(pega_cat_frame, height=4, font=("Courier New", 9))
        self.text_pega_catalogo.pack(fill="x", padx=6, pady=(0, 4))
        frame_btn_pega_cat = ttk.Frame(pega_cat_frame)
        frame_btn_pega_cat.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Label(frame_btn_pega_cat, text="Familia para estas filas:").pack(side="left")
        self.entry_familia_pega = ttk.Entry(frame_btn_pega_cat, width=20)
        self.entry_familia_pega.pack(side="left", padx=6)
        ttk.Button(frame_btn_pega_cat, text="Importar filas pegadas", command=self._importar_catalogo_pegado).pack(
            side="left", padx=10
        )

        self._refrescar_catalogo()

    def _refrescar_catalogo(self):
        for item in self.tree_catalogo.get_children():
            self.tree_catalogo.delete(item)
        texto_busqueda = self.entry_buscar_catalogo.get().strip() if self.entry_buscar_catalogo else ""
        if texto_busqueda:
            filas = self.db.buscar_catalogo_modelo(texto_busqueda)
        else:
            filas = self.db.listar_catalogo_modelos()
        for (cid, familia, nombre, lgit, pcb_type, deg, top_h, mess, bottom, opal) in filas:
            self.tree_catalogo.insert(
                "", "end", iid=str(cid),
                values=(familia, nombre, lgit, pcb_type, deg, top_h, mess, bottom, opal),
            )

    def _agregar_catalogo_manual(self):
        e = self.entries_catalogo_manual
        familia = e["Familia"].get().strip()
        nombre = e["Nombre"].get().strip()
        lgit = e["LGIT Model"].get().strip()
        if not lgit:
            messagebox.showwarning("Falta LGIT Model", "El código LGIT Model es obligatorio.")
            return
        self.db.agregar_catalogo_modelo(
            familia, nombre, lgit,
            e["PCB Type"].get().strip(), e["Degrees"].get().strip(),
            e["Top Housing"].get().strip(), e["Housing (mess)"].get().strip(),
            e["Bottom Cover"].get().strip(), e["Opal"].get().strip(),
        )
        for entry in e.values():
            entry.delete(0, tk.END)
        self._refrescar_catalogo()

    def _eliminar_catalogo_fila(self):
        seleccion = self.tree_catalogo.selection()
        if not seleccion:
            return
        for iid in seleccion:
            self.db.eliminar_catalogo_modelo(int(iid))
        self._refrescar_catalogo()

    def _importar_catalogo_pegado(self):
        texto = self.text_pega_catalogo.get("1.0", tk.END)
        if not texto.strip():
            messagebox.showwarning("Vacío", "Pega datos primero.")
            return
        familia = self.entry_familia_pega.get().strip()
        n = self.db.importar_catalogo_desde_texto(texto, familia)
        messagebox.showinfo("Importado", f"{n} fila(s) agregadas al catálogo.")
        self.text_pega_catalogo.delete("1.0", tk.END)
        self._refrescar_catalogo()

    def _cargar_catalogo_ejemplo(self):
        n = self.db.cargar_catalogo_ejemplo_thunder()
        if n == 0:
            messagebox.showinfo(
                "Catálogo ya cargado",
                "El catálogo ya tiene datos (no se sobreescribe). Si quieres recargar el ejemplo, "
                "elimina primero las filas existentes.",
            )
        else:
            messagebox.showinfo(
                "Catálogo cargado",
                f"{n} filas cargadas (Thunder Trinity + Thunder Cheetah).\n\n"
                "⚠ Un par de celdas estaban borrosas/con reflejo en la foto original "
                "(marcadas como '[VERIFICAR CODIGO]') — revísalas y corrígelas manualmente.",
            )
        self._refrescar_catalogo()

    def _agregar_equivalencia(self):
        variante = self.entry_variante.get().strip()
        base = self.entry_base.get().strip()
        if not variante or not base:
            messagebox.showwarning("Datos incompletos", "Ingresa el modelo variante y su modelo base.")
            return
        self.db.agregar_equivalencia(variante, base)
        self.entry_variante.delete(0, tk.END)
        self.entry_base.delete(0, tk.END)
        self._refrescar_equivalencias()

    def _eliminar_equivalencia(self):
        seleccion = self.tree_equivalencias.selection()
        if not seleccion:
            return
        variante = self.tree_equivalencias.item(seleccion[0], "values")[0]
        self.db.eliminar_equivalencia(variante)
        self._refrescar_equivalencias()

    def _refrescar_equivalencias(self):
        for item in self.tree_equivalencias.get_children():
            self.tree_equivalencias.delete(item)
        for variante, base in self.db.listar_equivalencias():
            self.tree_equivalencias.insert("", "end", values=(variante, base))

    def _construir_tab_historial(self):
        ttk.Label(
            self.tab_historial,
            text=f"Base de datos: {DB_PATH}",
            font=("Segoe UI", 8), foreground="#555555",
        ).pack(anchor="w", padx=10, pady=(8, 0))
        cols = ("evento", "detalle", "fecha")
        self.tree_historial = ttk.Treeview(self.tab_historial, columns=cols, show="headings")
        for c, w in zip(cols, (150, 500, 160)):
            self.tree_historial.heading(c, text=c.capitalize())
            self.tree_historial.column(c, width=w)
        self.tree_historial.pack(fill="both", expand=True, padx=10, pady=10)
        ttk.Button(self.tab_historial, text="Actualizar", command=self._refrescar_historial).pack(pady=4)
        self._refrescar_historial()

    # ---------------- Charolas ----------------
    def _dialogo_nueva_charola(self):
        ventana = tk.Toplevel(self)
        ventana.title("Nueva charola")
        ventana.geometry("360x400")

        siguiente_auto = self.db.siguiente_nombre_automatico()

        auto_nombre_var = tk.BooleanVar(value=True)
        lbl_preview = ttk.Label(ventana, text=f"Se nombrará: {siguiente_auto}", foreground="#2a7a2a")

        ttk.Label(ventana, text="Nombre / ID de charola:").pack(pady=(10, 0))
        entry_nombre = ttk.Entry(ventana)
        entry_nombre.pack(pady=4)
        entry_nombre.config(state="disabled")

        def on_toggle_auto():
            if auto_nombre_var.get():
                entry_nombre.config(state="disabled")
                lbl_preview.config(text=f"Se nombrará: {self.db.siguiente_nombre_automatico()}")
                lbl_preview.pack(pady=(0, 4))
            else:
                entry_nombre.config(state="normal")
                lbl_preview.pack_forget()

        chk_auto = ttk.Checkbutton(
            ventana, text="Nombrar automáticamente (A1, A2, A3... B1, B2...)",
            variable=auto_nombre_var, command=on_toggle_auto,
        )
        chk_auto.pack(pady=(4, 0))
        lbl_preview.pack(pady=(0, 4))

        ttk.Label(ventana, text="Tamano:").pack(pady=(10, 0))
        combo_tamano = ttk.Combobox(
            ventana, state="readonly",
            values=["48 (6x8)", "36 (6x6)", "Personalizado"],
        )
        combo_tamano.current(0)
        combo_tamano.pack(pady=4)

        volteada_var = tk.BooleanVar(value=False)
        chk_volteada = ttk.Checkbutton(
            ventana,
            text="Modo VOLTEADA en vivo (fila se invierte al escanear)",
            variable=volteada_var,
        )
        chk_volteada.pack(pady=(8, 0))

        auto_volteo_var = tk.BooleanVar(value=False)
        chk_auto_volteo = ttk.Checkbutton(
            ventana,
            text="Auto-voltear 180° al clasificar (recomendado en charolas 6x6)",
            variable=auto_volteo_var,
        )
        chk_auto_volteo.pack(pady=(4, 0))
        ttk.Label(
            ventana,
            text="Escaneas en el orden crudo tal cual ves la charola volteada;\n"
                 "al importar/pegar los datos de clasificación, se voltea sola.",
            font=("Segoe UI", 8), foreground="#666", justify="left",
        ).pack(pady=(0, 4))

        frame_custom = ttk.Frame(ventana)
        entry_f = ttk.Entry(frame_custom, width=5)
        entry_c = ttk.Entry(frame_custom, width=5)

        def on_tamano_change(event=None):
            seleccion = combo_tamano.get()
            if seleccion == "Personalizado":
                frame_custom.pack(pady=4)
                ttk.Label(frame_custom, text="Filas:").grid(row=0, column=0)
                entry_f.grid(row=0, column=1, padx=4)
                ttk.Label(frame_custom, text="Columnas:").grid(row=0, column=2)
                entry_c.grid(row=0, column=3, padx=4)
            else:
                frame_custom.pack_forget()
            # sugerencia automatica: 6x6 casi siempre necesita el auto-volteo de 180°
            auto_volteo_var.set(seleccion == "36 (6x6)")
            volteada_var.set(False)

        combo_tamano.bind("<<ComboboxSelected>>", on_tamano_change)

        def confirmar():
            if auto_nombre_var.get():
                nombre = self.db.siguiente_nombre_automatico()
            else:
                nombre = entry_nombre.get().strip()
                if not nombre:
                    messagebox.showwarning("Falta nombre", "Ingresa un nombre o ID para la charola.")
                    return
            seleccion = combo_tamano.get()
            if seleccion == "48 (6x8)":
                filas, columnas = 6, 8
            elif seleccion == "36 (6x6)":
                filas, columnas = 6, 6
            else:  # Personalizado
                try:
                    filas, columnas = int(entry_f.get()), int(entry_c.get())
                    if filas < 1 or columnas < 1 or filas > 20 or columnas > 20:
                        messagebox.showwarning("Tamaño inválido", "Filas y columnas deben estar entre 1 y 20.")
                        return
                except ValueError:
                    messagebox.showwarning("Tamaño inválido", "Filas y columnas deben ser números.")
                    return
            try:
                nueva_id = self.db.crear_charola(
                    nombre, filas, columnas, volteada_var.get(), auto_volteo_var.get()
                )
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Ya existe una charola con ese nombre.")
                return
            ventana.destroy()
            self._refrescar_combo_charolas()
            self.combo_charola.set(f"{nueva_id} - {nombre}")
            self._on_seleccionar_charola()

        ttk.Button(ventana, text="Crear", command=confirmar).pack(pady=15)

    def _refrescar_combo_charolas(self):
        charolas = self.db.listar_charolas()
        valores = [f"{cid} - {nombre} ({f}x{c})" for cid, nombre, f, c, _v, _av, _va in charolas]
        
        # actualizar combo de Escaneo
        if self.combo_charola is not None:
            self.combo_charola["values"] = valores
            if valores and not self.combo_charola.get():
                self.combo_charola.current(0)
                self._on_seleccionar_charola()
        
        # actualizar combo de Separacion
        if self.combo_charola_sep is not None:
            self.combo_charola_sep["values"] = valores

    def _on_seleccionar_charola(self, event=None):
        seleccion = self.combo_charola.get()
        if not seleccion:
            return
        charola_id = int(seleccion.split(" - ")[0])
        row = self.db.obtener_charola(charola_id)
        if not row:
            return
        _, _, filas, columnas, volteada, _auto_v, _v_aplicado = row
        self.charola_actual_id = charola_id
        self.filas, self.columnas = filas, columnas
        self.volteada_actual = volteada
        self._dibujar_grid()

    # ---------------- Grid / escaneo ----------------
    def _dibujar_grid(self):
        for w in self.frame_grid.winfo_children():
            w.destroy()
        self.celdas = {}
        if self.charola_actual_id is None:
            return

        posiciones = generar_posiciones(self.filas, self.columnas)
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        bloqueadas = self.db.posiciones_bloqueadas_de_charola(self.charola_actual_id)

        # encabezados de columna
        for c in range(self.columnas):
            ttk.Label(self.frame_grid, text=etiqueta_columna(c), width=10, anchor="center").grid(
                row=0, column=c + 1
            )
        for f in range(self.filas):
            ttk.Label(self.frame_grid, text=str(f + 1), width=4, anchor="center").grid(row=f + 1, column=0)

        for pos in posiciones:
            f = int("".join(ch for ch in pos if ch.isdigit())) - 1
            col_letra = "".join(ch for ch in pos if ch.isalpha())
            c = _col_letra_a_indice(col_letra)
            info = datos.get(pos)
            
            # determinar color
            if pos in bloqueadas:
                color = "#FFC0CB"  # rosado: bloqueado
                texto = "🔒"
            elif info:
                color = COLOR_ESCANEADO
                prefijo = "✓ " if info.get("separada") else ""
                texto = prefijo + info["lot_id"]
            else:
                color = COLOR_PENDIENTE
                texto = pos
            
            lbl = tk.Label(
                self.frame_grid, text=texto, width=12, height=2, relief="ridge",
                bg=color, wraplength=90, font=("Segoe UI", 8),
            )
            lbl.grid(row=f + 1, column=c + 1, padx=2, pady=2)
            self.celdas[pos] = lbl

        total = len(posiciones)
        escaneadas = len(datos)
        bloqueadas_count = len(bloqueadas)
        nota_volteo = "  |  Modo VOLTEADA: fila = (filas+1)-fila, columna igual" if self.volteada_actual else ""
        nota_bloqueadas = f"  |  {bloqueadas_count} posiciones bloqueadas" if bloqueadas_count > 0 else ""
        self.lbl_progreso.config(text=f"{escaneadas} / {total} posiciones escaneadas{nota_volteo}{nota_bloqueadas}")

    def _escanear(self):
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona o crea una charola primero.")
            return
        
        lot_id = self.entry_lot_id.get().strip()
        self.entry_lot_id.delete(0, tk.END)
        if not lot_id:
            return
        
        # Validacion 1: formato válido (solo alfanuméricos y guiones)
        if not all(c.isalnum() or c in '-_' for c in lot_id):
            messagebox.showerror("Formato inválido", "Lot ID contiene caracteres no válidos.")
            return

        # Validacion 2: duplicado en cualquier parte del sistema
        existente = self.db.lot_id_ya_escaneado(lot_id)
        if existente:
            charola_id, pos = existente
            messagebox.showerror(
                "Lot ID duplicado",
                f"Este Lot ID ya fue escaneado antes\n(charola {charola_id}, posición {pos}).",
            )
            return

        # Validacion 3: obtener siguiente posición libre
        siguiente = self._siguiente_posicion_libre()
        if siguiente is None:
            messagebox.showinfo("Charola completa", "Todas las posiciones ya fueron escaneadas o están bloqueadas.")
            return

        # Validacion 4: posición no está bloqueada
        posiciones_bloqueadas = self.db.posiciones_bloqueadas_de_charola(self.charola_actual_id)
        if siguiente in posiciones_bloqueadas:
            messagebox.showerror(
                "Posición bloqueada",
                f"La posición {siguiente} está bloqueada y no puede ser escaneada.",
            )
            return

        posicion = siguiente
        self.db.escanear(self.charola_actual_id, posicion, lot_id)
        self._dibujar_grid()
        self.entry_lot_id.focus()

    def _siguiente_posicion_libre(self):
        """
        Devuelve la siguiente posicion CANONICA a llenar, siguiendo el orden
        fisico de escaneo. Salta posiciones bloqueadas.
        """
        secuencia = orden_escaneo(self.filas, self.columnas, self.volteada_actual)
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        bloqueadas = self.db.posiciones_bloqueadas_de_charola(self.charola_actual_id)
        
        for pos in secuencia:
            if pos not in datos and pos not in bloqueadas:
                return pos
        return None

    def _limpiar_charola_dialogo(self):
        """Abre diálogo para confirmar limpieza de charola."""
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return
        
        confirm = messagebox.askyesno(
            "Limpiar charola",
            "¿Eliminar TODOS los PCBs escaneados de esta charola?\nEsta acción no se puede deshacer.",
        )
        if confirm:
            self.db.limpiar_charola_completa(self.charola_actual_id)
            self._dibujar_grid()
            messagebox.showinfo("Hecho", "Charola limpiada completamente.")

    def _voltear_charola_180_dialogo(self):
        """
        Gira 180° toda la charola escaneada de una sola vez (fila y columna
        invertidas): escaneas en el orden crudo/fisico tal cual ves la charola
        volteada frente a ti, y al terminar, un solo click reubica cada pieza
        en su posicion real. Conserva Lot ID, modelo, proceso y separada.
        """
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return

        confirm = messagebox.askyesno(
            "Voltear charola 180°",
            "Esto va a girar 180° TODAS las posiciones escaneadas de esta charola\n"
            "(la fila Y la columna se invierten, como girar un plato).\n\n"
            "Úsalo solo una vez, al terminar de escanear la charola completa volteada.\n"
            "¿Continuar?",
        )
        if not confirm:
            return

        n = self.db.voltear_charola_180(self.charola_actual_id)
        self._dibujar_grid()
        self._dibujar_grid_separacion()
        if n:
            messagebox.showinfo("Hecho", f"{n} pieza(s) giradas 180° a su posición real.")
        else:
            messagebox.showinfo("Sin datos", "Esta charola no tiene piezas escaneadas para voltear.")

    def _bloquear_posicion(self):
        """Bloquea una posición específica."""
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return
        pos = self.entry_posicion_bloqueo.get().strip().upper()
        if not pos:
            messagebox.showwarning("Posición vacía", "Ingresa una posición (ej: A1, B6).")
            return
        self.db.bloquear_posicion(self.charola_actual_id, pos)
        self.entry_posicion_bloqueo.delete(0, tk.END)
        self._dibujar_grid()
        messagebox.showinfo("Hecho", f"Posición {pos} bloqueada.")

    def _desbloquear_posicion(self):
        """Desbloquea una posición específica."""
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return
        pos = self.entry_posicion_bloqueo.get().strip().upper()
        if not pos:
            messagebox.showwarning("Posición vacía", "Ingresa una posición (ej: A1, B6).")
            return
        self.db.desbloquear_posicion(self.charola_actual_id, pos)
        self.entry_posicion_bloqueo.delete(0, tk.END)
        self._dibujar_grid()
        messagebox.showinfo("Hecho", f"Posición {pos} desbloqueada.")

    def _mapear_desde_texto(self):
        """Lee el texto pegado, muestra preview, y pide confirmación antes de mapear."""
        texto = self.text_pega.get("1.0", tk.END)
        if not texto.strip():
            messagebox.showwarning("Texto vacío", "Pega datos primero.")
            return
        
        preview, errores = self.db.parsear_y_mostrar_preview(texto)
        
        if not preview:
            msg = "No se pudo parsear ningún Lot ID.\n\nErrores:\n" + "\n".join(errores[:10])
            messagebox.showerror("Parsing fallido", msg)
            return
        
        # mostrar preview en un diálogo
        ventana_preview = tk.Toplevel(self)
        ventana_preview.title("Previsualización de mapeo")
        ventana_preview.geometry("700x400")
        
        ttk.Label(
            ventana_preview,
            text=f"Se detectaron {len(preview)} Lot IDs. Revisa que se haya parseado correctamente:",
            font=("Segoe UI", 9),
        ).pack(anchor="w", padx=10, pady=(10, 4))
        
        # tabla con scroll
        frame_tabla = ttk.Frame(ventana_preview)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        cols = ("Lot ID", "Modelo", "Proceso")
        tree = ttk.Treeview(frame_tabla, columns=cols, show="headings", height=12)
        for col, width in zip(cols, (150, 250, 250)):
            tree.heading(col, text=col)
            tree.column(col, width=width)
        
        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for lot_id, modelo, proceso in preview[:50]:  # mostrar máximo 50 para no saturar
            tree.insert("", "end", values=(lot_id, modelo, proceso))
        
        if len(preview) > 50:
            ttk.Label(ventana_preview, text=f"... y {len(preview) - 50} más", foreground="#666").pack(anchor="w", padx=10)
        
        # botones
        btn_frame = ttk.Frame(ventana_preview)
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        def confirmar():
            # mapear todos
            mapeo, _ = self.db.actualizar_modelo_proceso_desde_texto(texto)
            actualizados, no_encontrados = self.db.actualizar_modelo_proceso_batch(mapeo)
            
            estado = f"Mapeados: {actualizados} PCBs."
            if no_encontrados:
                estado += f" {len(no_encontrados)} Lot ID no encontrados en el sistema."
            
            messagebox.showinfo("Mapeo completado", estado)
            self._recalcular_colores()
            self._redibujar_grid()
            self._dibujar_grid_separacion()
            self.text_pega.delete("1.0", tk.END)
            self._refrescar_historial()
            ventana_preview.destroy()
            self._aplicar_volteos_automaticos_y_avisar()
        
        ttk.Button(btn_frame, text="✓ Confirmar y mapear", command=confirmar).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="✗ Cancelar", command=ventana_preview.destroy).pack(side="left", padx=4)
        ttk.Label(btn_frame, text="Si se ve bien, confirma. Si no, cancela y revisa el formato.", foreground="#666").pack(side="left", padx=20)

    def _undo_celda(self):
        if self.charola_actual_id is None:
            return
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        if not datos:
            return
        # deshace la ultima posicion ocupada, respetando el orden fisico de escaneo
        secuencia = orden_escaneo(self.filas, self.columnas, self.volteada_actual)
        ocupadas = [p for p in secuencia if p in datos]
        if not ocupadas:
            return
        ultima = ocupadas[-1]
        confirm = messagebox.askyesno("Confirmar", f"Deshacer escaneo de la posicion {ultima}?")
        if confirm:
            self.db.undo_ultimo_escaneo(self.charola_actual_id, ultima)
            self._dibujar_grid()

    def _copiar_lot_ids(self):
        if self.charola_actual_id is None:
            return
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        lot_ids = [info["lot_id"] for info in datos.values()]
        if not lot_ids:
            messagebox.showinfo("Sin datos", "No hay Lot ID escaneados en esta charola.")
            return
        self.clipboard_clear()
        self.clipboard_append("\n".join(lot_ids))
        messagebox.showinfo("Copiado", f"{len(lot_ids)} Lot ID copiados al portapapeles.")

    # ---------------- Importar Excel / clasificacion ----------------
    def _importar_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if not path:
            return
        try:
            df = pd.read_excel(path)
        except Exception as exc:
            messagebox.showerror("Error al leer Excel", str(exc))
            return

        columnas_lower = {c.lower().strip(): c for c in df.columns}
        col_lot = _buscar_columna(columnas_lower, ["lot id", "lot_id", "lotid"])
        col_modelo = _buscar_columna(columnas_lower, ["modelo", "model", "product", "producto"])
        col_proceso = _buscar_columna(
            columnas_lower,
            ["proceso", "process", "process name", "estacion", "station", "wip status name"],
        )

        if not col_lot:
            messagebox.showerror(
                "Columnas no encontradas",
                "No se encontro una columna de Lot ID en el Excel. "
                "Se esperaba algo como 'Lot ID', 'LotID' o similar.",
            )
            return

        # Construir mapeo: lot_id -> (modelo, proceso)
        mapeo = {}
        for _, row in df.iterrows():
            lot_id = str(row[col_lot]).strip()
            modelo = str(row[col_modelo]).strip() if col_modelo else "N/D"
            proceso = str(row[col_proceso]).strip() if col_proceso else "N/D"
            mapeo[lot_id] = (modelo, proceso)

        # BATCH UPDATE contra TODOS los Lot ID de TODAS las charolas
        actualizados, no_encontrados = self.db.actualizar_modelo_proceso_batch(mapeo)

        self.db.registrar_historial(
            "excel_importado_batch",
            f"{os.path.basename(path)}: {actualizados} actualizados, {len(no_encontrados)} sin match",
        )

        estado = f"Importado: {actualizados} PCBs actualizadas en todas las charolas."
        if no_encontrados:
            porcentaje = (len(no_encontrados) / len(mapeo)) * 100
            if porcentaje > 10:
                messagebox.showwarning(
                    "Advertencia: muchos sin match",
                    f"{len(no_encontrados)} Lot ID ({porcentaje:.1f}%) del Excel no se encontraron "
                    "en ninguna charola escaneada. Verifica que hayas escaneado todas las piezas.",
                )
            estado += f" ({len(no_encontrados)} sin match)"

        self.lbl_import_status.config(text=estado)
        self._recalcular_colores()
        self._redibujar_grid()
        self._dibujar_grid_separacion()
        self._refrescar_historial()
        self._aplicar_volteos_automaticos_y_avisar()

        messagebox.showinfo("Importación batch completada", estado)

    def _aplicar_volteos_automaticos_y_avisar(self):
        """
        Aplica el volteo de 180° a cualquier charola marcada como 'auto-voltear al
        clasificar' (tipicamente charolas 6x6) que aun no se le haya aplicado.
        Se llama justo despues de pasar/clasificar datos (importar Excel o pegar
        texto), asi el usuario no tiene que acordarse de presionar el boton manual.
        """
        resultados = self.db.aplicar_volteos_automaticos_pendientes()
        if resultados:
            self._dibujar_grid()
            self._dibujar_grid_separacion()
            self._refrescar_historial()
            detalle = "\n".join(f"  • {nombre}: {n} piezas" for _cid, nombre, n in resultados)
            messagebox.showinfo(
                "Volteo automático aplicado",
                f"Se giraron 180° automáticamente estas charolas (6x6 / auto-volteo):\n\n{detalle}",
            )

    def _exportar_consolidado(self):
        """Exporta TODAS las charolas a un Excel consolidado."""
        datos_todas = self.db.exportar_todas_charolas_a_dict()
        if not datos_todas:
            messagebox.showinfo("Sin datos", "No hay charolas con datos para exportar.")
            return

        # Usar openpyxl para crear Excel estructurado
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        wb.remove(wb.active)  # elimina hoja default

        for charola_id, charola_data in datos_todas.items():
            nombre = charola_data["nombre"]
            filas = charola_data["filas"]
            columnas = charola_data["columnas"]
            posiciones = charola_data["posiciones"]

            # crear hoja para esta charola
            ws = wb.create_sheet(title=nombre[:31])  # Excel limita nombre a 31 chars

            # encabezados
            ws["A1"] = "Posicion"
            ws["B1"] = "Lot ID"
            ws["C1"] = "Modelo"
            ws["D1"] = "Proceso"
            ws["E1"] = "Separada"

            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}1"].font = Font(bold=True, color="FFFFFF")
                ws[f"{col}1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            row = 2
            for pos in generar_posiciones(filas, columnas):
                info = posiciones.get(pos)
                if not info:
                    continue
                ws[f"A{row}"] = pos
                ws[f"B{row}"] = info["lot_id"]
                ws[f"C{row}"] = info.get("modelo", "")
                ws[f"D{row}"] = info.get("proceso", "")
                ws[f"E{row}"] = "✓" if info.get("separada") else ""
                row += 1

            # auto-ajustar ancho
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 12

        # guardar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path_salida = f"retrabajo_consolidado_{timestamp}.xlsx"
        try:
            wb.save(path_salida)
            messagebox.showinfo("Exportado", f"Archivo guardado como:\n{path_salida}")
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))

    def _recalcular_colores(self):
        """Asigna un color fijo de la paleta a cada modelo y a cada proceso nuevos."""
        cur = self.db.conn.cursor()
        cur.execute("SELECT DISTINCT modelo FROM pcbs WHERE modelo IS NOT NULL")
        for (modelo,) in cur.fetchall():
            if modelo not in self.color_map_modelo:
                self.color_map_modelo[modelo] = COLOR_PALETTE[len(self.color_map_modelo) % len(COLOR_PALETTE)]
        cur.execute("SELECT DISTINCT proceso FROM pcbs WHERE proceso IS NOT NULL")
        for (proceso,) in cur.fetchall():
            if proceso not in self.color_map_proceso:
                self.color_map_proceso[proceso] = COLOR_PALETTE[len(self.color_map_proceso) % len(COLOR_PALETTE)]

    def _redibujar_grid(self):
        """Recolorea el grid de la charola actual segun la vista activa (modelo/proceso)."""
        if self.charola_actual_id is None or not self.celdas:
            return
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        criterio = self.vista_color.get()
        mapa_color = self.color_map_modelo if criterio == "modelo" else self.color_map_proceso

        for pos, lbl in self.celdas.items():
            info = datos.get(pos)
            if not info:
                lbl.config(bg=COLOR_PENDIENTE, text=pos)
                continue
            prefijo = "\u2713 " if info.get("separada") else ""
            texto = prefijo + info["lot_id"]
            valor = info.get(criterio)
            if valor and valor in mapa_color:
                lbl.config(bg=mapa_color[valor], text=texto)
            elif info.get("modelo") is None and info.get("proceso") is None:
                lbl.config(bg=COLOR_ESCANEADO, text=texto)
            else:
                lbl.config(bg=COLOR_SIN_MATCH, text=texto)

        self._dibujar_leyenda(mapa_color)

    def _dibujar_leyenda(self, mapa_color):
        for w in self.frame_leyenda.winfo_children():
            w.destroy()
        if not mapa_color:
            ttk.Label(self.frame_leyenda, text="(sin datos importados)").pack(padx=8, pady=8)
            return
        for nombre, color in mapa_color.items():
            fila = ttk.Frame(self.frame_leyenda)
            fila.pack(fill="x", padx=6, pady=2)
            tk.Label(fila, bg=color, width=2, height=1).pack(side="left")
            ttk.Label(fila, text=nombre).pack(side="left", padx=6)

    # ---------------- Historial ----------------
    def _refrescar_historial(self):
        for item in self.tree_historial.get_children():
            self.tree_historial.delete(item)
        for evento, detalle, fecha in self.db.leer_historial():
            self.tree_historial.insert("", "end", values=(evento, detalle, fecha))


def _col_letra_a_indice(letras):
    idx = 0
    for ch in letras:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _buscar_columna(columnas_lower, candidatos):
    for cand in candidatos:
        if cand in columnas_lower:
            return columnas_lower[cand]
    return None


if __name__ == "__main__":
    app = App()
    app.mainloop()
