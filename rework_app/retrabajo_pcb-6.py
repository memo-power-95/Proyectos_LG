"""
Sistema de Retrabajo y Clasificación de PCBs
=============================================
Reusa la logica de mapeo tipo "ajedrez" del remapeo_4.0.1.py, pero aplicada
al flujo de retrabajo: escaneo con validacion de duplicados, cruce contra el
Excel que regresa el ERP (modelo/proceso por Lot ID), vista de clasificacion
por color (modelo o proceso) e historial persistente.

Autor: prototipo generado con Claude para Guillermo.
"""

import os
import sqlite3
import string
import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox, filedialog

import pandas as pd

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "retrabajo_pcb.db")

# Paleta fija para que los colores sean consistentes entre sesiones.
COLOR_PALETTE = [
    "#4CAF50", "#2196F3", "#FF9800", "#9C27B0", "#F44336",
    "#00BCD4", "#8BC34A", "#E91E63", "#3F51B5", "#FFC107",
    "#795548", "#607D8B",
]
COLOR_PENDIENTE = "#D9D9D9"   # gris: sin escanear
COLOR_ESCANEADO = "#A5D6A7"   # verde suave: escaneado, aun sin cruzar con Excel
COLOR_SIN_MATCH = "#FFF176"   # amarillo: escaneado pero el Excel no trajo dato


# --------------------------------------------------------------------------
# Capa de datos
# --------------------------------------------------------------------------
class Database:
    def __init__(self, path=DB_PATH):
        self.conn = sqlite3.connect(path)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._crear_tablas()

    def _crear_tablas(self):
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS charolas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE,
                filas INTEGER,
                columnas INTEGER,
                volteada INTEGER DEFAULT 0,
                creada TEXT
            )
        """)
        # Migracion suave por si la DB ya existia sin la columna 'volteada'
        cur.execute("PRAGMA table_info(charolas)")
        columnas_existentes = [c[1] for c in cur.fetchall()]
        if "volteada" not in columnas_existentes:
            cur.execute("ALTER TABLE charolas ADD COLUMN volteada INTEGER DEFAULT 0")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS pcbs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                charola_id INTEGER,
                posicion TEXT,
                lot_id TEXT,
                modelo TEXT,
                proceso TEXT,
                separada INTEGER DEFAULT 0,
                escaneado_en TEXT,
                actualizado_en TEXT,
                FOREIGN KEY(charola_id) REFERENCES charolas(id),
                UNIQUE(charola_id, posicion)
            )
        """)
        cur.execute("PRAGMA table_info(pcbs)")
        columnas_pcbs = [c[1] for c in cur.fetchall()]
        if "separada" not in columnas_pcbs:
            cur.execute("ALTER TABLE pcbs ADD COLUMN separada INTEGER DEFAULT 0")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS historial (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                evento TEXT,
                detalle TEXT,
                fecha TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS equivalencias_modelo (
                modelo_variante TEXT PRIMARY KEY,
                modelo_base TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS posiciones_bloqueadas (
                charola_id INTEGER,
                posicion TEXT,
                PRIMARY KEY (charola_id, posicion),
                FOREIGN KEY(charola_id) REFERENCES charolas(id)
            )
        """)
        self.conn.commit()

    # --- charolas ---
    def crear_charola(self, nombre, filas, columnas, volteada=0):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO charolas (nombre, filas, columnas, volteada, creada) VALUES (?,?,?,?,?)",
            (nombre, filas, columnas, int(volteada), datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()
        etiqueta_volteo = " [volteada]" if volteada else ""
        self.registrar_historial("charola_creada", f"{nombre} ({filas}x{columnas}){etiqueta_volteo}")
        return cur.lastrowid

    def listar_charolas(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, nombre, filas, columnas, volteada FROM charolas ORDER BY id DESC")
        return cur.fetchall()

    def obtener_charola(self, charola_id):
        cur = self.conn.cursor()
        cur.execute("SELECT id, nombre, filas, columnas, volteada FROM charolas WHERE id=?", (charola_id,))
        return cur.fetchone()

    # --- pcbs / escaneo ---
    def lot_id_ya_escaneado(self, lot_id):
        """Valida duplicados en TODO el sistema, no solo en la charola actual."""
        cur = self.conn.cursor()
        cur.execute("SELECT charola_id, posicion FROM pcbs WHERE lot_id=?", (lot_id,))
        return cur.fetchone()

    def posicion_ocupada(self, charola_id, posicion):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT lot_id FROM pcbs WHERE charola_id=? AND posicion=?",
            (charola_id, posicion),
        )
        return cur.fetchone()

    def escanear(self, charola_id, posicion, lot_id):
        cur = self.conn.cursor()
        ahora = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            """INSERT INTO pcbs (charola_id, posicion, lot_id, escaneado_en, actualizado_en)
               VALUES (?,?,?,?,?)""",
            (charola_id, posicion, lot_id, ahora, ahora),
        )
        self.conn.commit()
        self.registrar_historial("pcb_escaneado", f"{lot_id} -> charola {charola_id} pos {posicion}")

    def undo_ultimo_escaneo(self, charola_id, posicion):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM pcbs WHERE charola_id=? AND posicion=?", (charola_id, posicion))
        self.conn.commit()
        self.registrar_historial("undo_escaneo", f"charola {charola_id} pos {posicion}")

    def pcbs_de_charola(self, charola_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT posicion, lot_id, modelo, proceso, separada FROM pcbs WHERE charola_id=?",
            (charola_id,),
        )
        return {
            row[0]: {"lot_id": row[1], "modelo": row[2], "proceso": row[3], "separada": bool(row[4])}
            for row in cur.fetchall()
        }

    def marcar_separada(self, charola_id, posicion, valor=True):
        cur = self.conn.cursor()
        cur.execute(
            "UPDATE pcbs SET separada=? WHERE charola_id=? AND posicion=?",
            (int(bool(valor)), charola_id, posicion),
        )
        self.conn.commit()

    def pcbs_agrupados_por_modelo(self, charola_id, usar_modelo_base=False):
        """Para la lista de separacion: agrupa las piezas ya clasificadas por modelo.

        Si usar_modelo_base=True, agrupa usando la tabla de equivalencias
        (modelo_variante -> modelo_base). Si un modelo no tiene equivalencia
        registrada, se agrupa tal cual viene del ERP (no se pierde ni se inventa nada).
        """
        cur = self.conn.cursor()
        cur.execute(
            """SELECT posicion, lot_id, modelo, proceso, separada FROM pcbs
               WHERE charola_id=? AND modelo IS NOT NULL
               ORDER BY modelo, posicion""",
            (charola_id,),
        )
        filas = cur.fetchall()

        equivalencias = {}
        if usar_modelo_base:
            equivalencias = dict(self.listar_equivalencias())

        grupos = {}
        for pos, lot_id, modelo, proceso, separada in filas:
            clave = equivalencias.get(modelo, modelo) if usar_modelo_base else modelo
            grupos.setdefault(clave, []).append(
                {
                    "posicion": pos, "lot_id": lot_id, "modelo": modelo,
                    "proceso": proceso, "separada": bool(separada),
                }
            )
        return grupos

    # --- equivalencias de modelo base ---
    def agregar_equivalencia(self, variante, base):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT OR REPLACE INTO equivalencias_modelo (modelo_variante, modelo_base) VALUES (?,?)",
            (variante.strip(), base.strip()),
        )
        self.conn.commit()
        self.registrar_historial("equivalencia_modelo", f"{variante} -> {base}")

    def eliminar_equivalencia(self, variante):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM equivalencias_modelo WHERE modelo_variante=?", (variante,))
        self.conn.commit()
        self.registrar_historial("equivalencia_eliminada", variante)

    def listar_todas_las_charolas_clasificadas(self):
        """Lista charolas que tienen al menos algunos datos clasificados."""
        cur = self.conn.cursor()
        cur.execute(
            """SELECT DISTINCT c.id, c.nombre, c.filas, c.columnas, c.volteada
               FROM charolas c
               INNER JOIN pcbs p ON c.id = p.charola_id
               WHERE p.modelo IS NOT NULL
               ORDER BY c.id"""
        )
        return cur.fetchall()

    def contar_sin_clasificar_en_charola(self, charola_id):
        """Cuántas piezas en charola sin modelo/proceso."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM pcbs WHERE charola_id=? AND (modelo IS NULL OR proceso IS NULL)",
            (charola_id,),
        )
        return cur.fetchone()[0]

    def contar_totales_en_charola(self, charola_id):
        """Total de piezas escaneadas en charola."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM pcbs WHERE charola_id=?", (charola_id,))
        return cur.fetchone()[0]

    def actualizar_modelo_proceso_batch(self, mapeo_lot_id):
        """Actualiza modelo/proceso para múltiples Lot ID de una sola vez.
        
        mapeo_lot_id: dict {lot_id: (modelo, proceso)}
        Retorna: (actualizados, no_encontrados)
        """
        cur = self.conn.cursor()
        actualizados = 0
        no_encontrados = []
        ahora = datetime.now().isoformat(timespec="seconds")

        for lot_id, (modelo, proceso) in mapeo_lot_id.items():
            resultado = cur.execute(
                "UPDATE pcbs SET modelo=?, proceso=?, actualizado_en=? WHERE lot_id=?",
                (modelo, proceso, ahora, lot_id),
            )
            if resultado.rowcount > 0:
                actualizados += 1
            else:
                no_encontrados.append(lot_id)

        self.conn.commit()
        return actualizados, no_encontrados

    def exportar_charola_a_dict(self, charola_id):
        """Exporta una charola completa como dict."""
        cur = self.conn.cursor()
        cur.execute("SELECT nombre, filas, columnas FROM charolas WHERE id=?", (charola_id,))
        charola_info = cur.fetchone()
        if not charola_info:
            return None
        nombre, filas, columnas = charola_info
        
        datos = self.pcbs_de_charola(charola_id)
        return {
            "nombre": nombre,
            "filas": filas,
            "columnas": columnas,
            "posiciones": datos,
        }

    def limpiar_charola_completa(self, charola_id):
        """Elimina TODOS los PCBs escaneados de una charola (pero conserva la charola)."""
        cur = self.conn.cursor()
        cur.execute("DELETE FROM pcbs WHERE charola_id=?", (charola_id,))
        self.conn.commit()
        self.registrar_historial("charola_limpiada", f"Charola {charola_id}: {cur.rowcount} piezas eliminadas")

    def bloquear_posicion(self, charola_id, posicion):
        """Bloquea una posición para que no pueda ser escaneada."""
        cur = self.conn.cursor()
        try:
            cur.execute(
                "INSERT INTO posiciones_bloqueadas (charola_id, posicion) VALUES (?,?)",
                (charola_id, posicion),
            )
            self.conn.commit()
            self.registrar_historial("posicion_bloqueada", f"Charola {charola_id}, pos {posicion}")
        except sqlite3.IntegrityError:
            pass  # ya existe

    def desbloquear_posicion(self, charola_id, posicion):
        """Desbloquea una posición."""
        cur = self.conn.cursor()
        cur.execute(
            "DELETE FROM posiciones_bloqueadas WHERE charola_id=? AND posicion=?",
            (charola_id, posicion),
        )
        self.conn.commit()
        if cur.rowcount > 0:
            self.registrar_historial("posicion_desbloqueada", f"Charola {charola_id}, pos {posicion}")

    def posiciones_bloqueadas_de_charola(self, charola_id):
        """Retorna lista de posiciones bloqueadas."""
        cur = self.conn.cursor()
        cur.execute("SELECT posicion FROM posiciones_bloqueadas WHERE charola_id=?", (charola_id,))
        return {row[0] for row in cur.fetchall()}

    def actualizar_modelo_proceso_desde_texto(self, texto):
        """Parsea texto 'espagueti' del RFC (múltiples espacios/tabs).
        
        Normaliza espacios y detecta automáticamente columnas.
        Retorna: dict {lot_id: (modelo, proceso)}, list errores
        """
        lineas = texto.strip().split("\n")
        mapeo = {}
        errores = []
        
        # normalizar cada linea: reemplaza tabs y espacios múltiples con |
        lineas_normalizadas = []
        for linea in lineas:
            # reemplaza tabs y espacios múltiples con separador consistente
            import re
            linea_norm = re.sub(r'[\t ]+', '|', linea.strip())
            lineas_normalizadas.append(linea_norm)
        
        # detectar estructura: probablemente columnas son Lot_ID | Producto | ... | Proceso | ...
        # heuristica: Lot_ID suele ser alfanumerico con numeros, Proceso puede contener palabras como "Install", "Screw"
        
        for i, linea in enumerate(lineas_normalizadas, start=1):
            if not linea.strip():
                continue
            
            partes = linea.split("|")
            partes = [p.strip() for p in partes if p.strip()]  # eliminar vacios
            
            if len(partes) < 2:
                errores.append(f"Línea {i}: insuficientes datos (esperaba al menos 2 campos)")
                continue
            
            # heuristica: primer campo suele ser Lot_ID (alfanumérico con números)
            lot_id = partes[0]
            
            # validar que sea un Lot_ID razonable (no "Product", "MASS", etc)
            if any(palabra in lot_id.lower() for palabra in ["product", "mass", "process", "hold", "array"]):
                continue  # es una linea de encabezado o basura
            
            # el resto de campos: buscar uno que parezca descripcion (producto/modelo)
            # y otro que parezca proceso
            modelo_cand = None
            proceso_cand = None
            
            for j in range(1, len(partes)):
                parte = partes[j]
                # buscar campo que contenga patrones de código de producto (puntos, guiones)
                if "." in parte or "-" in parte:
                    modelo_cand = parte
                # buscar proceso: Install, Screw, etc
                elif any(proc in parte for proc in ["Install", "Screw", "AVI", "TOP", "WAIT", "PROCESSING"]):
                    proceso_cand = parte
            
            # fallback: si no encontró, usar posiciones fijas
            if not modelo_cand and len(partes) > 1:
                modelo_cand = partes[1]
            if not proceso_cand and len(partes) > 2:
                proceso_cand = partes[2]
            
            modelo = modelo_cand or "N/D"
            proceso = proceso_cand or "N/D"
            
            mapeo[lot_id] = (modelo, proceso)
        
        return mapeo, errores

    def parsear_y_mostrar_preview(self, texto):
        """Parsea texto y retorna lista de (lot_id, modelo, proceso) para previsualizar."""
        mapeo, errores = self.actualizar_modelo_proceso_desde_texto(texto)
        preview = [(lot_id, mod, proc) for lot_id, (mod, proc) in mapeo.items()]
        return preview, errores

    def actualizar_modelo_proceso(self, lot_id, modelo, proceso):
        cur = self.conn.cursor()
        ahora = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            "UPDATE pcbs SET modelo=?, proceso=?, actualizado_en=? WHERE lot_id=?",
            (modelo, proceso, ahora, lot_id),
        )
        self.conn.commit()
        return cur.rowcount

    def todos_los_lot_ids_pendientes_de_cruce(self):
        cur = self.conn.cursor()
        cur.execute("SELECT DISTINCT lot_id FROM pcbs WHERE modelo IS NULL")
        return [r[0] for r in cur.fetchall()]

    def todos_los_lot_ids(self):
        cur = self.conn.cursor()
        cur.execute("SELECT DISTINCT lot_id FROM pcbs")
        return [r[0] for r in cur.fetchall()]

    # --- historial ---
    def registrar_historial(self, evento, detalle):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO historial (evento, detalle, fecha) VALUES (?,?,?)",
            (evento, detalle, datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()

    def leer_historial(self, limite=200):
        cur = self.conn.cursor()
        cur.execute("SELECT evento, detalle, fecha FROM historial ORDER BY id DESC LIMIT ?", (limite,))
        return cur.fetchall()


# --------------------------------------------------------------------------
# Utilidades de mapeo tipo ajedrez
# --------------------------------------------------------------------------
def etiqueta_columna(idx):
    """0 -> A, 1 -> B ... 25 -> Z, 26 -> AA ..."""
    letras = string.ascii_uppercase
    resultado = ""
    idx += 1
    while idx > 0:
        idx, resto = divmod(idx - 1, 26)
        resultado = letras[resto] + resultado
    return resultado


def generar_posiciones(filas, columnas):
    """Genera posiciones canonicas tipo ajedrez recorriendo fila por fila."""
    posiciones = []
    for f in range(filas):
        for c in range(columnas):
            posiciones.append(f"{etiqueta_columna(c)}{f + 1}")
    return posiciones


def voltear_posicion(pos, filas):
    """
    Traduce una posicion cuando la charola se voltea 180 grados de arriba-abajo
    para escanear el QR que queda boca abajo (tipico en charolas de 36).

    La columna se mantiene igual; la fila se invierte:
        fila_nueva = (filas + 1) - fila_original
    Ej. con 6 filas: A1 <-> A6, C3 <-> C4, C5 <-> C2.
    """
    col_letra = "".join(ch for ch in pos if ch.isalpha())
    fila_num = int("".join(ch for ch in pos if ch.isdigit()))
    fila_nueva = (filas + 1) - fila_num
    return f"{col_letra}{fila_nueva}"


def orden_escaneo(filas, columnas, volteada):
    """
    Devuelve la lista de posiciones CANONICAS en el orden real en que se van
    llenando conforme el operador escanea fisicamente la charola.

    Si la charola esta volteada, el operador recorre la charola en su orden
    fisico normal (fila por fila) pero cada pieza que toca en realidad
    corresponde a la posicion canonica opuesta verticalmente.
    """
    fisico = generar_posiciones(filas, columnas)
    if not volteada:
        return fisico
    return [voltear_posicion(p, filas) for p in fisico]


# --------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Retrabajo y Clasificacion de PCBs")
        self.geometry("1150x720")
        self.db = Database()

        # atributos generales
        self.charola_actual_id = None
        self.filas = 6
        self.columnas = 6
        self.volteada_actual = 0
        self.celdas_sep_seleccionadas = set()
        self.combo_charola = None
        self.combo_charola_sep = None
        self.orden_estricto = tk.BooleanVar(value=False)
        self.vista_color = tk.StringVar(value="proceso")
        self.usar_modelo_base = tk.BooleanVar(value=False)
        self.vista_color_sep = tk.StringVar(value="modelo")
        self.celdas = {}
        self.celdas_sep = {}
        self.color_map_modelo = {}
        self.color_map_proceso = {}
        
        # atributos de UI que se crean en _construir_tab_*
        self.text_pega = None
        self.entry_lot_id = None
        self.entry_posicion_bloqueo = None
        self.frame_grid = None
        self.frame_grid_separacion = None
        self.frame_picking = None
        self.frame_leyenda = None
        self.frame_leyenda_sep = None
        self.lbl_progreso = None
        self.lbl_import_status = None
        self.canvas_picking = None
        self.tree_historial = None
        self.tree_equivalencias = None
        self.combo_charola_sep = None

        self._construir_ui()

    # ---------------- UI construction ----------------
    def _construir_ui(self):
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)

        self.tab_escaneo = ttk.Frame(notebook)
        self.tab_clasificacion = ttk.Frame(notebook)
        self.tab_separacion = ttk.Frame(notebook)
        self.tab_modelos_base = ttk.Frame(notebook)
        self.tab_historial = ttk.Frame(notebook)

        notebook.add(self.tab_escaneo, text="Escaneo")
        notebook.add(self.tab_clasificacion, text="Clasificacion / Excel")
        notebook.add(self.tab_separacion, text="Separacion")
        notebook.add(self.tab_modelos_base, text="Modelos base")
        notebook.add(self.tab_historial, text="Historial")

        self._construir_tab_escaneo()
        self._construir_tab_clasificacion()
        self._construir_tab_separacion()
        self._construir_tab_modelos_base()
        self._construir_tab_historial()
        
        # refrescar combos DESPUES de que todas las pestañas esten construidas
        self._refrescar_combo_charolas()

    def _construir_tab_escaneo(self):
        top = ttk.Frame(self.tab_escaneo)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="Charola:").grid(row=0, column=0, sticky="w")
        self.combo_charola = ttk.Combobox(top, width=30, state="readonly")
        self.combo_charola.grid(row=0, column=1, padx=5)
        self.combo_charola.bind("<<ComboboxSelected>>", self._on_seleccionar_charola)

        ttk.Button(top, text="Nueva charola", command=self._dialogo_nueva_charola).grid(row=0, column=2, padx=5)
        ttk.Checkbutton(top, text="Forzar orden estricto de escaneo", variable=self.orden_estricto).grid(
            row=0, column=3, padx=15
        )

        entry_frame = ttk.Frame(self.tab_escaneo)
        entry_frame.pack(fill="x", padx=10, pady=4)
        ttk.Label(entry_frame, text="Escanear Lot ID:").pack(side="left")
        self.entry_lot_id = ttk.Entry(entry_frame, width=35)
        self.entry_lot_id.pack(side="left", padx=6)
        self.entry_lot_id.bind("<Return>", lambda e: self._escanear())
        ttk.Button(entry_frame, text="Escanear", command=self._escanear).pack(side="left", padx=4)
        ttk.Button(entry_frame, text="Deshacer ultima celda", command=self._undo_celda).pack(side="left", padx=4)
        ttk.Button(entry_frame, text="Copiar todos los Lot ID", command=self._copiar_lot_ids).pack(
            side="left", padx=12
        )
        ttk.Button(entry_frame, text="Limpiar charola", command=self._limpiar_charola_dialogo).pack(
            side="left", padx=4
        )

        self.lbl_progreso = ttk.Label(self.tab_escaneo, text="0 / 0 posiciones escaneadas")
        self.lbl_progreso.pack(anchor="w", padx=10)

        # opciones de bloqueo
        block_frame = ttk.Frame(self.tab_escaneo)
        block_frame.pack(fill="x", padx=10, pady=4)
        ttk.Label(block_frame, text="Bloquear posición:").pack(side="left")
        self.entry_posicion_bloqueo = ttk.Entry(block_frame, width=8)
        self.entry_posicion_bloqueo.pack(side="left", padx=4)
        ttk.Button(block_frame, text="Bloquear", command=self._bloquear_posicion).pack(side="left", padx=2)
        ttk.Button(block_frame, text="Desbloquear", command=self._desbloquear_posicion).pack(side="left", padx=2)
        ttk.Label(block_frame, text="(ej: A1, B6)", foreground="#666").pack(side="left", padx=4)

        self.frame_grid = ttk.Frame(self.tab_escaneo)
        self.frame_grid.pack(padx=10, pady=10)

    def _construir_tab_clasificacion(self):
        top = ttk.Frame(self.tab_clasificacion)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Button(top, text="Importar Excel del ERP (BATCH - todas charolas)", command=self._importar_excel).pack(side="left")
        ttk.Button(top, text="Exportar consolidado (Excel)", command=self._exportar_consolidado).pack(side="left", padx=10)

        ttk.Label(top, text="Ver por:").pack(side="left", padx=(20, 4))
        ttk.Radiobutton(
            top, text="Modelo", variable=self.vista_color, value="modelo", command=self._redibujar_grid
        ).pack(side="left")
        ttk.Radiobutton(
            top, text="Proceso", variable=self.vista_color, value="proceso", command=self._redibujar_grid
        ).pack(side="left")

        self.lbl_import_status = ttk.Label(self.tab_clasificacion, text="Sin archivo importado todavia.")
        self.lbl_import_status.pack(anchor="w", padx=10)

        # area de pega masiva
        pega_frame = ttk.LabelFrame(self.tab_clasificacion, text="Pegar datos masivos (Lot ID + Proceso +Modelo)")
        pega_frame.pack(fill="x", padx=10, pady=(8, 4))
        ttk.Label(
            pega_frame,
            text="Pega los datos aqui (tab/espacio separados: Lot_ID Modelo Proceso). Una linea por pieza.",
            font=("Segoe UI", 8), foreground="#666",
        ).pack(anchor="w", padx=8, pady=(4, 2))

        self.text_pega = tk.Text(pega_frame, height=6, width=80, font=("Courier New", 9))
        self.text_pega.pack(fill="x", padx=8, pady=(0, 4))

        pega_btn_frame = ttk.Frame(pega_frame)
        pega_btn_frame.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(pega_btn_frame, text="Mapear desde portapapeles", command=self._mapear_desde_texto).pack(
            side="left", padx=2
        )
        ttk.Button(pega_btn_frame, text="Limpiar texto", command=lambda: self.text_pega.delete("1.0", tk.END)).pack(
            side="left", padx=2
        )

        cuerpo = ttk.Frame(self.tab_clasificacion)
        cuerpo.pack(fill="both", expand=True, padx=10, pady=10)

        self.frame_leyenda = ttk.LabelFrame(cuerpo, text="Leyenda")
        self.frame_leyenda.pack(side="right", fill="y", padx=(10, 0))

    def _construir_tab_separacion(self):
        top = ttk.Frame(self.tab_separacion)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="Charola:").pack(side="left")
        self.combo_charola_sep = ttk.Combobox(top, width=30, state="readonly")
        self.combo_charola_sep.pack(side="left", padx=5)
        self.combo_charola_sep.bind("<<ComboboxSelected>>", lambda e: self._refrescar_picking_list())
        ttk.Button(top, text="Actualizar", command=self._refrescar_picking_list).pack(side="left", padx=5)

        self.usar_modelo_base = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            top, text="Agrupar por modelo base (usa tabla de equivalencias)",
            variable=self.usar_modelo_base, command=self._refrescar_picking_list,
        ).pack(side="left", padx=15)

        # --- Vista grafica tipo mapa (igual que Clasificacion, pero con esta charola) ---
        barra_vista = ttk.Frame(self.tab_separacion)
        barra_vista.pack(fill="x", padx=10, pady=(4, 0))
        ttk.Label(barra_vista, text="Mapa de charola - ver por:").pack(side="left")
        self.vista_color_sep = tk.StringVar(value="modelo")
        ttk.Radiobutton(
            barra_vista, text="Modelo", variable=self.vista_color_sep, value="modelo",
            command=self._dibujar_grid_separacion,
        ).pack(side="left")
        ttk.Radiobutton(
            barra_vista, text="Proceso", variable=self.vista_color_sep, value="proceso",
            command=self._dibujar_grid_separacion,
        ).pack(side="left", padx=6)
        ttk.Button(barra_vista, text="Copiar seleccionados", command=self._copiar_seleccionados_sep).pack(
            side="left", padx=15
        )

        panel_mapa = ttk.Frame(self.tab_separacion)
        panel_mapa.pack(fill="x", padx=10, pady=6)
        self.frame_grid_separacion = ttk.Frame(panel_mapa)
        self.frame_grid_separacion.pack(side="left")
        self.frame_leyenda_sep = ttk.LabelFrame(panel_mapa, text="Leyenda")
        self.frame_leyenda_sep.pack(side="left", padx=15, fill="y")

        ttk.Label(
            self.tab_separacion,
            text=(
                "Marca cada pieza conforme la muevas fisicamente a su charola de destino por modelo. "
                "El check se guarda al instante."
            ),
            foreground="#555555",
        ).pack(anchor="w", padx=10, pady=(0, 6))

        # Area con scroll para las listas de picking por modelo
        contenedor = ttk.Frame(self.tab_separacion)
        contenedor.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        canvas = tk.Canvas(contenedor, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(contenedor, orient="vertical", command=canvas.yview)
        self.frame_picking = ttk.Frame(canvas)
        self.frame_picking.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.frame_picking, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.canvas_picking = canvas

    def _dibujar_grid_separacion(self):
        """Dibuja el mapa visual de charola en Separacion, mostrando modelo/proceso por celda.
        Las celdas son clickeables para seleccionar/deseleccionar para copiar.
        """
        for w in self.frame_grid_separacion.winfo_children():
            w.destroy()

        seleccion = self.combo_charola_sep.get()
        if not seleccion or self.charola_actual_id != int(seleccion.split(" - ")[0]):
            return

        charola_id = self.charola_actual_id
        filas, columnas = self.filas, self.columnas
        volteada = self.volteada_actual
        criterio = self.vista_color_sep.get()

        datos = self.db.pcbs_de_charola(charola_id)
        posiciones = generar_posiciones(filas, columnas)

        # generar mapas de color
        color_map = {}
        for pos, info in datos.items():
            valor = info.get(criterio)
            if valor and valor not in color_map:
                color_map[valor] = COLOR_PALETTE[len(color_map) % len(COLOR_PALETTE)]

        # encabezados columna
        for c in range(columnas):
            ttk.Label(self.frame_grid_separacion, text=etiqueta_columna(c), width=12, anchor="center").grid(
                row=0, column=c + 1
            )
        for f in range(filas):
            ttk.Label(self.frame_grid_separacion, text=str(f + 1), width=4, anchor="center").grid(
                row=f + 1, column=0
            )

        # celdas (clickeables)
        self.celdas_sep = {}
        for pos in posiciones:
            f = int("".join(ch for ch in pos if ch.isdigit())) - 1
            col_letra = "".join(ch for ch in pos if ch.isalpha())
            c = _col_letra_a_indice(col_letra)
            info = datos.get(pos)

            if not info:
                color = COLOR_PENDIENTE
                texto = pos
                fg = "black"
            else:
                valor = info.get(criterio)
                prefijo = "✓ " if info.get("separada") else ""
                texto = prefijo + (valor if valor else "N/D")
                color = color_map.get(valor, COLOR_SIN_MATCH) if valor else COLOR_ESCANEADO
                fg = "black"

            # resaltado si está seleccionada para copiar
            relief = "sunken" if pos in self.celdas_sep_seleccionadas else "ridge"
            width_label = 14 if pos not in self.celdas_sep_seleccionadas else 14
            bd = 3 if pos in self.celdas_sep_seleccionadas else 1

            lbl = tk.Label(
                self.frame_grid_separacion, text=texto, width=width_label, height=2,
                relief=relief, bg=color, fg=fg, wraplength=100, font=("Segoe UI", 8),
                bd=bd, cursor="hand2",
            )
            lbl.grid(row=f + 1, column=c + 1, padx=2, pady=2)
            lbl.bind("<Button-1>", lambda e, pos=pos: self._toggle_celda_sep(pos))
            self.celdas_sep[pos] = lbl

        # leyenda (clickeable también)
        self._dibujar_leyenda_separacion(color_map)

    def _dibujar_leyenda_separacion(self, mapa_color):
        for w in self.frame_leyenda_sep.winfo_children():
            w.destroy()
        if not mapa_color:
            ttk.Label(self.frame_leyenda_sep, text="(sin datos importados)").pack(padx=8, pady=8)
            return
        
        ttk.Label(self.frame_leyenda_sep, text="Haz click para seleccionar/deseleccionar", 
                  font=("Segoe UI", 7), foreground="#666").pack(padx=4, pady=2)
        
        for nombre, color in mapa_color.items():
            fila = tk.Frame(self.frame_leyenda_sep, bg="white", relief="raised", bd=1, cursor="hand2")
            fila.pack(fill="x", padx=4, pady=2, ipady=2)
            fila.bind("<Button-1>", lambda e, cat=nombre: self._toggle_color_sep(cat, mapa_color))
            
            tk.Label(fila, bg=color, width=2, height=1).pack(side="left", padx=2)
            tk.Label(fila, text=nombre, bg="white", fg="black", cursor="hand2").pack(side="left", padx=6, fill="x", expand=True)
            fila.bind("<Button-1>", lambda e, cat=nombre: self._toggle_color_sep(cat, mapa_color))

    def _toggle_celda_sep(self, posicion):
        """Marca/desmarca una celda individual para copiar."""
        if posicion in self.celdas_sep_seleccionadas:
            self.celdas_sep_seleccionadas.discard(posicion)
        else:
            self.celdas_sep_seleccionadas.add(posicion)
        self._dibujar_grid_separacion()

    def _toggle_color_sep(self, categoria, mapa_color):
        """Selecciona/deselecciona todas las celdas de un color (modelo/proceso)."""
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        criterio = self.vista_color_sep.get()
        
        # encuentra todas las posiciones que tienen este color
        posiciones_color = [pos for pos, info in datos.items() if info.get(criterio) == categoria]
        
        # si todas ya están seleccionadas, deselecciona; si no, selecciona todas
        si_todas_marcadas = all(pos in self.celdas_sep_seleccionadas for pos in posiciones_color)
        
        if si_todas_marcadas:
            for pos in posiciones_color:
                self.celdas_sep_seleccionadas.discard(pos)
        else:
            for pos in posiciones_color:
                self.celdas_sep_seleccionadas.add(pos)
        
        self._dibujar_grid_separacion()

    def _copiar_seleccionados_sep(self):
        """Copia los Lot ID de las celdas seleccionadas al portapapeles."""
        if not self.celdas_sep_seleccionadas:
            messagebox.showinfo("Sin selección", "Selecciona al menos una pieza haciendo click en el mapa.")
            return
        
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        lot_ids = [datos[pos]["lot_id"] for pos in self.celdas_sep_seleccionadas if pos in datos]
        
        if not lot_ids:
            messagebox.showinfo("Sin datos", "Las posiciones seleccionadas no tienen Lot ID.")
            return
        
        self.clipboard_clear()
        self.clipboard_append("\n".join(lot_ids))
        messagebox.showinfo("Copiado", f"{len(lot_ids)} Lot ID copiados al portapapeles.")

    def _refrescar_picking_list(self):
        for w in self.frame_picking.winfo_children():
            w.destroy()

        seleccion = self.combo_charola_sep.get()
        if not seleccion:
            ttk.Label(self.frame_picking, text="Selecciona una charola arriba.").pack(padx=10, pady=10)
            return

        charola_id = int(seleccion.split(" - ")[0])
        self.charola_actual_id = charola_id
        row = self.db.obtener_charola(charola_id)
        if row:
            _, _, self.filas, self.columnas, self.volteada_actual = row
        grupos = self.db.pcbs_agrupados_por_modelo(charola_id, usar_modelo_base=self.usar_modelo_base.get())

        if not grupos:
            ttk.Label(
                self.frame_picking,
                text="Esta charola aun no tiene piezas clasificadas.\n"
                     "Importa el Excel del ERP en la pestana 'Clasificacion / Excel' primero.",
                justify="left",
            ).pack(padx=10, pady=10, anchor="w")
            return

        for modelo, piezas in grupos.items():
            total = len(piezas)
            separadas = sum(1 for p in piezas if p["separada"])
            titulo = f"Modelo: {modelo}   ({separadas}/{total} separadas)"
            grupo_frame = ttk.LabelFrame(self.frame_picking, text=titulo)
            grupo_frame.pack(fill="x", padx=6, pady=6, anchor="w")

            barra = ttk.Frame(grupo_frame)
            barra.pack(fill="x", pady=(2, 6))
            ttk.Button(
                barra, text="Marcar todas",
                command=lambda cid=charola_id, m=modelo: self._marcar_grupo(cid, m, True),
            ).pack(side="left", padx=4)
            ttk.Button(
                barra, text="Desmarcar todas",
                command=lambda cid=charola_id, m=modelo: self._marcar_grupo(cid, m, False),
            ).pack(side="left", padx=4)
            ttk.Button(
                barra, text="Copiar Lot ID de este grupo",
                command=lambda piezas=piezas: self._copiar_grupo(piezas),
            ).pack(side="left", padx=12)

            for pieza in piezas:
                var = tk.BooleanVar(value=pieza["separada"])
                proceso_txt = pieza["proceso"] if pieza["proceso"] else "sin proceso"
                texto = f'{pieza["posicion"]}   {pieza["lot_id"]}   ({proceso_txt})'
                ttk.Checkbutton(
                    grupo_frame, text=texto, variable=var,
                    command=lambda cid=charola_id, pos=pieza["posicion"], var=var: self._toggle_separada(
                        cid, pos, var
                    ),
                ).pack(anchor="w", padx=20, pady=1)

        # validaciones de error proofing
        total_charola = self.db.contar_totales_en_charola(charola_id)
        sin_clasificar = self.db.contar_sin_clasificar_en_charola(charola_id)
        if total_charola > 0 and sin_clasificar > 0:
            porcentaje_sin = (sin_clasificar / total_charola) * 100
            if porcentaje_sin > 50:
                messagebox.showwarning(
                    "Alerta: mucho sin clasificar",
                    f"Esta charola tiene {sin_clasificar}/{total_charola} piezas ({porcentaje_sin:.1f}%) sin clasificar.\n"
                    "Asegúrate de que el Excel del ERP se importó correctamente.",
                )
            elif porcentaje_sin > 10:
                self.db.registrar_historial("validacion_alerta", f"Charola {charola_id}: {porcentaje_sin:.1f}% sin clasificar")

        # redibuja el grid visual tras refrescar la lista de picking
        self._dibujar_grid_separacion()

    def _toggle_separada(self, charola_id, posicion, var):
        self.db.marcar_separada(charola_id, posicion, var.get())
        estado = "separada" if var.get() else "sin separar"
        self.db.registrar_historial("separacion", f"charola {charola_id} pos {posicion} -> {estado}")
        self._actualizar_titulos_picking()
        self._redibujar_grid()
        self._dibujar_grid_separacion()  # actualiza el grid visual de Separacion

    def _marcar_grupo(self, charola_id, modelo, valor):
        grupos = self.db.pcbs_agrupados_por_modelo(charola_id, usar_modelo_base=self.usar_modelo_base.get())
        for pieza in grupos.get(modelo, []):
            self.db.marcar_separada(charola_id, pieza["posicion"], valor)
        estado = "todas separadas" if valor else "todas sin separar"
        self.db.registrar_historial("separacion_grupo", f"charola {charola_id} modelo {modelo} -> {estado}")
        self._refrescar_picking_list()
        self._redibujar_grid()

    def _actualizar_titulos_picking(self):
        # simplemente redibuja toda la lista para refrescar los contadores (N/N separadas)
        self._refrescar_picking_list()

    def _copiar_grupo(self, piezas):
        lot_ids = [p["lot_id"] for p in piezas]
        if not lot_ids:
            return
        self.clipboard_clear()
        self.clipboard_append("\n".join(lot_ids))
        messagebox.showinfo("Copiado", f"{len(lot_ids)} Lot ID de este modelo copiados al portapapeles.")

    def _construir_tab_modelos_base(self):
        ttk.Label(
            self.tab_modelos_base,
            text=(
                "Tabla de equivalencias: registra aqui como se reduce cada modelo variante a su "
                "modelo base al relocalizar en InstalPCB (ej. 6L10 -> 6L). Confirma la regla exacta "
                "con tu equipo antes de llenarla; mientras tanto, la separacion sigue funcionando "
                "con el modelo tal cual lo entrega el ERP."
            ),
            wraplength=780, justify="left", foreground="#555555",
        ).pack(anchor="w", padx=10, pady=10)

        form = ttk.Frame(self.tab_modelos_base)
        form.pack(fill="x", padx=10, pady=4)
        ttk.Label(form, text="Modelo variante (ERP):").grid(row=0, column=0, sticky="w")
        self.entry_variante = ttk.Entry(form, width=20)
        self.entry_variante.grid(row=0, column=1, padx=6)
        ttk.Label(form, text="Modelo base:").grid(row=0, column=2, sticky="w", padx=(15, 0))
        self.entry_base = ttk.Entry(form, width=20)
        self.entry_base.grid(row=0, column=3, padx=6)
        ttk.Button(form, text="Agregar / actualizar", command=self._agregar_equivalencia).grid(
            row=0, column=4, padx=10
        )

        cols = ("variante", "base")
        self.tree_equivalencias = ttk.Treeview(self.tab_modelos_base, columns=cols, show="headings", height=12)
        self.tree_equivalencias.heading("variante", text="Modelo variante")
        self.tree_equivalencias.heading("base", text="Modelo base")
        self.tree_equivalencias.column("variante", width=250)
        self.tree_equivalencias.column("base", width=250)
        self.tree_equivalencias.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Button(self.tab_modelos_base, text="Eliminar seleccionada", command=self._eliminar_equivalencia).pack(
            anchor="w", padx=10, pady=(0, 10)
        )

        self._refrescar_equivalencias()

    def _agregar_equivalencia(self):
        variante = self.entry_variante.get().strip()
        base = self.entry_base.get().strip()
        if not variante or not base:
            messagebox.showwarning("Datos incompletos", "Ingresa el modelo variante y su modelo base.")
            return
        self.db.agregar_equivalencia(variante, base)
        self.entry_variante.delete(0, tk.END)
        self.entry_base.delete(0, tk.END)
        self._refrescar_equivalencias()

    def _eliminar_equivalencia(self):
        seleccion = self.tree_equivalencias.selection()
        if not seleccion:
            return
        variante = self.tree_equivalencias.item(seleccion[0], "values")[0]
        self.db.eliminar_equivalencia(variante)
        self._refrescar_equivalencias()

    def _refrescar_equivalencias(self):
        for item in self.tree_equivalencias.get_children():
            self.tree_equivalencias.delete(item)
        for variante, base in self.db.listar_equivalencias():
            self.tree_equivalencias.insert("", "end", values=(variante, base))

    def _construir_tab_historial(self):
        cols = ("evento", "detalle", "fecha")
        self.tree_historial = ttk.Treeview(self.tab_historial, columns=cols, show="headings")
        for c, w in zip(cols, (150, 500, 160)):
            self.tree_historial.heading(c, text=c.capitalize())
            self.tree_historial.column(c, width=w)
        self.tree_historial.pack(fill="both", expand=True, padx=10, pady=10)
        ttk.Button(self.tab_historial, text="Actualizar", command=self._refrescar_historial).pack(pady=4)
        self._refrescar_historial()

    # ---------------- Charolas ----------------
    def _dialogo_nueva_charola(self):
        ventana = tk.Toplevel(self)
        ventana.title("Nueva charola")
        ventana.geometry("300x220")

        ttk.Label(ventana, text="Nombre / ID de charola:").pack(pady=(10, 0))
        entry_nombre = ttk.Entry(ventana)
        entry_nombre.pack(pady=4)

        ttk.Label(ventana, text="Tamano:").pack(pady=(10, 0))
        combo_tamano = ttk.Combobox(
            ventana, state="readonly",
            values=["48 (6x8)", "Personalizado"],
        )
        combo_tamano.current(0)
        combo_tamano.pack(pady=4)

        volteada_var = tk.BooleanVar(value=True)  # 36 (6x6) suele ir con QR boca abajo
        chk_volteada = ttk.Checkbutton(
            ventana,
            text="Charola volteada (QR boca abajo, tipico en charolas de 36)",
            variable=volteada_var,
        )
        chk_volteada.pack(pady=(8, 0))

        frame_custom = ttk.Frame(ventana)
        entry_f = ttk.Entry(frame_custom, width=5)
        entry_c = ttk.Entry(frame_custom, width=5)

        def on_tamano_change(event=None):
            seleccion = combo_tamano.get()
            if seleccion == "Personalizado":
                frame_custom.pack(pady=4)
                ttk.Label(frame_custom, text="Filas:").grid(row=0, column=0)
                entry_f.grid(row=0, column=1, padx=4)
                ttk.Label(frame_custom, text="Columnas:").grid(row=0, column=2)
                entry_c.grid(row=0, column=3, padx=4)
            else:
                frame_custom.pack_forget()
            # 48 (6x8) es la default, no tiene volteada
            volteada_var.set(False)

        combo_tamano.bind("<<ComboboxSelected>>", on_tamano_change)

        def confirmar():
            nombre = entry_nombre.get().strip()
            if not nombre:
                messagebox.showwarning("Falta nombre", "Ingresa un nombre o ID para la charola.")
                return
            seleccion = combo_tamano.get()
            if seleccion == "48 (6x8)":
                filas, columnas = 6, 8
            else:  # Personalizado
                try:
                    filas, columnas = int(entry_f.get()), int(entry_c.get())
                    if filas < 1 or columnas < 1 or filas > 20 or columnas > 20:
                        messagebox.showwarning("Tamaño inválido", "Filas y columnas deben estar entre 1 y 20.")
                        return
                except ValueError:
                    messagebox.showwarning("Tamaño inválido", "Filas y columnas deben ser números.")
                    return
            try:
                nueva_id = self.db.crear_charola(nombre, filas, columnas, volteada_var.get())
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Ya existe una charola con ese nombre.")
                return
            ventana.destroy()
            self._refrescar_combo_charolas()
            self.combo_charola.set(f"{nueva_id} - {nombre}")
            self._on_seleccionar_charola()

        ttk.Button(ventana, text="Crear", command=confirmar).pack(pady=15)

    def _refrescar_combo_charolas(self):
        charolas = self.db.listar_charolas()
        valores = [f"{cid} - {nombre} ({f}x{c})" for cid, nombre, f, c, _v in charolas]
        
        # actualizar combo de Escaneo
        if self.combo_charola is not None:
            self.combo_charola["values"] = valores
            if valores and not self.combo_charola.get():
                self.combo_charola.current(0)
                self._on_seleccionar_charola()
        
        # actualizar combo de Separacion
        if self.combo_charola_sep is not None:
            self.combo_charola_sep["values"] = valores

    def _on_seleccionar_charola(self, event=None):
        seleccion = self.combo_charola.get()
        if not seleccion:
            return
        charola_id = int(seleccion.split(" - ")[0])
        row = self.db.obtener_charola(charola_id)
        if not row:
            return
        _, _, filas, columnas, volteada = row
        self.charola_actual_id = charola_id
        self.filas, self.columnas = filas, columnas
        self.volteada_actual = volteada
        self._dibujar_grid()

    # ---------------- Grid / escaneo ----------------
    def _dibujar_grid(self):
        for w in self.frame_grid.winfo_children():
            w.destroy()
        self.celdas = {}
        if self.charola_actual_id is None:
            return

        posiciones = generar_posiciones(self.filas, self.columnas)
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        bloqueadas = self.db.posiciones_bloqueadas_de_charola(self.charola_actual_id)

        # encabezados de columna
        for c in range(self.columnas):
            ttk.Label(self.frame_grid, text=etiqueta_columna(c), width=10, anchor="center").grid(
                row=0, column=c + 1
            )
        for f in range(self.filas):
            ttk.Label(self.frame_grid, text=str(f + 1), width=4, anchor="center").grid(row=f + 1, column=0)

        for pos in posiciones:
            f = int("".join(ch for ch in pos if ch.isdigit())) - 1
            col_letra = "".join(ch for ch in pos if ch.isalpha())
            c = _col_letra_a_indice(col_letra)
            info = datos.get(pos)
            
            # determinar color
            if pos in bloqueadas:
                color = "#FFC0CB"  # rosado: bloqueado
                texto = "🔒"
            elif info:
                color = COLOR_ESCANEADO
                prefijo = "✓ " if info.get("separada") else ""
                texto = prefijo + info["lot_id"]
            else:
                color = COLOR_PENDIENTE
                texto = pos
            
            lbl = tk.Label(
                self.frame_grid, text=texto, width=12, height=2, relief="ridge",
                bg=color, wraplength=90, font=("Segoe UI", 8),
            )
            lbl.grid(row=f + 1, column=c + 1, padx=2, pady=2)
            self.celdas[pos] = lbl

        total = len(posiciones)
        escaneadas = len(datos)
        bloqueadas_count = len(bloqueadas)
        nota_volteo = "  |  Modo VOLTEADA: fila = (filas+1)-fila, columna igual" if self.volteada_actual else ""
        nota_bloqueadas = f"  |  {bloqueadas_count} posiciones bloqueadas" if bloqueadas_count > 0 else ""
        self.lbl_progreso.config(text=f"{escaneadas} / {total} posiciones escaneadas{nota_volteo}{nota_bloqueadas}")

    def _siguiente_posicion_libre(self):
        """
        Devuelve la siguiente posicion CANONICA a llenar, siguiendo el orden
        fisico de escaneo. Salta posiciones bloqueadas.
        """
        secuencia = orden_escaneo(self.filas, self.columnas, self.volteada_actual)
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        bloqueadas = self.db.posiciones_bloqueadas_de_charola(self.charola_actual_id)
        
        for pos in secuencia:
            if pos not in datos and pos not in bloqueadas:
                return pos
        return None

    def _limpiar_charola_dialogo(self):
        """Abre diálogo para confirmar limpieza de charola."""
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return
        
        confirm = messagebox.askyesno(
            "Limpiar charola",
            "¿Eliminar TODOS los PCBs escaneados de esta charola?\nEsta acción no se puede deshacer.",
        )
        if confirm:
            self.db.limpiar_charola_completa(self.charola_actual_id)
            self._dibujar_grid()
            messagebox.showinfo("Hecho", "Charola limpiada completamente.")

    def _bloquear_posicion(self):
        """Bloquea una posición específica."""
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return
        pos = self.entry_posicion_bloqueo.get().strip().upper()
        if not pos:
            messagebox.showwarning("Posición vacía", "Ingresa una posición (ej: A1, B6).")
            return
        self.db.bloquear_posicion(self.charola_actual_id, pos)
        self.entry_posicion_bloqueo.delete(0, tk.END)
        self._dibujar_grid()
        messagebox.showinfo("Hecho", f"Posición {pos} bloqueada.")

    def _desbloquear_posicion(self):
        """Desbloquea una posición específica."""
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona una charola primero.")
            return
        pos = self.entry_posicion_bloqueo.get().strip().upper()
        if not pos:
            messagebox.showwarning("Posición vacía", "Ingresa una posición (ej: A1, B6).")
            return
        self.db.desbloquear_posicion(self.charola_actual_id, pos)
        self.entry_posicion_bloqueo.delete(0, tk.END)
        self._dibujar_grid()
        messagebox.showinfo("Hecho", f"Posición {pos} desbloqueada.")

    def _mapear_desde_texto(self):
        """Lee el texto pegado, muestra preview, y pide confirmación antes de mapear."""
        texto = self.text_pega.get("1.0", tk.END)
        if not texto.strip():
            messagebox.showwarning("Texto vacío", "Pega datos primero.")
            return
        
        preview, errores = self.db.parsear_y_mostrar_preview(texto)
        
        if not preview:
            msg = "No se pudo parsear ningún Lot ID.\n\nErrores:\n" + "\n".join(errores[:10])
            messagebox.showerror("Parsing fallido", msg)
            return
        
        # mostrar preview en un diálogo
        ventana_preview = tk.Toplevel(self)
        ventana_preview.title("Previsualización de mapeo")
        ventana_preview.geometry("700x400")
        
        ttk.Label(
            ventana_preview,
            text=f"Se detectaron {len(preview)} Lot IDs. Revisa que se haya parseado correctamente:",
            font=("Segoe UI", 9),
        ).pack(anchor="w", padx=10, pady=(10, 4))
        
        # tabla con scroll
        frame_tabla = ttk.Frame(ventana_preview)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        cols = ("Lot ID", "Modelo", "Proceso")
        tree = ttk.Treeview(frame_tabla, columns=cols, show="headings", height=12)
        for col, width in zip(cols, (150, 250, 250)):
            tree.heading(col, text=col)
            tree.column(col, width=width)
        
        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for lot_id, modelo, proceso in preview[:50]:  # mostrar máximo 50 para no saturar
            tree.insert("", "end", values=(lot_id, modelo, proceso))
        
        if len(preview) > 50:
            ttk.Label(ventana_preview, text=f"... y {len(preview) - 50} más", foreground="#666").pack(anchor="w", padx=10)
        
        # botones
        btn_frame = ttk.Frame(ventana_preview)
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        def confirmar():
            # mapear todos
            mapeo, _ = self.db.actualizar_modelo_proceso_desde_texto(texto)
            actualizados, no_encontrados = self.db.actualizar_modelo_proceso_batch(mapeo)
            
            estado = f"Mapeados: {actualizados} PCBs."
            if no_encontrados:
                estado += f" {len(no_encontrados)} Lot ID no encontrados en el sistema."
            
            messagebox.showinfo("Mapeo completado", estado)
            self._recalcular_colores()
            self._redibujar_grid()
            self._dibujar_grid_separacion()
            self.text_pega.delete("1.0", tk.END)
            self._refrescar_historial()
            ventana_preview.destroy()
        
        ttk.Button(btn_frame, text="✓ Confirmar y mapear", command=confirmar).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="✗ Cancelar", command=ventana_preview.destroy).pack(side="left", padx=4)
        ttk.Label(btn_frame, text="Si se ve bien, confirma. Si no, cancela y revisa el formato.", foreground="#666").pack(side="left", padx=20)

    def _copiar_lot_ids(self):
        if self.charola_actual_id is None:
            messagebox.showwarning("Sin charola", "Selecciona o crea una charola primero.")
            return
        
        lot_id = self.entry_lot_id.get().strip()
        self.entry_lot_id.delete(0, tk.END)
        if not lot_id:
            return
        
        # Validacion 1: formato válido (solo alfanuméricos y guiones)
        if not all(c.isalnum() or c in '-_' for c in lot_id):
            messagebox.showerror("Formato inválido", "Lot ID contiene caracteres no válidos.")
            return

        # Validacion 2: duplicado en cualquier parte del sistema
        existente = self.db.lot_id_ya_escaneado(lot_id)
        if existente:
            charola_id, pos = existente
            messagebox.showerror(
                "Lot ID duplicado",
                f"Este Lot ID ya fue escaneado antes\n(charola {charola_id}, posición {pos}).",
            )
            return

        # Validacion 3: obtener siguiente posición libre
        siguiente = self._siguiente_posicion_libre()
        if siguiente is None:
            messagebox.showinfo("Charola completa", "Todas las posiciones ya fueron escaneadas o están bloqueadas.")
            return

        # Validacion 4: posición no está bloqueada
        posiciones_bloqueadas = self.db.posiciones_bloqueadas_de_charola(self.charola_actual_id)
        if siguiente in posiciones_bloqueadas:
            messagebox.showerror(
                "Posición bloqueada",
                f"La posición {siguiente} está bloqueada y no puede ser escaneada.",
            )
            return

        posicion = siguiente
        self.db.escanear(self.charola_actual_id, posicion, lot_id)
        self._dibujar_grid()
        self.entry_lot_id.focus()

    def _undo_celda(self):
        if self.charola_actual_id is None:
            return
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        if not datos:
            return
        # deshace la ultima posicion ocupada, respetando el orden fisico de escaneo
        secuencia = orden_escaneo(self.filas, self.columnas, self.volteada_actual)
        ocupadas = [p for p in secuencia if p in datos]
        if not ocupadas:
            return
        ultima = ocupadas[-1]
        confirm = messagebox.askyesno("Confirmar", f"Deshacer escaneo de la posicion {ultima}?")
        if confirm:
            self.db.undo_ultimo_escaneo(self.charola_actual_id, ultima)
            self._dibujar_grid()

    def _copiar_lot_ids(self):
        if self.charola_actual_id is None:
            return
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        lot_ids = [info["lot_id"] for info in datos.values()]
        if not lot_ids:
            messagebox.showinfo("Sin datos", "No hay Lot ID escaneados en esta charola.")
            return
        self.clipboard_clear()
        self.clipboard_append("\n".join(lot_ids))
        messagebox.showinfo("Copiado", f"{len(lot_ids)} Lot ID copiados al portapapeles.")

    # ---------------- Importar Excel / clasificacion ----------------
    def _importar_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if not path:
            return
        try:
            df = pd.read_excel(path)
        except Exception as exc:
            messagebox.showerror("Error al leer Excel", str(exc))
            return

        columnas_lower = {c.lower().strip(): c for c in df.columns}
        col_lot = _buscar_columna(columnas_lower, ["lot id", "lot_id", "lotid"])
        col_modelo = _buscar_columna(columnas_lower, ["modelo", "model", "product", "producto"])
        col_proceso = _buscar_columna(
            columnas_lower,
            ["proceso", "process", "process name", "estacion", "station", "wip status name"],
        )

        if not col_lot:
            messagebox.showerror(
                "Columnas no encontradas",
                "No se encontro una columna de Lot ID en el Excel. "
                "Se esperaba algo como 'Lot ID', 'LotID' o similar.",
            )
            return

        # Construir mapeo: lot_id -> (modelo, proceso)
        mapeo = {}
        for _, row in df.iterrows():
            lot_id = str(row[col_lot]).strip()
            modelo = str(row[col_modelo]).strip() if col_modelo else "N/D"
            proceso = str(row[col_proceso]).strip() if col_proceso else "N/D"
            mapeo[lot_id] = (modelo, proceso)

        # BATCH UPDATE contra TODOS los Lot ID de TODAS las charolas
        actualizados, no_encontrados = self.db.actualizar_modelo_proceso_batch(mapeo)

        self.db.registrar_historial(
            "excel_importado_batch",
            f"{os.path.basename(path)}: {actualizados} actualizados, {len(no_encontrados)} sin match",
        )

        estado = f"Importado: {actualizados} PCBs actualizadas en todas las charolas."
        if no_encontrados:
            porcentaje = (len(no_encontrados) / len(mapeo)) * 100
            if porcentaje > 10:
                messagebox.showwarning(
                    "Advertencia: muchos sin match",
                    f"{len(no_encontrados)} Lot ID ({porcentaje:.1f}%) del Excel no se encontraron "
                    "en ninguna charola escaneada. Verifica que hayas escaneado todas las piezas.",
                )
            estado += f" ({len(no_encontrados)} sin match)"

        self.lbl_import_status.config(text=estado)
        self._recalcular_colores()
        self._redibujar_grid()
        self._dibujar_grid_separacion()
        self._refrescar_historial()

        messagebox.showinfo("Importación batch completada", estado)

    def _exportar_consolidado(self):
        """Exporta TODAS las charolas a un Excel consolidado."""
        datos_todas = self.db.exportar_todas_charolas_a_dict()
        if not datos_todas:
            messagebox.showinfo("Sin datos", "No hay charolas con datos para exportar.")
            return

        # Usar openpyxl para crear Excel estructurado
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        wb.remove(wb.active)  # elimina hoja default

        for charola_id, charola_data in datos_todas.items():
            nombre = charola_data["nombre"]
            filas = charola_data["filas"]
            columnas = charola_data["columnas"]
            posiciones = charola_data["posiciones"]

            # crear hoja para esta charola
            ws = wb.create_sheet(title=nombre[:31])  # Excel limita nombre a 31 chars

            # encabezados
            ws["A1"] = "Posicion"
            ws["B1"] = "Lot ID"
            ws["C1"] = "Modelo"
            ws["D1"] = "Proceso"
            ws["E1"] = "Separada"

            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}1"].font = Font(bold=True, color="FFFFFF")
                ws[f"{col}1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            row = 2
            for pos in generar_posiciones(filas, columnas):
                info = posiciones.get(pos)
                if not info:
                    continue
                ws[f"A{row}"] = pos
                ws[f"B{row}"] = info["lot_id"]
                ws[f"C{row}"] = info.get("modelo", "")
                ws[f"D{row}"] = info.get("proceso", "")
                ws[f"E{row}"] = "✓" if info.get("separada") else ""
                row += 1

            # auto-ajustar ancho
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 12

        # guardar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path_salida = f"retrabajo_consolidado_{timestamp}.xlsx"
        try:
            wb.save(path_salida)
            messagebox.showinfo("Exportado", f"Archivo guardado como:\n{path_salida}")
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))

    def _recalcular_colores(self):
        """Asigna un color fijo de la paleta a cada modelo y a cada proceso nuevos."""
        cur = self.db.conn.cursor()
        cur.execute("SELECT DISTINCT modelo FROM pcbs WHERE modelo IS NOT NULL")
        for (modelo,) in cur.fetchall():
            if modelo not in self.color_map_modelo:
                self.color_map_modelo[modelo] = COLOR_PALETTE[len(self.color_map_modelo) % len(COLOR_PALETTE)]
        cur.execute("SELECT DISTINCT proceso FROM pcbs WHERE proceso IS NOT NULL")
        for (proceso,) in cur.fetchall():
            if proceso not in self.color_map_proceso:
                self.color_map_proceso[proceso] = COLOR_PALETTE[len(self.color_map_proceso) % len(COLOR_PALETTE)]

    def _redibujar_grid(self):
        """Recolorea el grid de la charola actual segun la vista activa (modelo/proceso)."""
        if self.charola_actual_id is None or not self.celdas:
            return
        datos = self.db.pcbs_de_charola(self.charola_actual_id)
        criterio = self.vista_color.get()
        mapa_color = self.color_map_modelo if criterio == "modelo" else self.color_map_proceso

        for pos, lbl in self.celdas.items():
            info = datos.get(pos)
            if not info:
                lbl.config(bg=COLOR_PENDIENTE, text=pos)
                continue
            prefijo = "\u2713 " if info.get("separada") else ""
            texto = prefijo + info["lot_id"]
            valor = info.get(criterio)
            if valor and valor in mapa_color:
                lbl.config(bg=mapa_color[valor], text=texto)
            elif info.get("modelo") is None and info.get("proceso") is None:
                lbl.config(bg=COLOR_ESCANEADO, text=texto)
            else:
                lbl.config(bg=COLOR_SIN_MATCH, text=texto)

        self._dibujar_leyenda(mapa_color)

    def _dibujar_leyenda(self, mapa_color):
        for w in self.frame_leyenda.winfo_children():
            w.destroy()
        if not mapa_color:
            ttk.Label(self.frame_leyenda, text="(sin datos importados)").pack(padx=8, pady=8)
            return
        for nombre, color in mapa_color.items():
            fila = ttk.Frame(self.frame_leyenda)
            fila.pack(fill="x", padx=6, pady=2)
            tk.Label(fila, bg=color, width=2, height=1).pack(side="left")
            ttk.Label(fila, text=nombre).pack(side="left", padx=6)

    # ---------------- Historial ----------------
    def _refrescar_historial(self):
        for item in self.tree_historial.get_children():
            self.tree_historial.delete(item)
        for evento, detalle, fecha in self.db.leer_historial():
            self.tree_historial.insert("", "end", values=(evento, detalle, fecha))


def _col_letra_a_indice(letras):
    idx = 0
    for ch in letras:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _buscar_columna(columnas_lower, candidatos):
    for cand in candidatos:
        if cand in columnas_lower:
            return columnas_lower[cand]
    return None


if __name__ == "__main__":
    app = App()
    app.mainloop()
